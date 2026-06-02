from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, WebSocket, WebSocketDisconnect, Query, UploadFile, File
from fastapi.responses import JSONResponse, FileResponse
import re
import uuid
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any, Set
from datetime import datetime, timezone, timedelta
from bson import ObjectId
import os
import logging
import bcrypt
import jwt
import secrets
import json
import asyncio
import hmac
import hashlib
import razorpay
import io
import openpyxl
import httpx
from emergentintegrations.llm.chat import LlmChat, UserMessage
from emergentintegrations.payments.stripe.checkout import StripeCheckout, CheckoutSessionResponse, CheckoutStatusResponse, CheckoutSessionRequest

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALGORITHM = "HS256"

app = FastAPI()
api_router = APIRouter(prefix="/api")

# Health endpoint — must be fast, no DB touch, used by K8s liveness/readiness probes
@api_router.get("/health")
async def health_check():
    return {"status": "ok"}

@app.get("/")
async def root():
    return {"status": "ok", "service": "cravitoo-api"}

# WebSocket Connection Manager
class ConnectionManager:
    def __init__(self):
        # user_id -> set of websockets (multiple devices per user)
        self.user_connections: Dict[str, Set[WebSocket]] = {}
        # vendor_id -> set of websockets
        self.vendor_connections: Dict[str, Set[WebSocket]] = {}

    async def connect_user(self, user_id: str, websocket: WebSocket):
        await websocket.accept()
        if user_id not in self.user_connections:
            self.user_connections[user_id] = set()
        self.user_connections[user_id].add(websocket)

    async def connect_vendor(self, vendor_id: str, websocket: WebSocket):
        await websocket.accept()
        if vendor_id not in self.vendor_connections:
            self.vendor_connections[vendor_id] = set()
        self.vendor_connections[vendor_id].add(websocket)

    def disconnect_user(self, user_id: str, websocket: WebSocket):
        if user_id in self.user_connections:
            self.user_connections[user_id].discard(websocket)
            if not self.user_connections[user_id]:
                del self.user_connections[user_id]

    def disconnect_vendor(self, vendor_id: str, websocket: WebSocket):
        if vendor_id in self.vendor_connections:
            self.vendor_connections[vendor_id].discard(websocket)
            if not self.vendor_connections[vendor_id]:
                del self.vendor_connections[vendor_id]

    async def send_to_user(self, user_id: str, message: dict):
        if user_id not in self.user_connections:
            return
        dead = set()
        for ws in list(self.user_connections[user_id]):
            try:
                await ws.send_json(message)
            except Exception:
                dead.add(ws)
        for ws in dead:
            self.user_connections[user_id].discard(ws)

    async def send_to_vendor(self, vendor_id: str, message: dict):
        if vendor_id not in self.vendor_connections:
            return
        dead = set()
        for ws in list(self.vendor_connections[vendor_id]):
            try:
                await ws.send_json(message)
            except Exception:
                dead.add(ws)
        for ws in dead:
            self.vendor_connections[vendor_id].discard(ws)

manager = ConnectionManager()

def verify_ws_token(token: str) -> Optional[dict]:
    """Validate JWT token from WebSocket query string. Returns payload dict or None."""
    if not token:
        return None
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            return None
        return payload
    except (jwt.ExpiredSignatureError, jwt.InvalidTokenError):
        return None

# Helper Functions
def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode("utf-8"), salt)
    return hashed.decode("utf-8")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode("utf-8"), hashed_password.encode("utf-8"))

def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {"sub": user_id, "email": email, "role": role, "exp": datetime.now(timezone.utc) + timedelta(minutes=15), "type": "access"}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": safe_objectid(payload["sub"], "User")})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user["_id"] = str(user["_id"])
        user["id"] = user["_id"]
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# Models — extracted to /app/backend/models.py during iteration 12 refactor
from models import (  # noqa: E402
    RegisterRequest, LoginRequest, UserResponse,
    CompanyCreate, CompanyResponse,
    VendorCreate, VendorResponse,
    MenuItemCreate, MenuItemResponse, MenuItemSiteUpdate,
    OrderItemInput, OrderCreate, OrderResponse, OrderStatus, CheckoutRequest,
    AIRecommendationRequest,
    SiteCreate, VendorSiteMappingCreate, MealScheduleEntry, MealScheduleUpdate,
    CityCreate, CityAdminCreate,
    VendorOnboardingBasic, VendorOnboardingUpdate, ChecklistUpdate, OnboardingDecision,
    CHECKLIST_FIELDS, DOC_TYPES, ONBOARDING_STATUSES,
    SiteAdminCreate, SuperAdminCreate, MasterAdminCreate,
    ReviewCreate, PreferencesUpdate, SubscriptionCreate,
    EmployeeCreate, BulkOrderItem, BulkOrderCreate, EventCateringCreate,
    NotificationCreate, LoyaltyRedeemRequest,
    RazorpayOrderCreate, RazorpayVerify,
    PushTokenRegister,
)

from enum import Enum  # kept for any local enums elsewhere

# Helper - safe ObjectId parsing
def safe_objectid(id_str: str, entity_name: str = "Resource") -> ObjectId:
    try:
        return ObjectId(id_str)
    except Exception:
        raise HTTPException(status_code=404, detail=f"{entity_name} not found")

# Helper - detect if request is HTTPS
def is_secure_request(request: Request) -> bool:
    forwarded_proto = request.headers.get("x-forwarded-proto", "")
    return forwarded_proto == "https" or request.url.scheme == "https"

# Helper - generate QR code data for order pickup
def generate_pickup_qr(order_id: str) -> str:
    import hashlib
    qr_hash = hashlib.sha256(f"{order_id}{JWT_SECRET}".encode()).hexdigest()[:16]
    return f"CRAVITOO-PICKUP-{order_id}-{qr_hash}"

def verify_pickup_qr(qr_code: str, order_id: str) -> bool:
    import hashlib
    qr_hash = hashlib.sha256(f"{order_id}{JWT_SECRET}".encode()).hexdigest()[:16]
    expected = f"CRAVITOO-PICKUP-{order_id}-{qr_hash}"
    return qr_code == expected

# Brute force protection
MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_MINUTES = 15

async def check_brute_force(identifier: str) -> bool:
    """Returns True if locked out, False if OK to proceed"""
    record = await db.login_attempts.find_one({"identifier": identifier})
    if not record:
        return False
    if record.get("attempts", 0) >= MAX_LOGIN_ATTEMPTS:
        locked_until = record.get("locked_until")
        if locked_until:
            # MongoDB returns naive datetime - convert to aware
            if locked_until.tzinfo is None:
                locked_until = locked_until.replace(tzinfo=timezone.utc)
            if datetime.now(timezone.utc) < locked_until:
                return True
            await db.login_attempts.delete_one({"identifier": identifier})
    return False

async def record_failed_login(identifier: str):
    record = await db.login_attempts.find_one({"identifier": identifier})
    if record:
        new_attempts = record.get("attempts", 0) + 1
        update = {"attempts": new_attempts, "last_attempt": datetime.now(timezone.utc)}
        if new_attempts >= MAX_LOGIN_ATTEMPTS:
            update["locked_until"] = datetime.now(timezone.utc) + timedelta(minutes=LOCKOUT_MINUTES)
        await db.login_attempts.update_one({"identifier": identifier}, {"$set": update})
    else:
        await db.login_attempts.insert_one({
            "identifier": identifier,
            "attempts": 1,
            "last_attempt": datetime.now(timezone.utc)
        })

async def clear_login_attempts(identifier: str):
    await db.login_attempts.delete_one({"identifier": identifier})

# Startup Events
@app.on_event("startup")
async def startup_event():
    """Resilient, NON-BLOCKING startup — never block FastAPI from serving requests.
    Seed/index work runs in background so the health probe responds fast."""
    import asyncio

    async def _index_and_seed():
        index_ops = [
            ("users", "email", {"unique": True}),
            ("password_reset_tokens", "expires_at", {"expireAfterSeconds": 0}),
            ("login_attempts", "identifier", {}),
            ("companies", "name", {}),
            ("vendors", "name", {}),
            ("menu_items", "vendor_id", {}),
            ("menu_items", "site_id", {}),
            ("orders", "user_id", {}),
            ("orders", "vendor_id", {}),
            ("orders", "site_id", {}),
            ("notifications", "user_id", {}),
            ("notifications", "created_at", {}),
            ("sites", "name", {}),
        ]
        for coll, field, opts in index_ops:
            try:
                await db[coll].create_index(field, **opts)
            except Exception as e:
                logger.warning(f"Index create skipped on {coll}.{field}: {e}")

        try:
            await db.vendor_site_mappings.create_index([("vendor_id", 1), ("site_id", 1)], unique=True)
        except Exception as e:
            logger.warning(f"vendor_site_mappings index skipped: {e}")
        try:
            await db.meal_schedules.create_index("site_id", unique=True)
        except Exception as e:
            logger.warning(f"meal_schedules index skipped: {e}")

        try:
            await seed_admin()
        except Exception as e:
            logger.error(f"seed_admin failed: {e}")
        try:
            await seed_demo_data()
        except Exception as e:
            logger.error(f"seed_demo_data failed: {e}")
        logger.info("Background startup tasks complete")

    # Fire-and-forget — does NOT block startup probe
    asyncio.create_task(_index_and_seed())

async def seed_admin():
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@cravitoo.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        hashed = hash_password(admin_password)
        await db.users.insert_one({
            "email": admin_email,
            "password_hash": hashed,
            "name": "Master Admin",
            "role": "master_admin",
            "created_at": datetime.now(timezone.utc)
        })
        logger.info(f"Master admin created: {admin_email}")
    else:
        updates = {}
        if existing.get("role") != "master_admin":
            updates["role"] = "master_admin"
            updates["name"] = "Master Admin"
        if not verify_password(admin_password, existing["password_hash"]):
            updates["password_hash"] = hash_password(admin_password)
        if updates:
            await db.users.update_one({"email": admin_email}, {"$set": updates})
            logger.info("Master admin updated")

async def seed_demo_data():
    demo_company_email = "demo@techcorp.com"
    demo_vendor_email = "vendor@spicekitchen.com"
    demo_employee_email = "employee@techcorp.com"
    
    if not await db.companies.find_one({"name": "Tech Corp"}):
        company_result = await db.companies.insert_one({
            "name": "Tech Corp",
            "address": "123 Tech Park, Bangalore",
            "contact_email": "contact@techcorp.com",
            "contact_phone": "+91-9876543210",
            "status": "active",
            "created_at": datetime.now(timezone.utc)
        })
        company_id = str(company_result.inserted_id)
        logger.info(f"Demo company created: Tech Corp")
    else:
        company_id = str((await db.companies.find_one({"name": "Tech Corp"}))['_id'])
    
    if not await db.users.find_one({"email": demo_company_email}):
        await db.users.insert_one({
            "email": demo_company_email,
            "password_hash": hash_password("demo123"),
            "name": "Corporate Admin",
            "role": "corporate_admin",
            "company_id": company_id,
            "created_at": datetime.now(timezone.utc)
        })
        logger.info(f"Demo corporate admin created")
    
    if not await db.vendors.find_one({"name": "Spice Kitchen"}):
        vendor_result = await db.vendors.insert_one({
            "name": "Spice Kitchen",
            "description": "Authentic North Indian Cuisine",
            "cuisine_type": "North Indian",
            "contact_email": "contact@spicekitchen.com",
            "contact_phone": "+91-9876543211",
            "rating": 4.5,
            "status": "active",
            "created_at": datetime.now(timezone.utc)
        })
        vendor_id = str(vendor_result.inserted_id)
        logger.info(f"Demo vendor created: Spice Kitchen")
    else:
        vendor_id = str((await db.vendors.find_one({"name": "Spice Kitchen"}))['_id'])
    
    if not await db.users.find_one({"email": demo_vendor_email}):
        await db.users.insert_one({
            "email": demo_vendor_email,
            "password_hash": hash_password("vendor123"),
            "name": "Vendor Manager",
            "role": "vendor",
            "vendor_id": vendor_id,
            "created_at": datetime.now(timezone.utc)
        })
        logger.info(f"Demo vendor user created")
    
    if not await db.users.find_one({"email": demo_employee_email}):
        await db.users.insert_one({
            "email": demo_employee_email,
            "password_hash": hash_password("employee123"),
            "name": "John Doe",
            "role": "employee",
            "company_id": company_id,
            "created_at": datetime.now(timezone.utc)
        })
        logger.info(f"Demo employee created")
    
    if await db.menu_items.count_documents({"vendor_id": vendor_id}) == 0:
        menu_items = [
            {"vendor_id": vendor_id, "name": "Paneer Tikka", "description": "Grilled cottage cheese with spices", "category": "Appetizer", "price": 180.0, "is_vegetarian": True, "is_available": True, "image_url": "https://images.unsplash.com/photo-1567188040759-fb8a883dc6d8", "created_at": datetime.now(timezone.utc)},
            {"vendor_id": vendor_id, "name": "Butter Chicken", "description": "Creamy tomato curry with tender chicken", "category": "Main Course", "price": 280.0, "is_vegetarian": False, "is_available": True, "image_url": "https://images.unsplash.com/photo-1603894584373-5ac82b2ae398", "created_at": datetime.now(timezone.utc)},
            {"vendor_id": vendor_id, "name": "Dal Makhani", "description": "Black lentils in rich creamy gravy", "category": "Main Course", "price": 220.0, "is_vegetarian": True, "is_available": True, "image_url": "https://images.unsplash.com/photo-1546833999-b9f581a1996d", "created_at": datetime.now(timezone.utc)},
            {"vendor_id": vendor_id, "name": "Garlic Naan", "description": "Soft bread with garlic butter", "category": "Bread", "price": 60.0, "is_vegetarian": True, "is_available": True, "image_url": "https://images.unsplash.com/photo-1628840042765-356cda07504e", "created_at": datetime.now(timezone.utc)},
            {"vendor_id": vendor_id, "name": "Gulab Jamun", "description": "Sweet milk dumplings in sugar syrup", "category": "Dessert", "price": 80.0, "is_vegetarian": True, "is_available": True, "image_url": "https://images.unsplash.com/photo-1589119908995-c6b5f3e3bf6a", "created_at": datetime.now(timezone.utc)}
        ]
        await db.menu_items.insert_many(menu_items)
        logger.info(f"Demo menu items created")
    
    # Seed a default site & vendor-site mapping
    site_id = None
    existing_site = await db.sites.find_one({"name": "Tech Corp - Bangalore HQ"})
    if not existing_site:
        site_result = await db.sites.insert_one({
            "name": "Tech Corp - Bangalore HQ",
            "company_id": company_id,
            "address": "123 Tech Park, Whitefield, Bangalore",
            "city": "Bangalore",
            "contact_email": "site-hq@techcorp.com",
            "contact_phone": "+91-9876543220",
            "allow_pre_order": True,
            "allow_cash_carry": True,
            "allow_company_paid": True,
            "allow_employee_paid": True,
            "status": "active",
            "created_at": datetime.now(timezone.utc)
        })
        site_id = str(site_result.inserted_id)
        # Default meal schedule
        await db.meal_schedules.insert_one({
            "site_id": site_id,
            "schedules": [
                {"meal_period": "breakfast", "start_time": "07:30", "end_time": "10:30", "enabled": True},
                {"meal_period": "lunch", "start_time": "12:00", "end_time": "15:00", "enabled": True},
                {"meal_period": "snacks", "start_time": "16:00", "end_time": "18:00", "enabled": True},
                {"meal_period": "dinner", "start_time": "19:00", "end_time": "22:00", "enabled": False},
            ],
            "updated_at": datetime.now(timezone.utc)
        })
        logger.info(f"Demo site created with meal schedules")
    else:
        site_id = str(existing_site["_id"])
    
    # Vendor-site mapping (Spice Kitchen at Bangalore HQ)
    if not await db.vendor_site_mappings.find_one({"vendor_id": vendor_id, "site_id": site_id}):
        await db.vendor_site_mappings.insert_one({
            "vendor_id": vendor_id,
            "site_id": site_id,
            "status": "active",
            "created_at": datetime.now(timezone.utc)
        })
    
    # Backfill site_id, meal_periods, show_price on existing menu items
    await db.menu_items.update_many(
        {"vendor_id": vendor_id, "$or": [{"site_id": {"$exists": False}}, {"site_id": None}]},
        {"$set": {"site_id": site_id, "meal_periods": ["breakfast", "lunch", "snacks"], "show_price": True}}
    )
    
    # Backfill employee with site_id
    await db.users.update_one(
        {"email": demo_employee_email, "site_id": {"$exists": False}},
        {"$set": {"site_id": site_id}}
    )
    
    # Demo Site Admin
    site_admin_email = "siteadmin@techcorp.com"
    if not await db.users.find_one({"email": site_admin_email}):
        await db.users.insert_one({
            "email": site_admin_email,
            "password_hash": hash_password("site123"),
            "name": "Site Admin",
            "role": "site_admin",
            "site_id": site_id,
            "company_id": company_id,
            "created_at": datetime.now(timezone.utc)
        })
        logger.info("Demo site admin created")
    
    test_creds_content = f"""# Cravitoo Test Credentials

## Master Admin
- Email: {os.environ.get('ADMIN_EMAIL', 'admin@cravitoo.com')}
- Password: {os.environ.get('ADMIN_PASSWORD', 'admin123')}
- Role: master_admin (full platform control, Partner App access)

## Corporate Admin
- Email: demo@techcorp.com
- Password: demo123
- Role: corporate_admin (web app)
- Company: Tech Corp

## Site Admin
- Email: siteadmin@techcorp.com
- Password: site123
- Role: site_admin (Partner App access)
- Site: Tech Corp - Bangalore HQ

## Vendor Manager
- Email: vendor@spicekitchen.com
- Password: vendor123
- Role: vendor (Partner App access)
- Vendor: Spice Kitchen

## Employee
- Email: employee@techcorp.com
- Password: employee123
- Role: employee (Customer App access)
- Company: Tech Corp
- Site: Tech Corp - Bangalore HQ

## Auth Endpoints
- POST /api/auth/register
- POST /api/auth/login
- GET /api/auth/me
- POST /api/auth/logout
"""
    try:
        Path("/app/memory").mkdir(exist_ok=True)
        Path("/app/memory/test_credentials.md").write_text(test_creds_content)
    except (PermissionError, OSError) as e:
        # Read-only filesystem in production — skip silently
        logger.debug(f"Skipped writing test_credentials.md: {e}")

# Auth Routes
@api_router.post("/auth/register")
async def register(data: RegisterRequest, request: Request, response: Response):
    email_lower = data.email.lower()
    existing = await db.users.find_one({"email": email_lower})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_doc = {
        "email": email_lower,
        "password_hash": hash_password(data.password),
        "name": data.name,
        "role": data.role,
        "created_at": datetime.now(timezone.utc)
    }
    
    if data.company_id:
        user_doc["company_id"] = data.company_id
    
    result = await db.users.insert_one(user_doc)
    user_id = str(result.inserted_id)
    
    access_token = create_access_token(user_id, email_lower, data.role)
    refresh_token = create_refresh_token(user_id)
    
    secure_cookie = is_secure_request(request)
    samesite_value = "none" if secure_cookie else "lax"
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=secure_cookie, samesite=samesite_value, max_age=900, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=secure_cookie, samesite=samesite_value, max_age=604800, path="/")
    
    return {"id": user_id, "email": email_lower, "name": data.name, "role": data.role, "access_token": access_token, "refresh_token": refresh_token}

@api_router.post("/auth/login")
async def login(data: LoginRequest, request: Request, response: Response):
    email_lower = data.email.lower()
    # Use X-Forwarded-For header if available (for behind proxy/LB), else fallback to client.host
    forwarded_for = request.headers.get("x-forwarded-for", "")
    client_ip = forwarded_for.split(",")[0].strip() if forwarded_for else (request.client.host if request.client else "unknown")
    # Also track by email-only to catch attacks from different IPs
    identifier = f"{client_ip}:{email_lower}"
    email_identifier = f"email:{email_lower}"
    
    if await check_brute_force(identifier) or await check_brute_force(email_identifier):
        raise HTTPException(status_code=429, detail=f"Too many failed attempts. Account locked for {LOCKOUT_MINUTES} minutes.")
    
    user = await db.users.find_one({"email": email_lower})
    
    if not user or not verify_password(data.password, user["password_hash"]):
        await record_failed_login(identifier)
        await record_failed_login(email_identifier)
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    await clear_login_attempts(identifier)
    await clear_login_attempts(email_identifier)
    
    user_id = str(user["_id"])
    access_token = create_access_token(user_id, email_lower, user["role"])
    refresh_token = create_refresh_token(user_id)
    
    secure_cookie = is_secure_request(request)
    samesite_value = "none" if secure_cookie else "lax"
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=secure_cookie, samesite=samesite_value, max_age=900, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=secure_cookie, samesite=samesite_value, max_age=604800, path="/")
    
    return {
        "id": user_id,
        "email": email_lower,
        "name": user["name"],
        "role": user["role"],
        "company_id": user.get("company_id"),
        "vendor_id": user.get("vendor_id"),
        "site_id": user.get("site_id"),
        "assigned_sites": user.get("assigned_sites", []),
        "access_token": access_token,
        "refresh_token": refresh_token
    }

@api_router.get("/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    return user

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logged out successfully"}


# ============== Email OTP login (channel-agnostic — SMS can be added later) ==============

class OTPRequest(BaseModel):
    email: EmailStr
    channel: Optional[str] = "email"  # 'email' | 'sms' (future) | 'whatsapp' (future)
    purpose: Optional[str] = "Login"  # 'Login' | 'Password Reset' | 'Account Verification'


class OTPVerify(BaseModel):
    email: EmailStr
    code: str


OTP_EXPIRY_MINUTES = 10
OTP_MAX_ATTEMPTS = 5
OTP_REQUEST_LIMIT_PER_HOUR = 3


@api_router.post("/auth/otp/request")
async def request_otp(data: OTPRequest, request: Request):
    """Generate a 6-digit OTP and send it via the chosen channel.
    Rate-limited: 3 per email per hour to prevent abuse.
    Does NOT reveal whether the email exists (anti-enumeration)."""
    import email_service  # local import to keep top-level cleaner

    email_lower = data.email.lower()
    channel = (data.channel or "email").lower()
    purpose = data.purpose or "Login"

    if channel not in ("email", "sms", "whatsapp"):
        raise HTTPException(status_code=400, detail="Invalid channel")
    if channel in ("sms", "whatsapp"):
        raise HTTPException(status_code=501, detail=f"{channel.upper()} OTP is not yet configured. Please use email.")

    # Rate-limit: count requests in the last 1 hour for this email
    one_hour_ago = datetime.now(timezone.utc) - timedelta(hours=1)
    recent_count = await db.otp_codes.count_documents({
        "identifier": email_lower,
        "channel": channel,
        "created_at": {"$gte": one_hour_ago},
    })
    if recent_count >= OTP_REQUEST_LIMIT_PER_HOUR:
        raise HTTPException(
            status_code=429,
            detail=f"Too many OTP requests for this email. Please wait an hour and try again.",
        )

    # Generate, hash, store
    code = email_service.generate_otp(6)
    code_hash = email_service.hash_otp(code)
    now = datetime.now(timezone.utc)
    expires_at = now + timedelta(minutes=OTP_EXPIRY_MINUTES)

    # Invalidate any existing active codes for the same email+channel+purpose
    await db.otp_codes.update_many(
        {"identifier": email_lower, "channel": channel, "purpose": purpose, "used": False, "expires_at": {"$gte": now}},
        {"$set": {"superseded": True}},
    )

    await db.otp_codes.insert_one({
        "identifier": email_lower,
        "channel": channel,
        "purpose": purpose,
        "code_hash": code_hash,
        "attempts": 0,
        "used": False,
        "superseded": False,
        "created_at": now,
        "expires_at": expires_at,
    })

    # Send via the requested channel (best-effort — failures don't leak to user)
    success, err = email_service.send_otp_channel(
        identifier=email_lower,
        code=code,
        channel=channel,
        purpose=purpose,
        expiry_minutes=OTP_EXPIRY_MINUTES,
    )

    if not success:
        logger.warning(f"OTP delivery failed for {email_lower} via {channel}: {err}")
        # We DO surface delivery failures so the user knows to retry,
        # but we don't leak whether the email is registered.
        raise HTTPException(
            status_code=502,
            detail="We couldn't send the code right now. Please try again in a moment.",
        )

    # Anti-enumeration: always return the same response whether or not the email exists
    return {
        "ok": True,
        "channel": channel,
        "expires_in_minutes": OTP_EXPIRY_MINUTES,
        "message": f"If an account exists for {email_lower}, a verification code has been sent.",
    }


@api_router.post("/auth/otp/verify")
async def verify_otp_login(data: OTPVerify, request: Request, response: Response):
    """Verify an OTP and issue Cravitoo JWT tokens. Also acts as an auto-register
    fallback ONLY for the 'employee' role — admin/vendor accounts must be created via the normal flow."""
    import email_service

    email_lower = data.email.lower()
    code = (data.code or "").strip()
    if not code or not code.isdigit() or len(code) < 4 or len(code) > 8:
        raise HTTPException(status_code=400, detail="Invalid code format")

    now = datetime.now(timezone.utc)

    # Find the most recent active OTP
    record = await db.otp_codes.find_one(
        {
            "identifier": email_lower,
            "used": False,
            "superseded": False,
            "expires_at": {"$gte": now},
        },
        sort=[("created_at", -1)],
    )
    if not record:
        raise HTTPException(status_code=400, detail="Code is invalid or has expired. Please request a new one.")

    # Increment attempts BEFORE verification (to prevent timing attacks)
    record_id = record["_id"]
    if record.get("attempts", 0) >= OTP_MAX_ATTEMPTS:
        await db.otp_codes.update_one({"_id": record_id}, {"$set": {"used": True}})
        raise HTTPException(status_code=429, detail="Too many incorrect attempts. Please request a new code.")

    if not email_service.verify_otp(code, record["code_hash"]):
        await db.otp_codes.update_one({"_id": record_id}, {"$inc": {"attempts": 1}})
        raise HTTPException(status_code=400, detail="Incorrect code. Please try again.")

    # Code is valid — mark used
    await db.otp_codes.update_one({"_id": record_id}, {"$set": {"used": True, "used_at": now}})

    # Find or auto-create the user
    user = await db.users.find_one({"email": email_lower})
    auto_created = False
    if not user:
        # Auto-register as employee (vendors and admins must be created by an admin)
        user_doc = {
            "email": email_lower,
            "name": email_lower.split("@")[0].replace(".", " ").title(),
            "role": "employee",
            "password_hash": hash_password(secrets.token_urlsafe(24)),  # random unguessable
            "phone": None,
            "company_id": None,
            "vendor_id": None,
            "email_verified": True,
            "created_at": now,
            "created_via": "otp",
        }
        result = await db.users.insert_one(user_doc)
        user = {**user_doc, "_id": result.inserted_id}
        auto_created = True

    user_id = str(user["_id"])
    access_token = create_access_token(user_id, email_lower, user["role"])
    refresh_token = create_refresh_token(user_id)

    secure_cookie = is_secure_request(request)
    samesite_value = "none" if secure_cookie else "lax"
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=secure_cookie, samesite=samesite_value, max_age=900, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=secure_cookie, samesite=samesite_value, max_age=604800, path="/")

    # Mark email as verified on the user record too
    await db.users.update_one({"_id": user["_id"]}, {"$set": {"email_verified": True, "last_login_at": now}})

    return {
        "id": user_id,
        "email": email_lower,
        "name": user["name"],
        "role": user["role"],
        "company_id": user.get("company_id"),
        "vendor_id": user.get("vendor_id"),
        "site_id": user.get("site_id"),
        "assigned_sites": user.get("assigned_sites", []),
        "access_token": access_token,
        "refresh_token": refresh_token,
        "auto_created": auto_created,
    }


# ============== DPDP / GDPR — Right to Access & Right to Erasure ==============

def _stringify_datetimes(doc: Any) -> Any:
    """Recursively convert datetime/ObjectId/bson types to JSON-safe primitives."""
    if isinstance(doc, dict):
        return {k: _stringify_datetimes(v) for k, v in doc.items()}
    if isinstance(doc, list):
        return [_stringify_datetimes(x) for x in doc]
    if isinstance(doc, datetime):
        return doc.isoformat()
    if isinstance(doc, ObjectId):
        return str(doc)
    return doc


@api_router.get("/me/data")
async def export_my_data(user: dict = Depends(get_current_user)):
    """DPDP Act / GDPR right-to-access. Returns a JSON snapshot of all personal data
    Cravitoo holds about the calling user. Sensitive fields (password_hash, tokens) are excluded."""
    uid = user["id"]
    uid_obj = safe_objectid(uid, "User")

    # User profile (drop sensitive fields)
    profile = await db.users.find_one(
        {"_id": uid_obj},
        {"password_hash": 0},
    ) or {}
    if profile:
        profile["id"] = str(profile.pop("_id", uid))

    # Orders + reviews + favorites + loyalty + subscriptions + notifications
    orders = await db.orders.find({"user_id": uid}).to_list(2000)
    for o in orders:
        o["id"] = str(o.pop("_id"))

    reviews = await db.reviews.find({"user_id": uid}).to_list(2000)
    for r in reviews:
        r["id"] = str(r.pop("_id"))

    favorites = await db.favorites.find({"user_id": uid}).to_list(2000)
    for f in favorites:
        f["id"] = str(f.pop("_id"))

    loyalty = await db.loyalty.find_one({"user_id": uid}) or {}
    if loyalty:
        loyalty["id"] = str(loyalty.pop("_id", ""))

    subscriptions = await db.subscriptions.find({"user_id": uid}).to_list(500)
    for s in subscriptions:
        s["id"] = str(s.pop("_id"))

    notifications = await db.notifications.find({"user_id": uid}).to_list(2000)
    for n in notifications:
        n["id"] = str(n.pop("_id"))

    preferences = await db.preferences.find_one({"user_id": uid}) or {}
    if preferences:
        preferences["id"] = str(preferences.pop("_id", ""))

    push_tokens = await db.push_tokens.find({"user_id": uid}, {"token": 0}).to_list(50)
    for pt in push_tokens:
        pt["id"] = str(pt.pop("_id"))
        pt["token"] = "[REDACTED]"  # don't expose actual tokens

    return _stringify_datetimes({
        "export_format_version": "1.0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "data_controller": "Cravitoo Foods Private Limited",
        "user": user["email"],
        "profile": profile,
        "orders": orders,
        "reviews": reviews,
        "favorites": favorites,
        "loyalty": loyalty,
        "subscriptions": subscriptions,
        "notifications": notifications,
        "preferences": preferences,
        "push_tokens": push_tokens,
        "_note": "Vendor KYC documents, payment processor metadata, and 7-year retention financial records are managed under separate compliance regimes (GST/Companies Act/RBI). For those, please email privacy@cravitoo.com.",
    })


@api_router.delete("/me/data")
async def delete_my_data(
    confirm: str = Query(..., description="Must equal 'DELETE' to confirm"),
    user: dict = Depends(get_current_user),
):
    """DPDP Act / GDPR right-to-erasure. Deletes personal data and anonymises tax-mandated records.

    Master_admin accounts cannot self-delete via this endpoint (would lock the platform) —
    they must contact another master_admin or escalate via privacy@cravitoo.com.
    """
    if confirm != "DELETE":
        raise HTTPException(status_code=400, detail="Pass ?confirm=DELETE to confirm")

    if user["role"] == "master_admin":
        raise HTTPException(
            status_code=403,
            detail="Master Admin accounts cannot be self-deleted via this endpoint. Please contact privacy@cravitoo.com to escalate.",
        )

    uid = user["id"]
    uid_obj = safe_objectid(uid, "User")
    email_lower = user["email"].lower()

    # 1) Anonymise orders (keep for tax/audit but strip PII)
    anon_marker = f"deleted_user_{secrets.token_hex(6)}"
    await db.orders.update_many(
        {"user_id": uid},
        {
            "$set": {
                "user_id": anon_marker,
                "user_email_anon": True,
                "deleted_at": datetime.now(timezone.utc),
            },
            "$unset": {"special_instructions": ""},
        },
    )

    # 2) Anonymise reviews (keep ratings for vendor reputation, drop user link)
    await db.reviews.update_many(
        {"user_id": uid},
        {"$set": {"user_id": anon_marker, "anonymised": True, "comment": "[Comment removed by user]"}},
    )

    # 3) Hard-delete personal records
    deletions = [
        db.favorites.delete_many({"user_id": uid}),
        db.preferences.delete_many({"user_id": uid}),
        db.subscriptions.delete_many({"user_id": uid}),
        db.notifications.delete_many({"user_id": uid}),
        db.push_tokens.delete_many({"user_id": uid}),
        db.loyalty.delete_many({"user_id": uid}),
        db.login_attempts.delete_many({"identifier": {"$regex": email_lower}}),
        db.audit_log.delete_many({"user_id": uid}),
    ]
    await asyncio.gather(*deletions)

    # 4) Delete the user account itself
    await db.users.delete_one({"_id": uid_obj})

    # 5) Record this deletion in a compliance log (no PII — just the fact it happened)
    await db.deletion_log.insert_one({
        "anon_id": anon_marker,
        "role": user["role"],
        "deleted_at": datetime.now(timezone.utc),
        "compliance_basis": "DPDP_2023_section_12",
    })

    return {
        "ok": True,
        "message": "Your account and personal data have been deleted. Order records have been anonymised for tax compliance (retained 7 years).",
        "anonymisation_id": anon_marker,
    }


# Company Routes
@api_router.post("/companies")
async def create_company(data: CompanyCreate, user: dict = Depends(get_current_user)):
    if user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Only super admin can create companies")
    
    company_doc = {
        **data.model_dump(),
        "status": "active",
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.companies.insert_one(company_doc)
    return {"id": str(result.inserted_id), **data.model_dump(), "status": "active"}

@api_router.get("/companies")
async def get_companies(user: dict = Depends(get_current_user)):
    if user["role"] not in ["super_admin", "corporate_admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    companies = await db.companies.find({}, {"_id": 1, "name": 1, "address": 1, "contact_email": 1, "contact_phone": 1, "status": 1, "created_at": 1}).to_list(1000)
    for company in companies:
        company["id"] = str(company.pop("_id"))
    return companies

# Vendor Routes
@api_router.post("/vendors")
async def create_vendor(data: VendorCreate, user: dict = Depends(get_current_user)):
    if user["role"] not in ["super_admin", "corporate_admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    vendor_doc = {
        **data.model_dump(),
        "rating": 0.0,
        "status": "active",
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.vendors.insert_one(vendor_doc)
    return {"id": str(result.inserted_id), **data.model_dump(), "rating": 0.0, "status": "active"}

@api_router.get("/vendors")
async def get_vendors():
    vendors = await db.vendors.find({"status": "active"}, {"_id": 1, "name": 1, "description": 1, "cuisine_type": 1, "rating": 1, "status": 1}).to_list(1000)
    for vendor in vendors:
        vendor["id"] = str(vendor.pop("_id"))
    return vendors

@api_router.get("/vendors/{vendor_id}")
async def get_vendor(vendor_id: str):
    vendor = await db.vendors.find_one({"_id": safe_objectid(vendor_id, "Vendor")})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    vendor["id"] = str(vendor.pop("_id"))
    return vendor

# Menu Routes
@api_router.post("/menu")
async def create_menu_item(data: MenuItemCreate, user: dict = Depends(get_current_user)):
    # Only Cravitoo (master_admin) can create menu items. Vendors are read-only.
    if user["role"] != "master_admin":
        raise HTTPException(status_code=403, detail="Only Cravitoo (Master Admin) can create menu items. Vendors cannot modify menus or pricing.")
    payload = data.model_dump()
    target_vendor_id = payload.pop("vendor_id", None)
    if not target_vendor_id:
        raise HTTPException(status_code=400, detail="vendor_id is required when creating menu items")
    # Validate vendor exists
    vendor_doc = await db.vendors.find_one({"_id": safe_objectid(target_vendor_id, "Vendor")})
    if not vendor_doc:
        raise HTTPException(status_code=404, detail="Vendor not found")
    menu_doc = {
        **payload,
        "vendor_id": target_vendor_id,
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.menu_items.insert_one(menu_doc)
    return {"id": str(result.inserted_id), **payload, "vendor_id": target_vendor_id}

@api_router.get("/menu/{vendor_id}")
async def get_menu(vendor_id: str):
    menu_items = await db.menu_items.find({"vendor_id": vendor_id, "is_available": True}, {"_id": 1, "name": 1, "description": 1, "category": 1, "price": 1, "image_url": 1, "is_vegetarian": 1, "is_available": 1}).to_list(1000)
    for item in menu_items:
        item["id"] = str(item.pop("_id"))
    return menu_items

@api_router.patch("/menu/{item_id}")
async def update_menu_item(item_id: str, data: Dict[str, Any], user: dict = Depends(get_current_user)):
    # Only Cravitoo (master_admin) can edit menu items / pricing. Vendors use /menu/{id}/availability for out-of-stock only.
    if user["role"] != "master_admin":
        raise HTTPException(status_code=403, detail="Only Cravitoo (Master Admin) can update menu items or pricing. Vendors can only toggle availability.")
    # vendor_id is immutable here — strip it from update payload
    data.pop("vendor_id", None)
    result = await db.menu_items.update_one({"_id": safe_objectid(item_id, "Menu item")}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Menu item not found")
    return {"message": "Menu item updated"}

# Order Routes
@api_router.post("/orders")
async def create_order(data: OrderCreate, user: dict = Depends(get_current_user)):
    if user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees can create orders")
    
    # Server-side price validation - look up actual prices from DB
    validated_items = []
    total_amount = 0.0
    for item in data.items:
        menu_item = await db.menu_items.find_one({"_id": safe_objectid(item.menu_item_id, "Menu item")})
        if not menu_item:
            raise HTTPException(status_code=400, detail=f"Menu item {item.menu_item_id} not found")
        if not menu_item.get("is_available", False):
            raise HTTPException(status_code=400, detail=f"Menu item {menu_item['name']} is not available")
        actual_price = menu_item["price"]
        validated_items.append({
            "menu_item_id": item.menu_item_id,
            "name": menu_item["name"],
            "quantity": item.quantity,
            "price": actual_price
        })
        total_amount += actual_price * item.quantity
    
    order_doc = {
        "user_id": user["id"],
        "vendor_id": data.vendor_id,
        "items": validated_items,
        "total_amount": total_amount,
        "status": "pending",
        "payment_status": "pending",
        "delivery_type": data.delivery_type,
        "special_instructions": data.special_instructions,
        "created_at": datetime.now(timezone.utc)
    }
    
    result = await db.orders.insert_one(order_doc)
    order_id = str(result.inserted_id)
    
    # Generate QR code for pickup
    qr_code = generate_pickup_qr(order_id)
    update_doc = {"pickup_qr": qr_code}
    
    # Auto-confirm if vendor has enabled it
    vendor_doc = await db.vendors.find_one({"_id": safe_objectid(data.vendor_id, "Vendor")})
    auto_confirmed = False
    if vendor_doc and vendor_doc.get("auto_confirm"):
        update_doc["status"] = "confirmed"
        auto_confirmed = True
    
    await db.orders.update_one({"_id": result.inserted_id}, {"$set": update_doc})
    final_status = "confirmed" if auto_confirmed else "pending"
    
    # Notify vendor of new order
    vendor_users = await db.users.find({"vendor_id": data.vendor_id, "role": "vendor"}).to_list(10)
    for vu in vendor_users:
        await create_notification(
            str(vu["_id"]),
            "New Order Received",
            f"You have a new order for ₹{total_amount:.2f}",
            "order"
        )

    # Notify employee if auto-confirmed (consistency with manual confirm path)
    if auto_confirmed:
        await create_notification(
            user["id"],
            "Order Confirmed",
            f"Your order has been auto-confirmed by {vendor_doc.get('name', 'the vendor')}",
            "order"
        )
        await manager.send_to_user(user["id"], {
            "type": "order_update",
            "order_id": order_id,
            "status": "confirmed",
        })

    # Broadcast WebSocket event to vendor
    await manager.send_to_vendor(data.vendor_id, {
        "type": "new_order",
        "order_id": order_id,
        "status": final_status,
        "amount": total_amount,
        "items_count": len(validated_items)
    })

    # Low-stock alert: count items sold today per menu_item_id, notify vendor if any drops below threshold
    try:
        if vendor_doc:
            threshold = int(vendor_doc.get("low_stock_threshold", 0) or 0)
            if threshold > 0:
                today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
                for it in validated_items:
                    pipe = [
                        {"$match": {"vendor_id": data.vendor_id, "created_at": {"$gte": today_start}, "status": {"$nin": ["cancelled"]}}},
                        {"$unwind": "$items"},
                        {"$match": {"items.menu_item_id": it["menu_item_id"]}},
                        {"$group": {"_id": None, "qty": {"$sum": "$items.quantity"}}},
                    ]
                    async for r in db.orders.aggregate(pipe):
                        cur_qty = r["qty"]
                        prev_qty = cur_qty - it.get("quantity", 0)
                        crossed_low = prev_qty < threshold <= cur_qty
                        crossed_critical = prev_qty < (threshold * 2) <= cur_qty
                        if crossed_low or crossed_critical:
                            level = "Critical low stock" if crossed_critical else "Low stock alert"
                            vendor_users = await db.users.find({"vendor_id": data.vendor_id, "role": "vendor"}).to_list(10)
                            for vu in vendor_users:
                                await create_notification(
                                    str(vu["_id"]),
                                    level,
                                    f"{it.get('name', 'An item')} has sold {cur_qty} units today",
                                    "stock"
                                )
    except Exception as e:
        logger.error(f"Low-stock check failed: {e}")

    return {"id": order_id, "total_amount": total_amount, "status": final_status, "pickup_qr": qr_code}

@api_router.get("/orders")
async def get_orders(user: dict = Depends(get_current_user)):
    query = {}
    if user["role"] == "employee":
        query["user_id"] = user["id"]
    elif user["role"] == "vendor":
        query["vendor_id"] = user.get("vendor_id")
    
    orders = await db.orders.find(query, {"_id": 1, "user_id": 1, "vendor_id": 1, "items": 1, "total_amount": 1, "status": 1, "payment_status": 1, "delivery_type": 1, "created_at": 1, "pickup_qr": 1}).sort("created_at", -1).to_list(1000)
    for order in orders:
        order["id"] = str(order.pop("_id"))
    return orders

@api_router.patch("/orders/{order_id}")
async def update_order_status(order_id: str, status: OrderStatus, user: dict = Depends(get_current_user)):
    if user["role"] != "vendor":
        raise HTTPException(status_code=403, detail="Only vendors can update order status")
    
    order = await db.orders.find_one({"_id": safe_objectid(order_id, "Order"), "vendor_id": user.get("vendor_id")})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found or not yours")
    
    await db.orders.update_one(
        {"_id": safe_objectid(order_id, "Order")},
        {"$set": {"status": status.value}}
    )

    # Notify employee with push + ws
    status_messages = {
        "confirmed": "Your order has been confirmed!",
        "preparing": "Your food is being prepared",
        "ready": "Your order is ready for pickup!",
        "completed": "Order completed. Enjoy your meal!",
        "cancelled": "Your order was cancelled"
    }
    msg = status_messages.get(status.value, f"Order status: {status.value}")
    await create_notification(
        order["user_id"],
        f"Order #{order_id[-8:]}",
        msg,
        "order"
    )

    # Broadcast WebSocket event
    await manager.send_to_user(order["user_id"], {
        "type": "order_update",
        "order_id": order_id,
        "status": status.value
    })

    return {"message": "Order status updated", "status": status.value}

@api_router.post("/orders/{order_id}/verify-pickup")
async def verify_pickup(order_id: str, qr_code: str, user: dict = Depends(get_current_user)):
    if user["role"] != "vendor":
        raise HTTPException(status_code=403, detail="Only vendors can verify pickup")
    
    order = await db.orders.find_one({"_id": safe_objectid(order_id, "Order"), "vendor_id": user.get("vendor_id")})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if not verify_pickup_qr(qr_code, order_id):
        raise HTTPException(status_code=400, detail="Invalid QR code")
    
    await db.orders.update_one({"_id": safe_objectid(order_id, "Order")}, {"$set": {"status": "completed"}})
    return {"message": "Pickup verified successfully", "order_id": order_id}

# Payment Routes
@api_router.post("/payments/checkout")
async def create_checkout_session(data: CheckoutRequest, request: Request, user: dict = Depends(get_current_user)):
    order = await db.orders.find_one({"_id": safe_objectid(data.order_id, "Order")})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if order["user_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Not your order")
    
    host_url = data.origin_url
    webhook_url = f"{host_url}/api/webhook/stripe"
    stripe_checkout = StripeCheckout(api_key=os.environ["STRIPE_API_KEY"], webhook_url=webhook_url)
    
    success_url = f"{host_url}/employee/orders?session_id={{{{CHECKOUT_SESSION_ID}}}}"
    cancel_url = f"{host_url}/employee/orders"
    
    checkout_request = CheckoutSessionRequest(
        amount=order["total_amount"],
        currency="inr",
        success_url=success_url,
        cancel_url=cancel_url,
        metadata={"order_id": data.order_id, "user_id": user["id"]}
    )
    
    session = await stripe_checkout.create_checkout_session(checkout_request)
    
    await db.payment_transactions.insert_one({
        "order_id": data.order_id,
        "user_id": user["id"],
        "session_id": session.session_id,
        "amount": order["total_amount"],
        "currency": "inr",
        "payment_status": "pending",
        "created_at": datetime.now(timezone.utc)
    })
    
    return {"url": session.url, "session_id": session.session_id}

@api_router.get("/payments/status/{session_id}")
async def get_checkout_status(session_id: str, user: dict = Depends(get_current_user)):
    transaction = await db.payment_transactions.find_one({"session_id": session_id})
    if not transaction:
        raise HTTPException(status_code=404, detail="Payment session not found")
    
    # Authorization scope - only owner or vendor of the order can check
    if transaction["user_id"] != user["id"] and user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Not authorized to view this payment")
    
    stripe_checkout = StripeCheckout(api_key=os.environ["STRIPE_API_KEY"], webhook_url="")
    status = await stripe_checkout.get_checkout_status(session_id)
    
    if transaction["payment_status"] != "paid" and status.payment_status == "paid":
        await db.payment_transactions.update_one(
            {"session_id": session_id},
            {"$set": {"payment_status": "paid"}}
        )
        await db.orders.update_one(
            {"_id": ObjectId(transaction["order_id"])},
            {"$set": {"payment_status": "paid", "status": "confirmed"}}
        )
    
    return {"payment_status": status.payment_status, "amount": status.amount_total / 100}

@api_router.post("/webhook/stripe")
async def stripe_webhook(request: Request):
    body = await request.body()
    signature = request.headers.get("Stripe-Signature")
    
    stripe_checkout = StripeCheckout(api_key=os.environ["STRIPE_API_KEY"], webhook_url="")
    try:
        webhook_response = await stripe_checkout.handle_webhook(body, signature)
        if webhook_response.payment_status == "paid":
            transaction = await db.payment_transactions.find_one({"session_id": webhook_response.session_id})
            if transaction and transaction["payment_status"] != "paid":
                await db.payment_transactions.update_one(
                    {"session_id": webhook_response.session_id},
                    {"$set": {"payment_status": "paid"}}
                )
                await db.orders.update_one(
                    {"_id": ObjectId(transaction["order_id"])},
                    {"$set": {"payment_status": "paid", "status": "confirmed"}}
                )
        return {"status": "success"}
    except Exception as e:
        logger.error(f"Webhook error: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))

# AI Recommendations
@api_router.post("/ai/recommendations")
async def get_ai_recommendations(data: AIRecommendationRequest, user: dict = Depends(get_current_user)):
    chat = LlmChat(
        api_key=os.environ["EMERGENT_LLM_KEY"],
        session_id=f"recommendations_{user['id']}",
        system_message="You are a helpful AI food recommendation assistant for Cravitoo, a corporate cafeteria platform. Provide personalized meal suggestions based on user preferences."
    ).with_model("openai", "gpt-5.2")
    
    menu_items = await db.menu_items.find({"is_available": True}, {"_id": 0, "name": 1, "description": 1, "category": 1, "price": 1, "is_vegetarian": 1}).to_list(100)
    
    prompt = f"""Based on the following available menu items, recommend 3 dishes for the user.
    
User preferences: {data.user_preferences or 'No specific preferences'}
Dietary restrictions: {data.dietary_restrictions or 'None'}

Available menu items:
{menu_items}

Provide your recommendations in a friendly, concise format."""
    
    user_message = UserMessage(text=prompt)
    response = await chat.send_message(user_message)
    
    return {"recommendations": response}

# Analytics Routes
@api_router.get("/analytics/vendor")
async def get_vendor_analytics(user: dict = Depends(get_current_user)):
    if user["role"] != "vendor":
        raise HTTPException(status_code=403, detail="Only vendors can access analytics")
    
    vendor_id = user.get("vendor_id")
    total_orders = await db.orders.count_documents({"vendor_id": vendor_id})
    
    pipeline = [
        {"$match": {"vendor_id": vendor_id, "payment_status": "paid"}},
        {"$group": {"_id": None, "total_revenue": {"$sum": "$total_amount"}}}
    ]
    revenue_result = await db.orders.aggregate(pipeline).to_list(1)
    total_revenue = revenue_result[0]["total_revenue"] if revenue_result else 0
    
    return {
        "total_orders": total_orders,
        "total_revenue": total_revenue,
        "average_order_value": total_revenue / total_orders if total_orders > 0 else 0
    }

@api_router.get("/analytics/corporate")
async def get_corporate_analytics(user: dict = Depends(get_current_user)):
    if user["role"] not in ["corporate_admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {}
    if user["role"] == "corporate_admin":
        company_id = user.get("company_id")
        employee_ids = [str(u["_id"]) for u in await db.users.find({"company_id": company_id, "role": "employee"}, {"_id": 1}).to_list(1000)]
        query["user_id"] = {"$in": employee_ids}
    
    total_orders = await db.orders.count_documents(query)
    pipeline = [
        {"$match": {**query, "payment_status": "paid"}},
        {"$group": {"_id": None, "total_spend": {"$sum": "$total_amount"}}}
    ]
    spend_result = await db.orders.aggregate(pipeline).to_list(1)
    total_spend = spend_result[0]["total_spend"] if spend_result else 0
    
    return {
        "total_orders": total_orders,
        "total_spend": total_spend
    }

# Review Routes
@api_router.post("/reviews")
async def create_review(data: ReviewCreate, user: dict = Depends(get_current_user)):
    if user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees can write reviews")
    
    order = await db.orders.find_one({"_id": safe_objectid(data.order_id, "Order"), "user_id": user["id"]})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if order.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Can only review completed orders")
    
    existing = await db.reviews.find_one({"order_id": data.order_id})
    if existing:
        raise HTTPException(status_code=400, detail="Review already exists for this order")
    
    review_doc = {
        "user_id": user["id"],
        "vendor_id": data.vendor_id,
        "order_id": data.order_id,
        "rating": data.rating,
        "comment": data.comment,
        "user_name": user.get("name", "Anonymous"),
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.reviews.insert_one(review_doc)
    
    # Update vendor average rating
    pipeline = [
        {"$match": {"vendor_id": data.vendor_id}},
        {"$group": {"_id": None, "avg_rating": {"$avg": "$rating"}, "count": {"$sum": 1}}}
    ]
    rating_result = await db.reviews.aggregate(pipeline).to_list(1)
    if rating_result:
        await db.vendors.update_one(
            {"_id": safe_objectid(data.vendor_id, "Vendor")},
            {"$set": {"rating": round(rating_result[0]["avg_rating"], 1)}}
        )
    
    return {"id": str(result.inserted_id), "message": "Review submitted successfully"}

@api_router.get("/reviews/vendor/{vendor_id}")
async def get_vendor_reviews(vendor_id: str):
    reviews = await db.reviews.find({"vendor_id": vendor_id}, {"_id": 1, "rating": 1, "comment": 1, "user_name": 1, "created_at": 1}).sort("created_at", -1).to_list(100)
    for review in reviews:
        review["id"] = str(review.pop("_id"))
    return reviews

# Preferences Routes
@api_router.get("/preferences")
async def get_preferences(user: dict = Depends(get_current_user)):
    if user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees can access preferences")
    
    prefs = await db.preferences.find_one({"user_id": user["id"]})
    if not prefs:
        return {"dietary_preferences": [], "allergies": [], "favorite_cuisines": []}
    
    return {
        "dietary_preferences": prefs.get("dietary_preferences", []),
        "allergies": prefs.get("allergies", []),
        "favorite_cuisines": prefs.get("favorite_cuisines", [])
    }

@api_router.post("/preferences")
async def update_preferences(data: PreferencesUpdate, user: dict = Depends(get_current_user)):
    if user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees can update preferences")
    
    update_doc = {
        "user_id": user["id"],
        "updated_at": datetime.now(timezone.utc)
    }
    if data.dietary_preferences is not None:
        update_doc["dietary_preferences"] = data.dietary_preferences
    if data.allergies is not None:
        update_doc["allergies"] = data.allergies
    if data.favorite_cuisines is not None:
        update_doc["favorite_cuisines"] = data.favorite_cuisines
    
    await db.preferences.update_one(
        {"user_id": user["id"]},
        {"$set": update_doc},
        upsert=True
    )
    return {"message": "Preferences updated successfully"}

# Subscription Routes
@api_router.post("/subscriptions")
async def create_subscription(data: SubscriptionCreate, user: dict = Depends(get_current_user)):
    if user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees can subscribe")
    
    start_date = datetime.now(timezone.utc)
    end_date = start_date + timedelta(days=data.duration_days)
    
    sub_doc = {
        "user_id": user["id"],
        "vendor_id": data.vendor_id,
        "plan_type": data.plan_type,
        "meal_type": data.meal_type,
        "duration_days": data.duration_days,
        "start_date": start_date,
        "end_date": end_date,
        "status": "active",
        "created_at": start_date
    }
    result = await db.subscriptions.insert_one(sub_doc)
    return {"id": str(result.inserted_id), "message": "Subscription created", "end_date": end_date.isoformat()}

@api_router.get("/subscriptions")
async def get_subscriptions(user: dict = Depends(get_current_user)):
    if user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees can view subscriptions")
    
    subs = await db.subscriptions.find({"user_id": user["id"]}).sort("created_at", -1).to_list(100)
    for sub in subs:
        sub["id"] = str(sub.pop("_id"))
        if isinstance(sub.get("start_date"), datetime):
            sub["start_date"] = sub["start_date"].isoformat()
        if isinstance(sub.get("end_date"), datetime):
            sub["end_date"] = sub["end_date"].isoformat()
        if isinstance(sub.get("created_at"), datetime):
            sub["created_at"] = sub["created_at"].isoformat()
    return subs

# Notifications Helper
async def create_notification(user_id: str, title: str, message: str, notif_type: str = "info", push_data: Optional[Dict[str, Any]] = None):
    """Persist an in-app notification AND fire a push notification (if user has a registered token)."""
    await db.notifications.insert_one({
        "user_id": user_id,
        "title": title,
        "message": message,
        "type": notif_type,
        "read": False,
        "created_at": datetime.now(timezone.utc)
    })
    # Best-effort push (failures never break the calling flow)
    try:
        await send_push_to_user(user_id, title, message, push_data or {"screen": "Notifications", "type": notif_type})
    except Exception as e:
        logger.warning(f"Push send failed for user {user_id}: {e}")


# ============== EXPO PUSH NOTIFICATIONS ==============

EXPO_PUSH_URL = "https://exp.host/--/api/v2/push/send"
_push_http_client: Optional[httpx.AsyncClient] = None


def _get_push_http_client() -> httpx.AsyncClient:
    global _push_http_client
    if _push_http_client is None:
        _push_http_client = httpx.AsyncClient(timeout=10.0)
    return _push_http_client


async def send_expo_push(messages: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Send a batch of Expo push messages. messages is a list of dicts with 'to', 'title', 'body', 'data'."""
    if not messages:
        return {}
    valid_messages = [m for m in messages if m.get("to", "").startswith("ExponentPushToken[")]
    if not valid_messages:
        return {}
    try:
        client_http = _get_push_http_client()
        resp = await client_http.post(
            EXPO_PUSH_URL,
            json=valid_messages,
            headers={
                "Content-Type": "application/json",
                "Accept": "application/json",
                "Accept-Encoding": "gzip, deflate",
            },
        )
        return resp.json()
    except Exception as e:
        logger.warning(f"Expo push failed: {e}")
        return {"error": str(e)}


async def send_push_to_user(user_id: str, title: str, body: str, data: Optional[Dict[str, Any]] = None):
    """Look up all active Expo push tokens for a user and send them a push notification."""
    tokens_cursor = db.push_tokens.find({"user_id": user_id, "active": True})
    messages = []
    async for t in tokens_cursor:
        messages.append({
            "to": t["token"],
            "title": title,
            "body": body,
            "data": data or {},
            "sound": "default",
            "priority": "high",
            "channelId": "default",
        })
    if messages:
        await send_expo_push(messages)


@api_router.post("/notifications/push-token")
async def register_push_token(data: PushTokenRegister, user: dict = Depends(get_current_user)):
    """Register or refresh an Expo push token for the authenticated user."""
    if not data.token or not data.token.startswith("ExponentPushToken["):
        raise HTTPException(status_code=400, detail="Invalid Expo push token format")
    now = datetime.now(timezone.utc)
    # Upsert by (user_id, token) — same physical device only stores one row per user
    await db.push_tokens.update_one(
        {"user_id": user["id"], "token": data.token},
        {
            "$set": {
                "user_id": user["id"],
                "token": data.token,
                "platform": data.platform,
                "variant": data.variant,
                "active": True,
                "last_seen_at": now,
            },
            "$setOnInsert": {"registered_at": now},
        },
        upsert=True,
    )
    return {"ok": True}


@api_router.delete("/notifications/push-token")
async def unregister_push_token(token: str = Query(...), user: dict = Depends(get_current_user)):
    """Mark a push token inactive (on logout / app uninstall)."""
    await db.push_tokens.update_one(
        {"user_id": user["id"], "token": token},
        {"$set": {"active": False, "deactivated_at": datetime.now(timezone.utc)}},
    )
    return {"ok": True}


@api_router.post("/notifications/test-push")
async def test_push(user: dict = Depends(get_current_user)):
    """Send a test push notification to the calling user. Useful for debugging in production."""
    await send_push_to_user(
        user["id"],
        "🍴 Cravitoo test notification",
        "If you can see this, push notifications are working!",
        {"screen": "Notifications"},
    )
    return {"ok": True, "sent_to": user["id"]}


# Menu CRUD - DELETE
@api_router.delete("/menu/{item_id}")
async def delete_menu_item(item_id: str, user: dict = Depends(get_current_user)):
    # Only Cravitoo (master_admin) can delete menu items.
    if user["role"] != "master_admin":
        raise HTTPException(status_code=403, detail="Only Cravitoo (Master Admin) can delete menu items.")
    result = await db.menu_items.delete_one({"_id": safe_objectid(item_id, "Menu item")})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Menu item not found")
    return {"message": "Menu item deleted"}

@api_router.get("/menu/vendor/all")
async def get_my_menu(user: dict = Depends(get_current_user)):
    # Vendors see their own menu (read-only). Master_admin can pass vendor_id (handled by /menu/{vendor_id}).
    if user["role"] != "vendor":
        raise HTTPException(status_code=403, detail="Only vendors can access this")
    items = await db.menu_items.find({"vendor_id": user.get("vendor_id")}).to_list(1000)
    for item in items:
        item["id"] = str(item.pop("_id"))
        if isinstance(item.get("created_at"), datetime):
            item["created_at"] = item["created_at"].isoformat()
    return items

# Employee Management (Corporate Admin)
@api_router.post("/companies/employees")
async def add_employee(data: EmployeeCreate, user: dict = Depends(get_current_user)):
    if user["role"] != "corporate_admin":
        raise HTTPException(status_code=403, detail="Only corporate admins can add employees")
    
    email_lower = data.email.lower()
    existing = await db.users.find_one({"email": email_lower})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    employee_doc = {
        "email": email_lower,
        "password_hash": hash_password(data.password),
        "name": data.name,
        "role": "employee",
        "company_id": user.get("company_id"),
        "department": data.department,
        "employee_id": data.employee_id,
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.users.insert_one(employee_doc)
    return {"id": str(result.inserted_id), "email": email_lower, "name": data.name, "department": data.department}

@api_router.get("/companies/employees")
async def list_employees(user: dict = Depends(get_current_user)):
    if user["role"] not in ["corporate_admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {"role": "employee"}
    if user["role"] == "corporate_admin":
        query["company_id"] = user.get("company_id")
    
    employees = await db.users.find(query, {"_id": 1, "email": 1, "name": 1, "department": 1, "employee_id": 1, "created_at": 1}).to_list(1000)
    for emp in employees:
        emp["id"] = str(emp.pop("_id"))
        if isinstance(emp.get("created_at"), datetime):
            emp["created_at"] = emp["created_at"].isoformat()
    return employees

@api_router.delete("/companies/employees/{employee_id}")
async def remove_employee(employee_id: str, user: dict = Depends(get_current_user)):
    if user["role"] != "corporate_admin":
        raise HTTPException(status_code=403, detail="Only corporate admins can remove employees")
    
    result = await db.users.delete_one({
        "_id": safe_objectid(employee_id, "Employee"),
        "company_id": user.get("company_id"),
        "role": "employee"
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"message": "Employee removed"}

# Bulk Team Ordering
@api_router.post("/orders/bulk")
async def create_bulk_order(data: BulkOrderCreate, user: dict = Depends(get_current_user)):
    if user["role"] not in ["corporate_admin", "employee"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    bulk_order_id = str(ObjectId())
    total_amount = 0.0
    individual_orders = []
    skipped = []
    
    for bulk_item in data.orders:
        target_user = await db.users.find_one({"email": bulk_item.user_email.lower()})
        if not target_user:
            skipped.append({"email": bulk_item.user_email, "reason": "user_not_found"})
            continue
        
        validated_items = []
        order_total = 0.0
        invalid_item_ids = []
        for item in bulk_item.items:
            try:
                menu_obj_id = ObjectId(item.menu_item_id)
            except Exception:
                invalid_item_ids.append(item.menu_item_id)
                continue
            menu_item = await db.menu_items.find_one({"_id": menu_obj_id})
            if not menu_item or not menu_item.get("is_available", False):
                invalid_item_ids.append(item.menu_item_id)
                continue
            actual_price = menu_item["price"]
            validated_items.append({
                "menu_item_id": item.menu_item_id,
                "name": menu_item["name"],
                "quantity": item.quantity,
                "price": actual_price
            })
            order_total += actual_price * item.quantity
        
        if not validated_items:
            skipped.append({"email": bulk_item.user_email, "reason": "no_valid_items", "invalid_item_ids": invalid_item_ids})
            continue
        
        order_doc = {
            "user_id": str(target_user["_id"]),
            "vendor_id": data.vendor_id,
            "items": validated_items,
            "total_amount": order_total,
            "status": "pending",
            "payment_status": "paid" if data.sponsored else "pending",
            "delivery_type": data.delivery_type,
            "bulk_order_id": bulk_order_id,
            "sponsored": data.sponsored,
            "sponsored_by": user["id"] if data.sponsored else None,
            "occasion": data.occasion,
            "created_at": datetime.now(timezone.utc)
        }
        result = await db.orders.insert_one(order_doc)
        order_id = str(result.inserted_id)
        qr_code = generate_pickup_qr(order_id)
        await db.orders.update_one({"_id": result.inserted_id}, {"$set": {"pickup_qr": qr_code, "status": "confirmed" if data.sponsored else "pending"}})
        
        await create_notification(
            str(target_user["_id"]),
            f"New {'Sponsored ' if data.sponsored else ''}Order",
            f"You have a new order{' (sponsored by company)' if data.sponsored else ''}{' - ' + data.occasion if data.occasion else ''}",
            "order"
        )
        
        total_amount += order_total
        individual_orders.append({"order_id": order_id, "user_email": bulk_item.user_email, "amount": order_total})
    
    await db.bulk_orders.insert_one({
        "_id": ObjectId(bulk_order_id),
        "created_by": user["id"],
        "vendor_id": data.vendor_id,
        "total_amount": total_amount,
        "order_count": len(individual_orders),
        "sponsored": data.sponsored,
        "occasion": data.occasion,
        "created_at": datetime.now(timezone.utc)
    })
    
    return {"bulk_order_id": bulk_order_id, "total_amount": total_amount, "orders": individual_orders, "skipped": skipped}

# Event Catering
@api_router.post("/events")
async def create_event_catering(data: EventCateringCreate, user: dict = Depends(get_current_user)):
    if user["role"] not in ["corporate_admin", "employee"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    total_amount = 0.0
    validated_items = []
    for item in data.menu_items:
        menu_item = await db.menu_items.find_one({"_id": safe_objectid(item.menu_item_id, "Menu item")})
        if not menu_item:
            continue
        actual_price = menu_item["price"]
        qty_for_event = item.quantity * data.headcount
        validated_items.append({
            "menu_item_id": item.menu_item_id,
            "name": menu_item["name"],
            "quantity_per_person": item.quantity,
            "total_quantity": qty_for_event,
            "price": actual_price
        })
        total_amount += actual_price * qty_for_event
    
    event_doc = {
        "created_by": user["id"],
        "company_id": user.get("company_id"),
        "vendor_id": data.vendor_id,
        "event_name": data.event_name,
        "event_date": data.event_date,
        "headcount": data.headcount,
        "menu_items": validated_items,
        "total_amount": total_amount,
        "notes": data.notes,
        "status": "pending_approval",
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.events.insert_one(event_doc)
    return {"id": str(result.inserted_id), "total_amount": total_amount, "status": "pending_approval"}

@api_router.get("/events")
async def list_events(user: dict = Depends(get_current_user)):
    query = {}
    if user["role"] == "employee" or user["role"] == "corporate_admin":
        if user.get("company_id"):
            query["company_id"] = user.get("company_id")
        else:
            query["created_by"] = user["id"]
    elif user["role"] == "vendor":
        query["vendor_id"] = user.get("vendor_id")
    
    events = await db.events.find(query).sort("created_at", -1).to_list(500)
    for event in events:
        event["id"] = str(event.pop("_id"))
        if isinstance(event.get("created_at"), datetime):
            event["created_at"] = event["created_at"].isoformat()
    return events

@api_router.patch("/events/{event_id}/approve")
async def approve_event(event_id: str, user: dict = Depends(get_current_user)):
    if user["role"] not in ["corporate_admin", "vendor"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    event = await db.events.find_one({"_id": safe_objectid(event_id, "Event")})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    # Ownership scoping
    if user["role"] == "vendor" and event.get("vendor_id") != user.get("vendor_id"):
        raise HTTPException(status_code=403, detail="Not your event")
    if user["role"] == "corporate_admin" and event.get("company_id") != user.get("company_id"):
        raise HTTPException(status_code=403, detail="Not your company's event")
    
    await db.events.update_one(
        {"_id": safe_objectid(event_id, "Event")},
        {"$set": {"status": "approved", "approved_by": user["id"]}}
    )
    return {"message": "Event approved"}

# Notifications
@api_router.get("/notifications")
async def list_notifications(user: dict = Depends(get_current_user)):
    notifs = await db.notifications.find({"user_id": user["id"]}).sort("created_at", -1).limit(50).to_list(50)
    for n in notifs:
        n["id"] = str(n.pop("_id"))
        if isinstance(n.get("created_at"), datetime):
            n["created_at"] = n["created_at"].isoformat()
    return notifs

@api_router.patch("/notifications/{notif_id}/read")
async def mark_notification_read(notif_id: str, user: dict = Depends(get_current_user)):
    await db.notifications.update_one(
        {"_id": safe_objectid(notif_id, "Notification"), "user_id": user["id"]},
        {"$set": {"read": True}}
    )
    return {"message": "Marked as read"}

@api_router.post("/notifications/mark-all-read")
async def mark_all_notifications_read(user: dict = Depends(get_current_user)):
    await db.notifications.update_many({"user_id": user["id"], "read": False}, {"$set": {"read": True}})
    return {"message": "All notifications marked as read"}

# AI Demand Forecasting
@api_router.post("/ai/demand-forecast")
async def get_demand_forecast(user: dict = Depends(get_current_user)):
    if user["role"] != "vendor":
        raise HTTPException(status_code=403, detail="Only vendors can access demand forecasting")
    
    vendor_id = user.get("vendor_id")
    # Aggregate item-level orders
    pipeline = [
        {"$match": {"vendor_id": vendor_id, "payment_status": "paid"}},
        {"$unwind": "$items"},
        {"$group": {
            "_id": "$items.name",
            "total_quantity": {"$sum": "$items.quantity"},
            "total_revenue": {"$sum": {"$multiply": ["$items.price", "$items.quantity"]}}
        }},
        {"$sort": {"total_quantity": -1}},
        {"$limit": 10}
    ]
    top_items = await db.orders.aggregate(pipeline).to_list(10)
    
    total_orders = await db.orders.count_documents({"vendor_id": vendor_id})
    
    if total_orders < 1:
        return {"forecast": "Not enough data for forecasting. Need at least a few orders to generate predictions.", "top_items": []}
    
    chat = LlmChat(
        api_key=os.environ["EMERGENT_LLM_KEY"],
        session_id=f"forecast_{vendor_id}",
        system_message="You are an AI demand forecasting analyst for a corporate cafeteria. Provide actionable demand predictions and recommendations."
    ).with_model("openai", "gpt-5.2")
    
    prompt = f"""Based on the following order history data, provide a demand forecast for next week.

Top selling items (last period):
{top_items}

Total orders in history: {total_orders}

Provide:
1. Top 3 items expected to be in highest demand next week
2. Suggested inventory levels (low/medium/high) for each top item
3. One actionable insight to maximize revenue

Keep response concise and bullet-point friendly (under 200 words)."""
    
    response = await chat.send_message(UserMessage(text=prompt))
    
    return {
        "forecast": response,
        "top_items": [{"name": item["_id"], "quantity": item["total_quantity"], "revenue": item["total_revenue"]} for item in top_items]
    }

# AI Food Wastage Analysis
@api_router.post("/ai/wastage-analysis")
async def get_wastage_analysis(user: dict = Depends(get_current_user)):
    if user["role"] != "vendor":
        raise HTTPException(status_code=403, detail="Only vendors can access wastage analysis")
    
    vendor_id = user.get("vendor_id")
    cancelled_orders = await db.orders.count_documents({"vendor_id": vendor_id, "status": "cancelled"})
    completed_orders = await db.orders.count_documents({"vendor_id": vendor_id, "status": "completed"})
    total_orders = await db.orders.count_documents({"vendor_id": vendor_id})
    
    cancellation_rate = (cancelled_orders / total_orders * 100) if total_orders > 0 else 0
    
    chat = LlmChat(
        api_key=os.environ["EMERGENT_LLM_KEY"],
        session_id=f"wastage_{vendor_id}",
        system_message="You are a food wastage reduction expert for corporate cafeterias. Provide actionable strategies."
    ).with_model("openai", "gpt-5.2")
    
    prompt = f"""Vendor metrics:
- Total orders: {total_orders}
- Completed: {completed_orders}
- Cancelled: {cancelled_orders}
- Cancellation rate: {cancellation_rate:.1f}%

Provide 3 actionable strategies to reduce food wastage and improve order fulfillment. Keep under 150 words."""
    
    response = await chat.send_message(UserMessage(text=prompt))
    
    return {
        "analysis": response,
        "metrics": {
            "total_orders": total_orders,
            "completed_orders": completed_orders,
            "cancelled_orders": cancelled_orders,
            "cancellation_rate": round(cancellation_rate, 2)
        }
    }

# Loyalty System
@api_router.get("/loyalty")
async def get_loyalty(user: dict = Depends(get_current_user)):
    if user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees have loyalty programs")
    
    # Calculate points: 1 point per ₹100 spent on paid orders
    pipeline = [
        {"$match": {"user_id": user["id"], "payment_status": "paid"}},
        {"$group": {"_id": None, "total_spent": {"$sum": "$total_amount"}, "order_count": {"$sum": 1}}}
    ]
    result = await db.orders.aggregate(pipeline).to_list(1)
    
    total_spent = result[0]["total_spent"] if result else 0
    order_count = result[0]["order_count"] if result else 0
    
    # Calculate points earned (1 per 100 INR)
    points_earned = int(total_spent / 100)
    
    # Get redeemed points
    redeemed = await db.loyalty_redemptions.find({"user_id": user["id"]}).to_list(1000)
    points_redeemed = sum(r.get("points", 0) for r in redeemed)
    
    available_points = points_earned - points_redeemed
    
    # Tier calculation
    if total_spent >= 10000:
        tier = "Gold"
        next_tier_at = None
    elif total_spent >= 5000:
        tier = "Silver"
        next_tier_at = 10000 - total_spent
    elif total_spent >= 1000:
        tier = "Bronze"
        next_tier_at = 5000 - total_spent
    else:
        tier = "Starter"
        next_tier_at = 1000 - total_spent
    
    return {
        "tier": tier,
        "total_spent": total_spent,
        "order_count": order_count,
        "points_earned": points_earned,
        "points_redeemed": points_redeemed,
        "available_points": available_points,
        "next_tier_at": next_tier_at,
        "point_value_inr": 1  # 1 point = 1 INR discount
    }

@api_router.post("/loyalty/redeem")
async def redeem_loyalty(data: LoyaltyRedeemRequest, user: dict = Depends(get_current_user)):
    if user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees can redeem points")

    # Validate order first (better error message ordering)
    order = await db.orders.find_one({"_id": safe_objectid(data.order_id, "Order"), "user_id": user["id"]})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("payment_status") == "paid":
        raise HTTPException(status_code=400, detail="Cannot redeem on already paid order")
    existing_redemption = await db.loyalty_redemptions.find_one({"order_id": data.order_id})
    if existing_redemption:
        raise HTTPException(status_code=400, detail="Points already redeemed for this order")

    # Now validate points
    loyalty = await get_loyalty(user)
    if data.points < 100:
        raise HTTPException(status_code=400, detail="Minimum 100 points to redeem")
    if data.points > loyalty["available_points"]:
        raise HTTPException(status_code=400, detail="Insufficient points")

    discount = min(data.points, order["total_amount"])
    new_total = max(0, order["total_amount"] - discount)

    await db.loyalty_redemptions.insert_one({
        "user_id": user["id"],
        "order_id": data.order_id,
        "points": discount,
        "discount_inr": discount,
        "created_at": datetime.now(timezone.utc)
    })

    await db.orders.update_one(
        {"_id": safe_objectid(data.order_id, "Order")},
        {"$set": {"total_amount": new_total, "loyalty_discount": discount}}
    )

    return {"message": f"{discount} points redeemed", "discount_inr": discount, "new_total": new_total}

# ============== RAZORPAY ==============

RAZORPAY_MOCK_MODE = os.environ.get('RAZORPAY_MOCK_MODE', 'true').lower() == 'true'

def get_razorpay_client():
    """Get razorpay client if not in mock mode, else None"""
    if RAZORPAY_MOCK_MODE:
        return None
    return razorpay.Client(auth=(os.environ['RAZORPAY_KEY_ID'], os.environ['RAZORPAY_KEY_SECRET']))


@api_router.post("/payments/razorpay/create-order")
async def razorpay_create_order(data: RazorpayOrderCreate, user: dict = Depends(get_current_user)):
    """Create a Razorpay order linked to a Cravitoo order. Works in mock mode."""
    order = await db.orders.find_one({"_id": safe_objectid(data.order_id, "Order"), "user_id": user["id"]})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("payment_status") == "paid":
        raise HTTPException(status_code=400, detail="Order already paid")

    amount_paise = int(order["total_amount"] * 100)

    if RAZORPAY_MOCK_MODE:
        # Mock: generate a fake razorpay_order_id
        razorpay_order_id = f"order_mock_{secrets.token_hex(8)}"
        await db.payment_transactions.insert_one({
            "order_id": data.order_id,
            "user_id": user["id"],
            "provider": "razorpay",
            "razorpay_order_id": razorpay_order_id,
            "amount": amount_paise,
            "currency": "INR",
            "payment_status": "created",
            "mock_mode": True,
            "created_at": datetime.now(timezone.utc)
        })
        return {
            "razorpay_order_id": razorpay_order_id,
            "amount": amount_paise,
            "currency": "INR",
            "key_id": os.environ['RAZORPAY_KEY_ID'],
            "mock_mode": True,
            "cravitoo_order_id": data.order_id
        }
    else:
        client_rzp = get_razorpay_client()
        razor_order = client_rzp.order.create({
            "amount": amount_paise,
            "currency": "INR",
            "payment_capture": 1,
            "receipt": data.order_id[:40]
        })
        await db.payment_transactions.insert_one({
            "order_id": data.order_id,
            "user_id": user["id"],
            "provider": "razorpay",
            "razorpay_order_id": razor_order["id"],
            "amount": amount_paise,
            "currency": "INR",
            "payment_status": "created",
            "mock_mode": False,
            "created_at": datetime.now(timezone.utc)
        })
        return {
            "razorpay_order_id": razor_order["id"],
            "amount": amount_paise,
            "currency": "INR",
            "key_id": os.environ['RAZORPAY_KEY_ID'],
            "mock_mode": False,
            "cravitoo_order_id": data.order_id
        }

@api_router.post("/payments/razorpay/verify")
async def razorpay_verify(data: RazorpayVerify, user: dict = Depends(get_current_user)):
    """Verify Razorpay payment signature and mark order as paid."""
    order = await db.orders.find_one({"_id": safe_objectid(data.order_id, "Order"), "user_id": user["id"]})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    if RAZORPAY_MOCK_MODE:
        # Mock: skip signature verification, accept any payment_id
        pass
    else:
        # Verify HMAC signature
        body = f"{data.razorpay_order_id}|{data.razorpay_payment_id}"
        expected_signature = hmac.new(
            os.environ['RAZORPAY_KEY_SECRET'].encode(),
            body.encode(),
            hashlib.sha256
        ).hexdigest()
        if not hmac.compare_digest(expected_signature, data.razorpay_signature):
            raise HTTPException(status_code=400, detail="Invalid payment signature")

    # Mark order as paid
    await db.payment_transactions.update_one(
        {"razorpay_order_id": data.razorpay_order_id},
        {"$set": {
            "payment_status": "paid",
            "razorpay_payment_id": data.razorpay_payment_id,
            "razorpay_signature": data.razorpay_signature,
            "paid_at": datetime.now(timezone.utc)
        }}
    )
    await db.orders.update_one(
        {"_id": safe_objectid(data.order_id, "Order")},
        {"$set": {"payment_status": "paid", "status": "confirmed"}}
    )

    # Notify vendor (push + websocket)
    vendor_users = await db.users.find({"vendor_id": order["vendor_id"], "role": "vendor"}).to_list(10)
    for vu in vendor_users:
        await create_notification(
            str(vu["_id"]),
            "Order Paid & Confirmed",
            f"Order #{data.order_id[-8:]} has been paid. Total ₹{order['total_amount']:.2f}",
            "order"
        )

    # Broadcast via WebSocket
    await manager.send_to_user(user["id"], {
        "type": "order_update",
        "order_id": data.order_id,
        "status": "confirmed",
        "payment_status": "paid"
    })
    await manager.send_to_vendor(order["vendor_id"], {
        "type": "new_order",
        "order_id": data.order_id,
        "status": "confirmed",
        "amount": order["total_amount"]
    })

    return {"verified": True, "payment_status": "paid", "order_status": "confirmed"}

# ============== ORDER CANCELLATION & REFUND ==============

CANCEL_WINDOW_SECONDS = int(os.environ.get('CANCEL_WINDOW_SECONDS', '300'))  # 5 min default

@api_router.post("/orders/{order_id}/cancel")
async def cancel_order(order_id: str, user: dict = Depends(get_current_user)):
    """Customer cancels their own order — only within CANCEL_WINDOW_SECONDS and before vendor confirms."""
    order = await db.orders.find_one({"_id": safe_objectid(order_id, "Order")})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if user["role"] != "employee" or order.get("user_id") != user["id"]:
        raise HTTPException(status_code=403, detail="You can only cancel your own orders")
    if order.get("status") == "cancelled":
        raise HTTPException(status_code=400, detail="Order is already cancelled")
    if order.get("status") not in ("pending",):
        raise HTTPException(status_code=400, detail=f"Cannot cancel order with status '{order.get('status')}' — vendor has already started preparing it")

    created_at = order.get("created_at")
    if isinstance(created_at, datetime):
        # MongoDB returns naive datetimes — assume UTC
        if created_at.tzinfo is None:
            created_at = created_at.replace(tzinfo=timezone.utc)
        elapsed = (datetime.now(timezone.utc) - created_at).total_seconds()
        if elapsed > CANCEL_WINDOW_SECONDS:
            raise HTTPException(
                status_code=400,
                detail=f"Cancellation window of {CANCEL_WINDOW_SECONDS // 60} minutes has passed",
            )

    # If paid → mark for refund (mock-mode auto-refund; real-mode would call Razorpay refund API)
    refund_status = None
    if order.get("payment_status") == "paid":
        if RAZORPAY_MOCK_MODE:
            refund_status = "refunded_mock"
        else:
            try:
                tx = await db.payment_transactions.find_one({"cravitoo_order_id": order_id})
                pay_id = tx and tx.get("razorpay_payment_id")
                if pay_id:
                    client_rp = get_razorpay_client()
                    client_rp.payment.refund(pay_id, {"amount": int(order["total_amount"] * 100)})
                    refund_status = "refunded"
                else:
                    refund_status = "refund_pending"
            except Exception as e:
                logger.error(f"Razorpay refund failed for order {order_id}: {e}")
                refund_status = "refund_failed"

    await db.orders.update_one(
        {"_id": safe_objectid(order_id, "Order")},
        {"$set": {
            "status": "cancelled",
            "cancelled_at": datetime.now(timezone.utc),
            "cancelled_by": "customer",
            **({"refund_status": refund_status} if refund_status else {}),
        }}
    )

    # Notify vendor
    vendor_users = await db.users.find({"vendor_id": order["vendor_id"], "role": "vendor"}).to_list(10)
    for vu in vendor_users:
        await create_notification(
            str(vu["_id"]),
            "Order Cancelled",
            f"Customer cancelled order #{order_id[-8:]}",
            "order"
        )
    await manager.send_to_vendor(order["vendor_id"], {
        "type": "order_update",
        "order_id": order_id,
        "status": "cancelled",
    })
    await manager.send_to_user(user["id"], {
        "type": "order_update",
        "order_id": order_id,
        "status": "cancelled",
    })

    return {"message": "Order cancelled", "refund_status": refund_status}

@api_router.post("/orders/{order_id}/refund")
async def refund_order(order_id: str, user: dict = Depends(get_current_user)):
    """Vendor or master_admin issues a refund (e.g. customer no-show, food unavailable)."""
    if user["role"] not in ("vendor", "master_admin"):
        raise HTTPException(status_code=403, detail="Only vendors or master admin can issue refunds")
    order = await db.orders.find_one({"_id": safe_objectid(order_id, "Order")})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if user["role"] == "vendor" and order.get("vendor_id") != user.get("vendor_id"):
        raise HTTPException(status_code=403, detail="Not your order")
    if order.get("payment_status") != "paid":
        raise HTTPException(status_code=400, detail="Order is not paid — nothing to refund")
    if order.get("refund_status") in ("refunded", "refunded_mock"):
        raise HTTPException(status_code=400, detail="Order already refunded")

    refund_status = "refunded_mock"
    if not RAZORPAY_MOCK_MODE:
        try:
            tx = await db.payment_transactions.find_one({"cravitoo_order_id": order_id})
            pay_id = tx and tx.get("razorpay_payment_id")
            if pay_id:
                client_rp = get_razorpay_client()
                client_rp.payment.refund(pay_id, {"amount": int(order["total_amount"] * 100)})
                refund_status = "refunded"
            else:
                refund_status = "refund_pending"
        except Exception as e:
            logger.error(f"Razorpay refund failed for order {order_id}: {e}")
            raise HTTPException(status_code=500, detail="Refund failed at gateway")

    await db.orders.update_one(
        {"_id": safe_objectid(order_id, "Order")},
        {"$set": {
            "status": "cancelled",
            "refund_status": refund_status,
            "refunded_at": datetime.now(timezone.utc),
            "cancelled_by": user["role"],
        }}
    )

    await create_notification(
        order["user_id"],
        "Order Refunded",
        f"Your order #{order_id[-8:]} has been refunded (₹{order['total_amount']:.2f})",
        "order"
    )
    await manager.send_to_user(order["user_id"], {
        "type": "order_update",
        "order_id": order_id,
        "status": "cancelled",
        "refund_status": refund_status,
    })

    return {"message": "Refund issued", "refund_status": refund_status, "amount": order["total_amount"]}

# ============== VENDOR EARNINGS, SETTLEMENT, SETTINGS ==============

@api_router.get("/vendor/today-earnings")
async def vendor_today_earnings(user: dict = Depends(get_current_user)):
    if user["role"] != "vendor":
        raise HTTPException(status_code=403, detail="Only vendors")
    vendor_id = user.get("vendor_id")
    if not vendor_id:
        return {"orders": 0, "revenue": 0.0, "completed": 0, "pending": 0}

    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    pipe = [
        {"$match": {"vendor_id": vendor_id, "created_at": {"$gte": today_start}}},
        {"$group": {
            "_id": "$status",
            "count": {"$sum": 1},
            "revenue": {"$sum": {"$cond": [{"$eq": ["$payment_status", "paid"]}, "$total_amount", 0]}},
        }}
    ]
    rows = await db.orders.aggregate(pipe).to_list(20)
    total_orders, total_rev, completed, pending = 0, 0.0, 0, 0
    for r in rows:
        total_orders += r["count"]
        total_rev += r["revenue"]
        if r["_id"] in ("completed", "ready"):
            completed += r["count"]
        elif r["_id"] in ("pending", "confirmed", "preparing"):
            pending += r["count"]
    return {"orders": total_orders, "revenue": round(total_rev, 2), "completed": completed, "pending": pending}


@api_router.get("/vendor/settlement")
async def vendor_settlement(days: int = 7, user: dict = Depends(get_current_user)):
    """Vendor's daily settlement: revenue, commission, net payout for last N days."""
    if user["role"] != "vendor":
        raise HTTPException(status_code=403, detail="Only vendors")
    vendor_id = user.get("vendor_id")
    days = max(1, min(days, 90))
    since = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=days - 1)

    # Vendor's commission % (default 15)
    vendor_doc = await db.vendors.find_one({"_id": safe_objectid(vendor_id, "Vendor")}) if vendor_id else None
    commission_pct = float(vendor_doc.get("commission_pct", 15.0)) if vendor_doc else 15.0

    pipe = [
        {"$match": {"vendor_id": vendor_id, "payment_status": "paid", "created_at": {"$gte": since}}},
        {"$group": {
            "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$created_at"}},
            "orders": {"$sum": 1},
            "gross": {"$sum": "$total_amount"},
        }},
        {"$sort": {"_id": 1}},
    ]
    rows = await db.orders.aggregate(pipe).to_list(100)
    daily = []
    total_gross = 0.0
    total_orders = 0
    for r in rows:
        gross = r["gross"]
        commission = round(gross * commission_pct / 100, 2)
        payout = round(gross - commission, 2)
        daily.append({
            "date": r["_id"],
            "orders": r["orders"],
            "gross": round(gross, 2),
            "commission": commission,
            "payout": payout,
        })
        total_gross += gross
        total_orders += r["orders"]
    total_commission = round(total_gross * commission_pct / 100, 2)
    total_payout = round(total_gross - total_commission, 2)
    return {
        "commission_pct": commission_pct,
        "days": days,
        "daily": daily,
        "total_orders": total_orders,
        "total_gross": round(total_gross, 2),
        "total_commission": total_commission,
        "total_payout": total_payout,
    }


@api_router.get("/vendor/settings")
async def get_vendor_settings(user: dict = Depends(get_current_user)):
    if user["role"] != "vendor":
        raise HTTPException(status_code=403, detail="Only vendors")
    vendor_id = user.get("vendor_id")
    if not vendor_id:
        raise HTTPException(status_code=404, detail="No vendor linked to this account")
    vendor = await db.vendors.find_one({"_id": safe_objectid(vendor_id, "Vendor")})
    return {
        "auto_confirm": bool(vendor.get("auto_confirm", False)),
        "low_stock_threshold": int(vendor.get("low_stock_threshold", 5)),
        "commission_pct": float(vendor.get("commission_pct", 15.0)),
    }


@api_router.patch("/vendor/settings")
async def update_vendor_settings(updates: Dict[str, Any], user: dict = Depends(get_current_user)):
    if user["role"] != "vendor":
        raise HTTPException(status_code=403, detail="Only vendors")
    vendor_id = user.get("vendor_id")
    allowed = {"auto_confirm", "low_stock_threshold"}
    cleaned = {k: v for k, v in updates.items() if k in allowed}
    if not cleaned:
        raise HTTPException(status_code=400, detail="No valid fields")
    if "auto_confirm" in cleaned:
        cleaned["auto_confirm"] = bool(cleaned["auto_confirm"])
    if "low_stock_threshold" in cleaned:
        cleaned["low_stock_threshold"] = max(0, int(cleaned["low_stock_threshold"]))
    await db.vendors.update_one({"_id": safe_objectid(vendor_id, "Vendor")}, {"$set": cleaned})
    return {"message": "Settings updated", **cleaned}


@api_router.patch("/menu/{item_id}/availability")
async def quick_toggle_menu_availability(item_id: str, user: dict = Depends(get_current_user)):
    """Vendor-only quick toggle for own menu items (different from site-control)."""
    if user["role"] != "vendor":
        raise HTTPException(status_code=403, detail="Only vendors")
    vendor_id = user.get("vendor_id")
    item = await db.menu_items.find_one({"_id": safe_objectid(item_id, "Menu item"), "vendor_id": vendor_id})
    if not item:
        raise HTTPException(status_code=404, detail="Menu item not found")
    new_avail = not bool(item.get("is_available", True))
    await db.menu_items.update_one({"_id": item["_id"]}, {"$set": {"is_available": new_avail}})
    return {"id": item_id, "is_available": new_avail}


# ============== FILE UPLOAD (LOCAL STORAGE) ==============

UPLOAD_DIR = Path(os.environ.get('UPLOAD_DIR', '/tmp/cravitoo_uploads'))
try:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
except (PermissionError, OSError) as e:
    # Fallback to /tmp if primary dir not writable (read-only k8s rootfs)
    UPLOAD_DIR = Path('/tmp/cravitoo_uploads')
    try:
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    except Exception as e2:
        logger.error(f"Could not create UPLOAD_DIR: {e2}. File uploads will fail until fixed.")

@api_router.post("/upload/menu-image")
async def upload_menu_image(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    # Only Cravitoo (master/site_admin) can upload menu images. Vendors are read-only.
    if user["role"] not in ("master_admin", "site_admin"):
        raise HTTPException(status_code=403, detail="Only Cravitoo admins can upload menu images.")
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="File must be an image")
    ext = (file.filename or "img").rsplit(".", 1)[-1].lower()[:5]
    if ext not in ("png", "jpg", "jpeg", "webp", "gif"):
        ext = "png"
    fname = f"{uuid.uuid4().hex}.{ext}"
    fpath = UPLOAD_DIR / fname
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Image must be under 5 MB")
    # Validate it's actually an image (not just a relabeled file)
    try:
        from PIL import Image as PILImage
        import io as _io
        img = PILImage.open(_io.BytesIO(content))
        img.verify()
    except Exception:
        raise HTTPException(status_code=400, detail="File is not a valid image")
    with open(fpath, "wb") as f:
        f.write(content)
    base = os.environ.get('PUBLIC_BACKEND_URL', '').rstrip('/')
    url = f"{base}/api/uploads/{fname}" if base else f"/api/uploads/{fname}"
    return {"url": url, "filename": fname, "size": len(content)}


@api_router.get("/uploads/{filename}")
async def serve_upload(filename: str):
    # Sanitize filename — only allow alphanumeric + .
    if not re.match(r'^[a-f0-9]+\.[a-z]+$', filename):
        raise HTTPException(status_code=400, detail="Invalid filename")
    fpath = UPLOAD_DIR / filename
    if not fpath.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(str(fpath))


# ============== MASTER ADMIN: VENDOR COMMISSION ==============

@api_router.patch("/admin/vendors/{vendor_id}/commission")
async def set_vendor_commission(vendor_id: str, payload: Dict[str, Any], user: dict = Depends(get_current_user)):
    if not is_master_admin(user):
        raise HTTPException(status_code=403, detail="Only master admin")
    pct = payload.get("commission_pct")
    if pct is None or not isinstance(pct, (int, float)) or pct < 0 or pct > 50:
        raise HTTPException(status_code=400, detail="commission_pct must be between 0 and 50")
    await db.vendors.update_one({"_id": safe_objectid(vendor_id, "Vendor")}, {"$set": {"commission_pct": float(pct)}})
    return {"vendor_id": vendor_id, "commission_pct": float(pct)}


@api_router.patch("/admin/vendors/{vendor_id}")
async def update_vendor(vendor_id: str, payload: Dict[str, Any], user: dict = Depends(get_current_user)):
    """Master admin edits vendor profile (name, cuisine, contact, address, status)."""
    if not is_master_admin(user):
        raise HTTPException(status_code=403, detail="Only master admin")
    allowed = {"name", "description", "cuisine_type", "phone", "email", "address",
               "status", "commission_pct", "image_url"}
    cleaned = {k: v for k, v in payload.items() if k in allowed}
    if not cleaned:
        raise HTTPException(status_code=400, detail="No valid fields")
    if "commission_pct" in cleaned:
        v = float(cleaned["commission_pct"])
        if v < 0 or v > 50:
            raise HTTPException(status_code=400, detail="commission_pct must be between 0 and 50")
        cleaned["commission_pct"] = v
    if "status" in cleaned and cleaned["status"] not in ("active", "inactive", "suspended"):
        raise HTTPException(status_code=400, detail="status must be active|inactive|suspended")
    await db.vendors.update_one({"_id": safe_objectid(vendor_id, "Vendor")}, {"$set": cleaned})
    await audit_log(user, "vendor", vendor_id, "updated", cleaned)
    return {"message": "Vendor updated"}


# ============== EMPLOYEE: REFUNDS, FAVOURITES ==============

@api_router.get("/refunds")
async def employee_refunds(user: dict = Depends(get_current_user)):
    """Employee sees their own refunded/cancelled orders."""
    if user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees")
    cursor = db.orders.find({
        "user_id": user["id"],
        "$or": [
            {"status": "cancelled"},
            {"refund_status": {"$exists": True, "$nin": [None, ""]}},
        ],
    }).sort("created_at", -1).limit(100)
    out = []
    async for o in cursor:
        out.append({
            "order_id": str(o["_id"]),
            "vendor_id": o.get("vendor_id"),
            "total_amount": o.get("total_amount", 0),
            "status": o.get("status"),
            "refund_status": o.get("refund_status"),
            "payment_status": o.get("payment_status"),
            "cancelled_at": o.get("cancelled_at").isoformat() if o.get("cancelled_at") else None,
            "refunded_at": o.get("refunded_at").isoformat() if o.get("refunded_at") else None,
            "cancelled_by": o.get("cancelled_by"),
            "created_at": o.get("created_at").isoformat() if o.get("created_at") else None,
        })
    return out


@api_router.get("/favorites")
async def list_favorites(user: dict = Depends(get_current_user)):
    """List employee's favorite vendors."""
    if user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees")
    favs = []
    async for f in db.favorites.find({"user_id": user["id"]}).sort("created_at", -1):
        vendor = await db.vendors.find_one({"_id": safe_objectid(f["vendor_id"], "Vendor")})
        if vendor:
            favs.append({
                "vendor_id": f["vendor_id"],
                "name": vendor.get("name"),
                "cuisine_type": vendor.get("cuisine_type"),
                "rating": vendor.get("rating", 0),
                "image_url": vendor.get("image_url"),
                "favorited_at": f.get("created_at").isoformat() if f.get("created_at") else None,
            })
    return favs


@api_router.post("/favorites/{vendor_id}")
async def add_favorite(vendor_id: str, user: dict = Depends(get_current_user)):
    if user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees")
    vendor = await db.vendors.find_one({"_id": safe_objectid(vendor_id, "Vendor")})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    if await db.favorites.find_one({"user_id": user["id"], "vendor_id": vendor_id}):
        return {"message": "Already favorited"}
    await db.favorites.insert_one({
        "user_id": user["id"],
        "vendor_id": vendor_id,
        "created_at": datetime.now(timezone.utc),
    })
    return {"message": "Added to favorites"}


@api_router.delete("/favorites/{vendor_id}")
async def remove_favorite(vendor_id: str, user: dict = Depends(get_current_user)):
    if user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees")
    await db.favorites.delete_one({"user_id": user["id"], "vendor_id": vendor_id})
    return {"message": "Removed from favorites"}


@api_router.get("/orders/last")
async def get_last_order(user: dict = Depends(get_current_user)):
    """Returns the employee's most recent order to enable 'reorder my usual'."""
    if user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees")
    order = await db.orders.find_one(
        {"user_id": user["id"], "status": {"$in": ["completed", "ready", "confirmed", "preparing"]}},
        sort=[("created_at", -1)]
    )
    if not order:
        raise HTTPException(status_code=404, detail="No previous orders to reorder")
    return {
        "vendor_id": order.get("vendor_id"),
        "items": order.get("items", []),
        "total_amount": order.get("total_amount"),
    }


# ============== ONBOARDING: BULK EXCEL IMPORT ==============

@api_router.post("/onboarding/vendors/bulk-import")
async def bulk_import_onboardings(
    file: UploadFile = File(...),
    site_id: str = "",
    user: dict = Depends(get_current_user),
):
    """Master/Site/City admin uploads an Excel sheet to bulk-create onboarding records.
    Columns: vendor_name, company_name, contact_person, mobile_number, email, business_address, cuisine_type"""
    if user["role"] not in ("master_admin", "city_admin", "site_admin"):
        raise HTTPException(status_code=403, detail="Access denied")
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")
    if user["role"] == "site_admin":
        site_id = user.get("site_id") or site_id
    if not site_id:
        raise HTTPException(status_code=400, detail="site_id required")
    site = await db.sites.find_one({"_id": safe_objectid(site_id, "Site")})
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    if user["role"] == "site_admin" and site_id != user.get("site_id"):
        raise HTTPException(status_code=403, detail="Not your site")

    content = await file.read()
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File must be under 2 MB")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel: {e}")

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel must have header + at least 1 data row")
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    required = ["vendor_name", "company_name", "contact_person", "mobile_number", "email", "business_address"]
    missing = [r for r in required if r not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(missing)}")

    inserted, errors = 0, []
    for idx, row in enumerate(rows[1:], start=2):
        try:
            rec = {headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(min(len(headers), len(row)))}
            if not rec.get("vendor_name") or not rec.get("email"):
                errors.append({"row": idx, "error": "vendor_name and email required"})
                continue
            doc = {
                "vendor_name": rec.get("vendor_name"),
                "company_name": rec.get("company_name", ""),
                "contact_person": rec.get("contact_person", ""),
                "mobile_number": rec.get("mobile_number", ""),
                "email": rec.get("email"),
                "business_address": rec.get("business_address", ""),
                "cuisine_type": rec.get("cuisine_type", "Multi-cuisine"),
                "site_id": site_id,
                "city_id": site.get("city_id"),
                "status": "draft",
                "checklist": {},
                "documents": {},
                "remarks": [],
                "created_by": user["id"],
                "created_at": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc),
            }
            res = await db.vendor_onboarding.insert_one(doc)
            await audit_log(user, "vendor_onboarding", str(res.inserted_id), "bulk_imported", {"vendor_name": rec.get("vendor_name")})
            inserted += 1
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})

    return {"inserted": inserted, "errors": errors, "total_attempted": len(rows) - 1}


# ============== ONBOARDING: MENU PRE-LOAD ==============

@api_router.post("/onboarding/vendors/{onb_id}/menu/upload-excel")
async def onboarding_menu_excel(
    onb_id: str,
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    """Pre-load menu items as 'draft' under onboarding — get activated when vendor is approved."""
    if user["role"] not in ("master_admin", "city_admin", "site_admin"):
        raise HTTPException(status_code=403, detail="Access denied")
    o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
    if not o:
        raise HTTPException(status_code=404, detail="Onboarding not found")
    if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
        raise HTTPException(status_code=403, detail="Not your site")
    if o.get("status") in ("approved", "active", "rejected"):
        raise HTTPException(status_code=400, detail=f"Cannot edit menu — status is '{o.get('status')}'")
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel: {e}")
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel must have header + at least 1 data row")
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    required = ["name", "category", "price"]
    missing = [r for r in required if r not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(missing)}")
    items = []
    errors = []
    for idx, row in enumerate(rows[1:], start=2):
        try:
            rec = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            name = (str(rec.get("name") or "")).strip()
            if not name:
                errors.append({"row": idx, "error": "Missing name"})
                continue
            price = float(rec.get("price") or 0)
            if price <= 0:
                errors.append({"row": idx, "error": f"Invalid price for {name}"})
                continue
            items.append({
                "name": name,
                "description": str(rec.get("description") or ""),
                "category": str(rec.get("category") or "Main").strip(),
                "price": price,
                "is_vegetarian": str(rec.get("is_vegetarian", "")).lower() in ("true", "yes", "1", "veg"),
                "image_url": str(rec.get("image_url") or "") or None,
                "is_available": True,
            })
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    await db.vendor_onboarding.update_one({"_id": o["_id"]}, {"$set": {"draft_menu": items, "updated_at": datetime.now(timezone.utc)}})
    # Auto-tick "menu_uploaded" in checklist
    new_checklist = {**(o.get("checklist", {})), "menu_uploaded": True}
    await db.vendor_onboarding.update_one({"_id": o["_id"]}, {"$set": {"checklist": new_checklist}})
    await audit_log(user, "vendor_onboarding", onb_id, "menu_uploaded", {"items": len(items)})
    return {"inserted": len(items), "errors": errors, "total_attempted": len(rows) - 1}


# ============== EMPLOYEE: CURRENT MEAL PERIOD ==============

def get_current_meal_period_default():
    """Return the current meal period based on IST time (no per-site schedule)."""
    from datetime import timezone as tz, timedelta as td
    ist = datetime.now(tz(td(hours=5, minutes=30)))
    h = ist.hour + ist.minute / 60
    if 6 <= h < 11:
        return "breakfast"
    if 11 <= h < 16:
        return "lunch"
    if 16 <= h < 19:
        return "snacks"
    if 19 <= h < 23:
        return "dinner"
    return None


@api_router.get("/meal-period/current")
async def get_current_meal_period_api():
    """Public endpoint — clients use this to filter menu by meal type."""
    return {"period": get_current_meal_period_default()}


# ============== CITIES & CITY ADMINS ==============

def is_city_admin(user):
    return user.get("role") == "city_admin"

def is_city_or_above(user):
    return user.get("role") in ("master_admin", "city_admin")

async def can_access_city(user, city_id):
    if is_master_admin(user):
        return True
    if user.get("role") == "city_admin" and user.get("city_id") == city_id:
        return True
    return False

async def audit_log(user, entity_type, entity_id, action, details=None):
    """Persist an audit trail entry."""
    await db.audit_log.insert_one({
        "user_id": user.get("id"),
        "user_email": user.get("email"),
        "user_role": user.get("role"),
        "entity_type": entity_type,  # "vendor_onboarding" | "city" | "vendor" etc
        "entity_id": entity_id,
        "action": action,  # "created" | "updated" | "approved" | "rejected" | "uploaded_doc" etc
        "details": details or {},
        "created_at": datetime.now(timezone.utc),
    })

@api_router.post("/cities")
async def create_city(data: CityCreate, user: dict = Depends(get_current_user)):
    if not is_master_admin(user):
        raise HTTPException(status_code=403, detail="Only master admin")
    if await db.cities.find_one({"name": data.name, "state": data.state}):
        raise HTTPException(status_code=400, detail="City already exists")
    doc = {
        "name": data.name,
        "state": data.state,
        "country": data.country,
        "status": "active",
        "created_at": datetime.now(timezone.utc),
    }
    res = await db.cities.insert_one(doc)
    city_id = str(res.inserted_id)
    await audit_log(user, "city", city_id, "created", {"name": data.name, "state": data.state})
    return {"id": city_id, **data.model_dump()}

@api_router.get("/cities")
async def list_cities(user: dict = Depends(get_current_user)):
    """Master sees all, City Admin sees only their city, others see all active cities (for site selection)."""
    if is_master_admin(user):
        cursor = db.cities.find({})
    elif is_city_admin(user):
        cid = user.get("city_id")
        cursor = db.cities.find({"_id": safe_objectid(cid, "City")}) if cid else db.cities.find({"_id": None})
    else:
        cursor = db.cities.find({"status": "active"})
    cities = []
    async for c in cursor:
        cities.append({
            "id": str(c["_id"]),
            "name": c.get("name"),
            "state": c.get("state"),
            "country": c.get("country", "India"),
            "status": c.get("status", "active"),
        })
    # Include site count per city
    for c in cities:
        c["site_count"] = await db.sites.count_documents({"city_id": c["id"]})
        c["vendor_count"] = await db.vendor_onboarding.count_documents({"city_id": c["id"], "status": "active"})
    return cities

@api_router.get("/cities/{city_id}")
async def get_city(city_id: str, user: dict = Depends(get_current_user)):
    if not await can_access_city(user, city_id):
        raise HTTPException(status_code=403, detail="Access denied")
    city = await db.cities.find_one({"_id": safe_objectid(city_id, "City")})
    if not city:
        raise HTTPException(status_code=404, detail="City not found")
    return {
        "id": str(city["_id"]),
        "name": city.get("name"),
        "state": city.get("state"),
        "country": city.get("country", "India"),
        "status": city.get("status", "active"),
    }

@api_router.patch("/cities/{city_id}")
async def update_city(city_id: str, payload: Dict[str, Any], user: dict = Depends(get_current_user)):
    if not is_master_admin(user):
        raise HTTPException(status_code=403, detail="Only master admin")
    allowed = {"name", "state", "country", "status"}
    cleaned = {k: v for k, v in payload.items() if k in allowed}
    if not cleaned:
        raise HTTPException(status_code=400, detail="No valid fields")
    await db.cities.update_one({"_id": safe_objectid(city_id, "City")}, {"$set": cleaned})
    await audit_log(user, "city", city_id, "updated", cleaned)
    return {"message": "City updated"}

@api_router.post("/admin/city-admins")
async def create_city_admin(data: CityAdminCreate, user: dict = Depends(get_current_user)):
    if not is_master_admin(user):
        raise HTTPException(status_code=403, detail="Only master admin")
    email_lower = data.email.lower()
    if await db.users.find_one({"email": email_lower}):
        raise HTTPException(status_code=400, detail="Email already registered")
    city = await db.cities.find_one({"_id": safe_objectid(data.city_id, "City")})
    if not city:
        raise HTTPException(status_code=404, detail="City not found")
    res = await db.users.insert_one({
        "email": email_lower,
        "password_hash": hash_password(data.password),
        "name": data.name,
        "role": "city_admin",
        "city_id": data.city_id,
        "created_at": datetime.now(timezone.utc),
        "failed_attempts": 0,
    })
    user_id = str(res.inserted_id)
    await audit_log(user, "user", user_id, "created_city_admin", {"city_id": data.city_id, "email": email_lower})
    return {"id": user_id, "email": email_lower, "role": "city_admin", "city_id": data.city_id}


# ============== VENDOR ONBOARDING ==============

def calc_checklist_pct(checklist: dict) -> int:
    if not checklist:
        return 0
    done = sum(1 for f in CHECKLIST_FIELDS if checklist.get(f))
    return int(done * 100 / len(CHECKLIST_FIELDS))

def onboarding_to_dict(o):
    return {
        "id": str(o["_id"]),
        "vendor_name": o.get("vendor_name"),
        "company_name": o.get("company_name"),
        "contact_person": o.get("contact_person"),
        "mobile_number": o.get("mobile_number"),
        "email": o.get("email"),
        "business_address": o.get("business_address"),
        "cuisine_type": o.get("cuisine_type"),
        "site_id": o.get("site_id"),
        "city_id": o.get("city_id"),
        "status": o.get("status", "draft"),
        "checklist": o.get("checklist", {}),
        "checklist_pct": calc_checklist_pct(o.get("checklist", {})),
        "documents": o.get("documents", {}),
        "draft_menu": o.get("draft_menu", []),
        "vendor_id": o.get("vendor_id"),  # set when approved
        "remarks": o.get("remarks", []),
        "created_by": o.get("created_by"),
        "created_at": o.get("created_at").isoformat() if o.get("created_at") else None,
        "updated_at": o.get("updated_at").isoformat() if o.get("updated_at") else None,
    }

@api_router.post("/onboarding/vendors")
async def create_vendor_onboarding(data: VendorOnboardingBasic, user: dict = Depends(get_current_user)):
    if user["role"] not in ("site_admin", "master_admin", "city_admin"):
        raise HTTPException(status_code=403, detail="Only site_admin, city_admin, or master_admin can onboard vendors")
    # Site admin can only onboard for their own site
    if user["role"] == "site_admin" and user.get("site_id") != data.site_id:
        raise HTTPException(status_code=403, detail="You can only onboard vendors for your own site")
    site = await db.sites.find_one({"_id": safe_objectid(data.site_id, "Site")})
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    if user["role"] == "city_admin" and site.get("city_id") != user.get("city_id"):
        raise HTTPException(status_code=403, detail="Site is not in your city")
    doc = {
        **data.model_dump(),
        "city_id": site.get("city_id"),
        "status": "draft",
        "checklist": {},
        "documents": {},
        "remarks": [],
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc),
    }
    res = await db.vendor_onboarding.insert_one(doc)
    onb_id = str(res.inserted_id)
    await audit_log(user, "vendor_onboarding", onb_id, "created", {"vendor_name": data.vendor_name, "site_id": data.site_id})
    doc["_id"] = res.inserted_id
    return onboarding_to_dict(doc)

@api_router.get("/onboarding/vendors")
async def list_vendor_onboardings(
    status: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """List based on role:
    - master_admin: all
    - city_admin: only in their city
    - site_admin: only for their site
    - others: 403"""
    filt = {}
    if user["role"] == "master_admin":
        pass
    elif user["role"] == "city_admin":
        filt["city_id"] = user.get("city_id")
    elif user["role"] == "site_admin":
        filt["site_id"] = user.get("site_id")
    else:
        raise HTTPException(status_code=403, detail="Access denied")
    if status and status in ONBOARDING_STATUSES:
        filt["status"] = status
    cursor = db.vendor_onboarding.find(filt).sort("created_at", -1).limit(200)
    return [onboarding_to_dict(o) async for o in cursor]

@api_router.get("/onboarding/vendors/{onb_id}")
async def get_vendor_onboarding(onb_id: str, user: dict = Depends(get_current_user)):
    o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
    if not o:
        raise HTTPException(status_code=404, detail="Onboarding not found")
    # Role-based access
    if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
        raise HTTPException(status_code=403, detail="Not your site")
    if user["role"] == "city_admin" and o.get("city_id") != user.get("city_id"):
        raise HTTPException(status_code=403, detail="Not your city")
    if user["role"] not in ("master_admin", "city_admin", "site_admin"):
        raise HTTPException(status_code=403, detail="Access denied")
    return onboarding_to_dict(o)

@api_router.patch("/onboarding/vendors/{onb_id}")
async def update_vendor_onboarding(onb_id: str, data: VendorOnboardingUpdate, user: dict = Depends(get_current_user)):
    o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
    if not o:
        raise HTTPException(status_code=404, detail="Onboarding not found")
    if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
        raise HTTPException(status_code=403, detail="Not your site")
    if user["role"] not in ("master_admin", "city_admin", "site_admin"):
        raise HTTPException(status_code=403, detail="Access denied")
    if o.get("status") in ("approved", "active", "rejected"):
        raise HTTPException(status_code=400, detail=f"Cannot edit onboarding with status '{o.get('status')}'")
    updates = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    updates["updated_at"] = datetime.now(timezone.utc)
    await db.vendor_onboarding.update_one({"_id": o["_id"]}, {"$set": updates})
    await audit_log(user, "vendor_onboarding", onb_id, "updated", updates)
    return {"message": "Updated"}

@api_router.patch("/onboarding/vendors/{onb_id}/checklist")
async def update_checklist(onb_id: str, data: ChecklistUpdate, user: dict = Depends(get_current_user)):
    o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
    if not o:
        raise HTTPException(status_code=404, detail="Onboarding not found")
    if user["role"] not in ("master_admin", "city_admin", "site_admin"):
        raise HTTPException(status_code=403, detail="Access denied")
    if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
        raise HTTPException(status_code=403, detail="Not your site")
    updates = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    new_checklist = {**(o.get("checklist", {})), **{k: v for k, v in updates.items() if k != "notes"}}
    set_doc = {"checklist": new_checklist, "updated_at": datetime.now(timezone.utc)}
    if "notes" in updates:
        set_doc["checklist_notes"] = updates["notes"]
    await db.vendor_onboarding.update_one({"_id": o["_id"]}, {"$set": set_doc})
    pct = calc_checklist_pct(new_checklist)
    await audit_log(user, "vendor_onboarding", onb_id, "checklist_updated", {"updates": updates, "pct": pct})
    return {"checklist": new_checklist, "checklist_pct": pct}

@api_router.post("/onboarding/vendors/{onb_id}/documents/{doc_type}")
async def upload_onboarding_doc(
    onb_id: str,
    doc_type: str,
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    if doc_type not in DOC_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid doc_type. Must be one of: {', '.join(DOC_TYPES)}")
    o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
    if not o:
        raise HTTPException(status_code=404, detail="Onboarding not found")
    if user["role"] not in ("master_admin", "city_admin", "site_admin"):
        raise HTTPException(status_code=403, detail="Access denied")
    if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
        raise HTTPException(status_code=403, detail="Not your site")

    # File size limit 10 MB for compliance docs (can be larger PDFs)
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File must be under 10 MB")
    ext = (file.filename or "doc").rsplit(".", 1)[-1].lower()[:5]
    if ext not in ("pdf", "png", "jpg", "jpeg", "webp"):
        raise HTTPException(status_code=400, detail="Allowed types: PDF, PNG, JPG, JPEG, WEBP")
    fname = f"onb_{uuid.uuid4().hex}.{ext}"
    fpath = UPLOAD_DIR / fname
    with open(fpath, "wb") as f:
        f.write(content)
    base = os.environ.get('PUBLIC_BACKEND_URL', '').rstrip('/')
    url = f"{base}/api/uploads/{fname}" if base else f"/api/uploads/{fname}"
    docs = o.get("documents", {})
    docs[doc_type] = {
        "url": url,
        "filename": fname,
        "original_name": file.filename,
        "uploaded_by": user["email"],
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "size": len(content),
    }
    set_doc = {"documents": docs, "updated_at": datetime.now(timezone.utc)}
    # Auto-flip status from draft → documents_pending after first upload
    if o.get("status") == "draft":
        set_doc["status"] = "documents_pending"
    await db.vendor_onboarding.update_one({"_id": o["_id"]}, {"$set": set_doc})
    await audit_log(user, "vendor_onboarding", onb_id, "uploaded_doc", {"doc_type": doc_type, "filename": fname})
    return {"doc_type": doc_type, "url": url, "filename": fname}

@api_router.delete("/onboarding/vendors/{onb_id}/documents/{doc_type}")
async def delete_onboarding_doc(onb_id: str, doc_type: str, user: dict = Depends(get_current_user)):
    o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
    if not o:
        raise HTTPException(status_code=404, detail="Onboarding not found")
    if user["role"] not in ("master_admin", "city_admin", "site_admin"):
        raise HTTPException(status_code=403, detail="Access denied")
    if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
        raise HTTPException(status_code=403, detail="Not your site")
    if o.get("status") in ("approved", "active"):
        raise HTTPException(status_code=400, detail="Cannot delete docs after approval")
    docs = o.get("documents", {})
    if doc_type not in docs:
        raise HTTPException(status_code=404, detail="Document not found")
    docs.pop(doc_type, None)
    await db.vendor_onboarding.update_one({"_id": o["_id"]}, {"$set": {"documents": docs, "updated_at": datetime.now(timezone.utc)}})
    await audit_log(user, "vendor_onboarding", onb_id, "deleted_doc", {"doc_type": doc_type})
    return {"message": "Deleted"}

@api_router.post("/onboarding/vendors/{onb_id}/submit-to-master")
async def submit_to_master(onb_id: str, user: dict = Depends(get_current_user)):
    """Site admin submits to master after their review."""
    o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
    if not o:
        raise HTTPException(status_code=404, detail="Onboarding not found")
    if user["role"] not in ("site_admin", "city_admin", "master_admin"):
        raise HTTPException(status_code=403, detail="Access denied")
    if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
        raise HTTPException(status_code=403, detail="Not your site")
    if o.get("status") not in ("documents_pending", "under_site_review", "changes_requested"):
        raise HTTPException(status_code=400, detail=f"Cannot submit from status '{o.get('status')}'")
    pct = calc_checklist_pct(o.get("checklist", {}))
    if pct < 80:
        raise HTTPException(status_code=400, detail=f"Checklist must be at least 80% complete (currently {pct}%)")
    await db.vendor_onboarding.update_one(
        {"_id": o["_id"]},
        {"$set": {"status": "under_master_review", "site_reviewed_by": user["id"], "site_reviewed_at": datetime.now(timezone.utc), "updated_at": datetime.now(timezone.utc)}}
    )
    await audit_log(user, "vendor_onboarding", onb_id, "submitted_to_master", {"checklist_pct": pct})
    return {"message": "Submitted to master admin for final approval", "status": "under_master_review"}

@api_router.post("/onboarding/vendors/{onb_id}/site-review")
async def site_review(onb_id: str, data: OnboardingDecision, user: dict = Depends(get_current_user)):
    """Site admin/city admin reviews — Approve→sub_to_master, Reject, Request changes."""
    o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
    if not o:
        raise HTTPException(status_code=404, detail="Onboarding not found")
    if user["role"] not in ("site_admin", "city_admin", "master_admin"):
        raise HTTPException(status_code=403, detail="Access denied")
    if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
        raise HTTPException(status_code=403, detail="Not your site")
    if data.decision not in ("approve", "reject", "request_changes"):
        raise HTTPException(status_code=400, detail="decision must be approve|reject|request_changes")
    remark = {
        "stage": "site_review",
        "by": user["email"],
        "decision": data.decision,
        "remarks": data.remarks or "",
        "at": datetime.now(timezone.utc).isoformat(),
    }
    new_remarks = o.get("remarks", []) + [remark]
    set_doc = {"remarks": new_remarks, "updated_at": datetime.now(timezone.utc)}
    if data.decision == "approve":
        pct = calc_checklist_pct(o.get("checklist", {}))
        if pct < 80:
            raise HTTPException(status_code=400, detail=f"Checklist must be at least 80% complete to approve (currently {pct}%)")
        set_doc["status"] = "under_master_review"
        set_doc["site_reviewed_by"] = user["id"]
        set_doc["site_reviewed_at"] = datetime.now(timezone.utc)
    elif data.decision == "reject":
        set_doc["status"] = "rejected"
    else:  # request_changes
        set_doc["status"] = "changes_requested"
    await db.vendor_onboarding.update_one({"_id": o["_id"]}, {"$set": set_doc})
    await audit_log(user, "vendor_onboarding", onb_id, f"site_{data.decision}", {"remarks": data.remarks})
    return {"message": f"Site review: {data.decision}", "status": set_doc["status"]}

@api_router.post("/onboarding/vendors/{onb_id}/master-decision")
async def master_decision(onb_id: str, data: OnboardingDecision, user: dict = Depends(get_current_user)):
    """Master admin final approval/rejection."""
    if not is_master_admin(user):
        raise HTTPException(status_code=403, detail="Only master admin")
    o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
    if not o:
        raise HTTPException(status_code=404, detail="Onboarding not found")
    if o.get("status") != "under_master_review":
        raise HTTPException(status_code=400, detail=f"Cannot finalize — current status is '{o.get('status')}', must be 'under_master_review'")
    if data.decision not in ("approve", "reject"):
        raise HTTPException(status_code=400, detail="decision must be approve|reject")
    remark = {
        "stage": "master_review",
        "by": user["email"],
        "decision": data.decision,
        "remarks": data.remarks or "",
        "at": datetime.now(timezone.utc).isoformat(),
    }
    new_remarks = o.get("remarks", []) + [remark]
    set_doc = {"remarks": new_remarks, "updated_at": datetime.now(timezone.utc),
               "master_reviewed_by": user["id"], "master_reviewed_at": datetime.now(timezone.utc)}
    if data.decision == "approve":
        # Create real Vendor business record + map to site
        vendor_doc = {
            "name": o.get("vendor_name"),
            "description": o.get("company_name", ""),
            "cuisine_type": o.get("cuisine_type", "Multi-cuisine"),
            "phone": o.get("mobile_number"),
            "email": o.get("email"),
            "address": o.get("business_address"),
            "rating": 0.0,
            "status": "active",
            "commission_pct": 15.0,
            "onboarding_id": str(o["_id"]),
            "created_at": datetime.now(timezone.utc),
        }
        vres = await db.vendors.insert_one(vendor_doc)
        vendor_id = str(vres.inserted_id)
        # Site-vendor mapping (uses canonical collection name `vendor_site_mappings`)
        await db.vendor_site_mappings.insert_one({
            "site_id": o.get("site_id"),
            "vendor_id": vendor_id,
            "status": "active",
            "created_at": datetime.now(timezone.utc),
        })
        set_doc["status"] = "active"
        set_doc["vendor_id"] = vendor_id
    else:
        set_doc["status"] = "rejected"
    await db.vendor_onboarding.update_one({"_id": o["_id"]}, {"$set": set_doc})
    await audit_log(user, "vendor_onboarding", onb_id, f"master_{data.decision}", {"remarks": data.remarks, "vendor_id": set_doc.get("vendor_id")})
    return {"message": f"Master decision: {data.decision}", "status": set_doc["status"], "vendor_id": set_doc.get("vendor_id")}

@api_router.get("/onboarding/vendors/{onb_id}/audit-trail")
async def onboarding_audit_trail(onb_id: str, user: dict = Depends(get_current_user)):
    o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
    if not o:
        raise HTTPException(status_code=404, detail="Onboarding not found")
    if user["role"] not in ("master_admin", "city_admin", "site_admin"):
        raise HTTPException(status_code=403, detail="Access denied")
    if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
        raise HTTPException(status_code=403, detail="Not your site")
    cursor = db.audit_log.find({"entity_type": "vendor_onboarding", "entity_id": onb_id}).sort("created_at", 1)
    log = []
    async for entry in cursor:
        log.append({
            "user_email": entry.get("user_email"),
            "user_role": entry.get("user_role"),
            "action": entry.get("action"),
            "details": entry.get("details", {}),
            "created_at": entry.get("created_at").isoformat() if entry.get("created_at") else None,
        })
    return {"audit_trail": log}

@api_router.get("/onboarding/dashboard")
async def onboarding_dashboard(user: dict = Depends(get_current_user)):
    """Dashboard stats based on role."""
    filt = {}
    if user["role"] == "master_admin":
        pass
    elif user["role"] == "city_admin":
        filt["city_id"] = user.get("city_id")
    elif user["role"] == "site_admin":
        filt["site_id"] = user.get("site_id")
    else:
        raise HTTPException(status_code=403, detail="Access denied")
    pipe = [
        {"$match": filt} if filt else {"$match": {}},
        {"$group": {"_id": "$status", "count": {"$sum": 1}}},
    ]
    by_status = {row["_id"]: row["count"] async for row in db.vendor_onboarding.aggregate(pipe)}
    total = sum(by_status.values())
    pending = by_status.get("documents_pending", 0) + by_status.get("under_site_review", 0) + by_status.get("under_master_review", 0)
    # Average checklist pct of in-progress onboardings
    cursor = db.vendor_onboarding.find({**filt, "status": {"$nin": ["approved", "active", "rejected"]}})
    pcts = []
    async for o in cursor:
        pcts.append(calc_checklist_pct(o.get("checklist", {})))
    avg_pct = round(sum(pcts) / len(pcts), 1) if pcts else 0.0
    return {
        "total": total,
        "by_status": by_status,
        "pending_approvals": pending,
        "approved": by_status.get("approved", 0) + by_status.get("active", 0),
        "rejected": by_status.get("rejected", 0),
        "avg_checklist_pct": avg_pct,
    }


# ============== CITY PERFORMANCE LEADERBOARD ==============

@api_router.get("/reports/city-leaderboard")
async def city_leaderboard(days: int = 30, user: dict = Depends(get_current_user)):
    """Ranked list of cities by revenue, orders, vendor count, avg checklist."""
    if not is_master_admin(user):
        raise HTTPException(status_code=403, detail="Only master admin")
    days = max(1, min(days, 365))
    since = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=days - 1)

    cities_cursor = db.cities.find({})
    rows = []
    async for c in cities_cursor:
        cid = str(c["_id"])
        site_ids = [str(s["_id"]) async for s in db.sites.find({"city_id": cid}, {"_id": 1})]
        vendor_count = await db.vendor_onboarding.count_documents({"city_id": cid, "status": "active"})
        pending = await db.vendor_onboarding.count_documents({
            "city_id": cid,
            "status": {"$in": ["documents_pending", "under_site_review", "under_master_review", "changes_requested"]}
        })
        in_progress_cursor = db.vendor_onboarding.find({"city_id": cid, "status": {"$nin": ["approved", "active", "rejected"]}})
        pcts = []
        async for o in in_progress_cursor:
            pcts.append(calc_checklist_pct(o.get("checklist", {})))
        avg_pct = round(sum(pcts) / len(pcts), 1) if pcts else 0.0

        revenue, orders = 0.0, 0
        if site_ids:
            user_ids = [str(u["_id"]) async for u in db.users.find({"site_id": {"$in": site_ids}, "role": "employee"}, {"_id": 1})]
            if user_ids:
                pipe = [
                    {"$match": {"user_id": {"$in": user_ids}, "payment_status": "paid", "created_at": {"$gte": since}}},
                    {"$group": {"_id": None, "revenue": {"$sum": "$total_amount"}, "orders": {"$sum": 1}}}
                ]
                async for r in db.orders.aggregate(pipe):
                    revenue = round(r.get("revenue", 0), 2)
                    orders = r.get("orders", 0)

        rows.append({
            "city_id": cid,
            "name": c.get("name", "Unknown"),
            "state": c.get("state", ""),
            "site_count": len(site_ids),
            "vendor_count": vendor_count,
            "pending_onboardings": pending,
            "avg_checklist_pct": avg_pct,
            "orders": orders,
            "revenue": revenue,
        })

    rows.sort(key=lambda r: r["revenue"], reverse=True)
    return {"days": days, "cities": rows, "total_revenue": round(sum(r["revenue"] for r in rows), 2)}


# ============== MASTER ADMIN: ANALYTICS CHARTS ==============

@api_router.get("/reports/charts")
async def master_charts(days: int = 14, user: dict = Depends(get_current_user)):
    """Time-series for charts on master dashboard."""
    if not is_master_admin(user) and user.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="Access denied")
    days = max(7, min(days, 90))
    since = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=days - 1)

    # Daily revenue
    rev_pipe = [
        {"$match": {"payment_status": "paid", "created_at": {"$gte": since}}},
        {"$group": {
            "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$created_at"}},
            "revenue": {"$sum": "$total_amount"},
            "orders": {"$sum": 1},
        }},
        {"$sort": {"_id": 1}},
    ]
    daily = await db.orders.aggregate(rev_pipe).to_list(100)
    daily_revenue = [{"date": d["_id"], "revenue": round(d["revenue"], 2), "orders": d["orders"]} for d in daily]

    # Top dishes by quantity (last N days)
    items_pipe = [
        {"$match": {"payment_status": "paid", "created_at": {"$gte": since}}},
        {"$unwind": "$items"},
        {"$group": {
            "_id": "$items.menu_item_id",
            "qty": {"$sum": "$items.quantity"},
            "revenue": {"$sum": {"$multiply": ["$items.quantity", "$items.price"]}},
        }},
        {"$sort": {"qty": -1}},
        {"$limit": 5},
    ]
    top_items_raw = await db.orders.aggregate(items_pipe).to_list(5)
    top_dishes = []
    for it in top_items_raw:
        mid = it.get("_id")
        name = "Unknown"
        if mid and ObjectId.is_valid(mid):
            mi = await db.menu_items.find_one({"_id": ObjectId(mid)})
            if mi:
                name = mi.get("name", "Unknown")
        top_dishes.append({"menu_item_id": mid, "name": name, "qty": it["qty"], "revenue": round(it["revenue"], 2)})

    return {"days": days, "daily_revenue": daily_revenue, "top_dishes": top_dishes}


# ============== BULK EMPLOYEE CSV UPLOAD ==============

@api_router.post("/admin/employees/bulk-csv")
async def bulk_employee_csv(
    file: UploadFile = File(...),
    company_id: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """Corporate admin or master uploads CSV with columns: email,name,password,phone (optional)."""
    if user["role"] not in ("corporate_admin", "master_admin"):
        raise HTTPException(status_code=403, detail="Only corporate or master admin")
    if not (file.filename or "").lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="File must be a .csv")
    content = (await file.read()).decode("utf-8", errors="ignore")
    if len(content) > 1 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="CSV must be under 1 MB")

    cid = company_id or user.get("company_id")
    if not cid:
        raise HTTPException(status_code=400, detail="company_id required")

    import csv, io
    reader = csv.DictReader(io.StringIO(content))
    inserted, errors = 0, []
    for idx, row in enumerate(reader, start=2):
        email = (row.get("email") or "").strip().lower()
        name = (row.get("name") or "").strip()
        pwd = (row.get("password") or "").strip()
        if not email or not name or len(pwd) < 6:
            errors.append({"row": idx, "error": "missing email/name or password<6"})
            continue
        if await db.users.find_one({"email": email}):
            errors.append({"row": idx, "error": f"email {email} exists"})
            continue
        try:
            await db.users.insert_one({
                "email": email,
                "name": name,
                "password_hash": hash_password(pwd),
                "role": "employee",
                "company_id": cid,
                "site_id": (row.get("site_id") or "").strip() or None,
                "phone": (row.get("phone") or "").strip() or None,
                "preferences": {"vegetarian": False, "vegan": False, "gluten_free": False, "dairy_free": False, "nut_free": False, "spicy_preference": "medium", "allergies": [], "preferred_cuisines": []},
                "created_at": datetime.now(timezone.utc),
                "failed_attempts": 0,
            })
            inserted += 1
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})

    return {"inserted": inserted, "errors": errors, "total_attempted": inserted + len(errors)}


# ============== SITES & MULTI-LEVEL ADMIN ==============

def is_master_admin(user: dict) -> bool:
    return user.get("role") == "master_admin"

def is_master_or_super(user: dict) -> bool:
    return user.get("role") in ("master_admin", "super_admin")

def can_access_site(user: dict, site_id: str) -> bool:
    role = user.get("role")
    if role == "master_admin":
        return True
    if role == "super_admin":
        return site_id in (user.get("assigned_sites") or [])
    if role == "site_admin":
        return user.get("site_id") == site_id
    return False

# Sites CRUD (Master Admin)
@api_router.post("/sites")
async def create_site(data: SiteCreate, user: dict = Depends(get_current_user)):
    if not is_master_admin(user):
        raise HTTPException(status_code=403, detail="Only master admin can create sites")
    doc = {
        **data.model_dump(),
        "status": "active",
        "created_at": datetime.now(timezone.utc),
    }
    result = await db.sites.insert_one(doc)
    site_id = str(result.inserted_id)
    # Default meal schedule
    await db.meal_schedules.insert_one({
        "site_id": site_id,
        "schedules": [
            {"meal_period": "breakfast", "start_time": "07:30", "end_time": "10:30", "enabled": True},
            {"meal_period": "lunch", "start_time": "12:00", "end_time": "15:00", "enabled": True},
            {"meal_period": "snacks", "start_time": "16:00", "end_time": "18:00", "enabled": True},
            {"meal_period": "dinner", "start_time": "19:00", "end_time": "22:00", "enabled": False},
        ],
        "updated_at": datetime.now(timezone.utc),
    })
    return {"id": site_id, **data.model_dump()}

@api_router.get("/sites")
async def list_sites(user: dict = Depends(get_current_user)):
    query = {}
    if user.get("role") == "super_admin":
        ids = [safe_objectid(s, "Site") for s in (user.get("assigned_sites") or [])]
        if not ids:
            return []
        query["_id"] = {"$in": ids}
    elif user.get("role") == "site_admin":
        sid = user.get("site_id")
        if not sid:
            return []
        query["_id"] = safe_objectid(sid, "Site")
    elif user.get("role") == "employee":
        sid = user.get("site_id")
        if not sid:
            return []
        query["_id"] = safe_objectid(sid, "Site")
    elif user.get("role") == "vendor":
        # Vendor sees sites they're mapped to
        mappings = await db.vendor_site_mappings.find({"vendor_id": user.get("vendor_id")}).to_list(500)
        site_ids = [safe_objectid(m["site_id"], "Site") for m in mappings]
        if not site_ids:
            return []
        query["_id"] = {"$in": site_ids}
    # master_admin sees all sites
    
    sites = await db.sites.find(query).sort("name", 1).to_list(1000)
    for s in sites:
        s["id"] = str(s.pop("_id"))
        if isinstance(s.get("created_at"), datetime):
            s["created_at"] = s["created_at"].isoformat()
    return sites

@api_router.get("/sites/{site_id}")
async def get_site(site_id: str, user: dict = Depends(get_current_user)):
    if not (is_master_admin(user) or can_access_site(user, site_id) or
            user.get("role") == "employee" and user.get("site_id") == site_id):
        raise HTTPException(status_code=403, detail="Access denied")
    site = await db.sites.find_one({"_id": safe_objectid(site_id, "Site")})
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    site["id"] = str(site.pop("_id"))
    if isinstance(site.get("created_at"), datetime):
        site["created_at"] = site["created_at"].isoformat()
    return site

@api_router.patch("/sites/{site_id}")
async def update_site(site_id: str, updates: Dict[str, Any], user: dict = Depends(get_current_user)):
    if not can_access_site(user, site_id):
        raise HTTPException(status_code=403, detail="Access denied")
    allowed = {"name", "address", "city", "city_id", "contact_email", "contact_phone",
               "allow_pre_order", "allow_cash_carry", "allow_company_paid", "allow_employee_paid"}
    # `status` field can only be changed by master_admin
    if is_master_admin(user):
        allowed = allowed | {"status"}
    cleaned = {k: v for k, v in updates.items() if k in allowed}
    if not cleaned:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    await db.sites.update_one({"_id": safe_objectid(site_id, "Site")}, {"$set": cleaned})
    return {"message": "Site updated"}

# Vendor-Site Mapping (Master/Super Admin)
@api_router.post("/sites/{site_id}/vendors")
async def map_vendor_to_site(site_id: str, data: VendorSiteMappingCreate, user: dict = Depends(get_current_user)):
    if not can_access_site(user, site_id):
        raise HTTPException(status_code=403, detail="Access denied")
    existing = await db.vendor_site_mappings.find_one({"vendor_id": data.vendor_id, "site_id": site_id})
    if existing:
        raise HTTPException(status_code=400, detail="Vendor already mapped to this site")
    await db.vendor_site_mappings.insert_one({
        "vendor_id": data.vendor_id,
        "site_id": site_id,
        "status": "active",
        "created_at": datetime.now(timezone.utc),
    })
    return {"message": "Vendor mapped to site"}

@api_router.get("/sites/{site_id}/vendors")
async def list_site_vendors(site_id: str, user: dict = Depends(get_current_user)):
    # Allow employees of this site to list vendors too
    if not (can_access_site(user, site_id) or
            (user.get("role") == "employee" and user.get("site_id") == site_id)):
        raise HTTPException(status_code=403, detail="Access denied")
    mappings = await db.vendor_site_mappings.find({"site_id": site_id, "status": "active"}).to_list(500)
    vendor_ids = [safe_objectid(m["vendor_id"], "Vendor") for m in mappings]
    if not vendor_ids:
        return []
    vendors = await db.vendors.find({"_id": {"$in": vendor_ids}, "status": "active"}).to_list(500)
    for v in vendors:
        v["id"] = str(v.pop("_id"))
    return vendors

@api_router.delete("/sites/{site_id}/vendors/{vendor_id}")
async def unmap_vendor(site_id: str, vendor_id: str, user: dict = Depends(get_current_user)):
    if not can_access_site(user, site_id):
        raise HTTPException(status_code=403, detail="Access denied")
    await db.vendor_site_mappings.delete_one({"vendor_id": vendor_id, "site_id": site_id})
    return {"message": "Vendor unmapped"}

# Meal Schedules per Site
@api_router.get("/sites/{site_id}/schedule")
async def get_site_schedule(site_id: str, user: dict = Depends(get_current_user)):
    if not (can_access_site(user, site_id) or
            (user.get("role") == "employee" and user.get("site_id") == site_id)):
        raise HTTPException(status_code=403, detail="Access denied")
    sched = await db.meal_schedules.find_one({"site_id": site_id})
    if not sched:
        return {"site_id": site_id, "schedules": []}
    return {
        "site_id": site_id,
        "schedules": sched.get("schedules", []),
    }

@api_router.put("/sites/{site_id}/schedule")
async def update_site_schedule(site_id: str, data: MealScheduleUpdate, user: dict = Depends(get_current_user)):
    if not can_access_site(user, site_id):
        raise HTTPException(status_code=403, detail="Access denied")
    await db.meal_schedules.update_one(
        {"site_id": site_id},
        {"$set": {"schedules": [s.model_dump() for s in data.schedules], "updated_at": datetime.now(timezone.utc)}},
        upsert=True,
    )
    return {"message": "Schedule updated"}

# Current meal period helper
def current_meal_period(schedules: list) -> Optional[str]:
    from datetime import time as dt_time
    now = datetime.now(timezone.utc).astimezone()
    current = now.strftime("%H:%M")
    for s in schedules:
        if s.get("enabled") and s.get("start_time") <= current <= s.get("end_time"):
            return s["meal_period"]
    return None

# Site Menu (Site Admin / Employee dynamic)
@api_router.get("/sites/{site_id}/menu")
async def get_site_menu(
    site_id: str,
    meal_period: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    # Allow employees of this site
    if not (can_access_site(user, site_id) or
            (user.get("role") == "employee" and user.get("site_id") == site_id) or
            user.get("role") == "vendor"):
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {"site_id": site_id}
    if meal_period:
        query["meal_periods"] = meal_period
    # Site admin / Master sees all; Employee sees only available
    if user.get("role") == "employee":
        query["is_available"] = True
    
    items = await db.menu_items.find(query).to_list(2000)
    for item in items:
        item["id"] = str(item.pop("_id"))
        if isinstance(item.get("created_at"), datetime):
            item["created_at"] = item["created_at"].isoformat()
    return items

@api_router.patch("/menu/{item_id}/site-control")
async def site_admin_menu_control(item_id: str, data: MenuItemSiteUpdate, user: dict = Depends(get_current_user)):
    """Site admin (or master/super) toggles availability, pricing, show_price, or meal_periods on a menu item."""
    item = await db.menu_items.find_one({"_id": safe_objectid(item_id, "Menu item")})
    if not item:
        raise HTTPException(status_code=404, detail="Menu item not found")
    if not (is_master_admin(user) or can_access_site(user, item.get("site_id", ""))):
        raise HTTPException(status_code=403, detail="Access denied")
    cleaned = {k: v for k, v in data.model_dump().items() if v is not None}
    if not cleaned:
        raise HTTPException(status_code=400, detail="No fields to update")
    await db.menu_items.update_one({"_id": safe_objectid(item_id, "Menu item")}, {"$set": cleaned})
    return {"message": "Menu item updated"}

# Excel Menu Upload (Site Admin)
@api_router.post("/sites/{site_id}/menu/upload-excel")
async def upload_menu_excel(
    site_id: str,
    vendor_id: str = Query(...),
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    """Upload an Excel (.xlsx) file with menu items.
    Expected columns: name, description, category, price, is_vegetarian, image_url (optional), meal_periods (comma-separated)
    """
    if not (is_master_admin(user) or can_access_site(user, site_id)):
        raise HTTPException(status_code=403, detail="Access denied")
    
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xls files are supported")
    
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 5 MB)")
    
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")
    
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel must contain at least a header row and one data row")
    
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    required = ["name", "description", "category", "price"]
    missing = [c for c in required if c not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}")
    
    name_idx = headers.index("name")
    desc_idx = headers.index("description")
    cat_idx = headers.index("category")
    price_idx = headers.index("price")
    veg_idx = headers.index("is_vegetarian") if "is_vegetarian" in headers else None
    img_idx = headers.index("image_url") if "image_url" in headers else None
    meal_idx = headers.index("meal_periods") if "meal_periods" in headers else None
    
    inserted = 0
    errors = []
    for i, row in enumerate(rows[1:], start=2):
        try:
            if not row[name_idx]:
                continue
            meal_periods = ["lunch"]
            if meal_idx is not None and row[meal_idx]:
                meal_periods = [m.strip().lower() for m in str(row[meal_idx]).split(",") if m.strip()]
            doc = {
                "vendor_id": vendor_id,
                "site_id": site_id,
                "name": str(row[name_idx]).strip(),
                "description": str(row[desc_idx] or "").strip(),
                "category": str(row[cat_idx] or "Main Course").strip(),
                "price": float(row[price_idx] or 0),
                "is_vegetarian": bool(row[veg_idx]) if veg_idx is not None else True,
                "is_available": True,
                "show_price": True,
                "meal_periods": meal_periods,
                "image_url": str(row[img_idx]).strip() if img_idx is not None and row[img_idx] else None,
                "created_at": datetime.now(timezone.utc),
            }
            await db.menu_items.insert_one(doc)
            inserted += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")
    
    return {"inserted": inserted, "errors": errors, "site_id": site_id, "vendor_id": vendor_id}

# Master Admin: Create Site Admin / Super Admin
@api_router.post("/admin/site-admins")
async def create_site_admin(data: SiteAdminCreate, user: dict = Depends(get_current_user)):
    if not is_master_admin(user):
        raise HTTPException(status_code=403, detail="Only master admin can create site admins")
    email_lower = data.email.lower()
    if await db.users.find_one({"email": email_lower}):
        raise HTTPException(status_code=400, detail="Email already registered")
    # Validate site exists
    site = await db.sites.find_one({"_id": safe_objectid(data.site_id, "Site")})
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    result = await db.users.insert_one({
        "email": email_lower,
        "password_hash": hash_password(data.password),
        "name": data.name,
        "role": "site_admin",
        "site_id": data.site_id,
        "created_at": datetime.now(timezone.utc),
    })
    return {"id": str(result.inserted_id), "email": email_lower, "role": "site_admin", "site_id": data.site_id}

@api_router.post("/admin/super-admins")
async def create_super_admin(data: SuperAdminCreate, user: dict = Depends(get_current_user)):
    if not is_master_admin(user):
        raise HTTPException(status_code=403, detail="Only master admin can create super admins")
    email_lower = data.email.lower()
    if await db.users.find_one({"email": email_lower}):
        raise HTTPException(status_code=400, detail="Email already registered")
    # Validate assigned_sites exist
    if data.assigned_sites:
        site_oids = [safe_objectid(sid, "Site") for sid in data.assigned_sites]
        existing = await db.sites.count_documents({"_id": {"$in": site_oids}})
        if existing != len(data.assigned_sites):
            raise HTTPException(status_code=404, detail="One or more assigned_sites not found")
    result = await db.users.insert_one({
        "email": email_lower,
        "password_hash": hash_password(data.password),
        "name": data.name,
        "role": "super_admin",
        "assigned_sites": data.assigned_sites,
        "created_at": datetime.now(timezone.utc),
    })
    return {"id": str(result.inserted_id), "email": email_lower, "role": "super_admin", "assigned_sites": data.assigned_sites}

@api_router.post("/admin/master-admins")
async def create_master_admin(data: MasterAdminCreate, user: dict = Depends(get_current_user)):
    if not is_master_admin(user):
        raise HTTPException(status_code=403, detail="Only master admin can create master admins")
    email_lower = data.email.lower()
    if not email_lower.endswith("@cravitoo.com"):
        raise HTTPException(status_code=400, detail="Master admin email must be @cravitoo.com")
    if await db.users.find_one({"email": email_lower}):
        raise HTTPException(status_code=400, detail="Email already registered")
    result = await db.users.insert_one({
        "email": email_lower,
        "password_hash": hash_password(data.password),
        "name": data.name,
        "role": "master_admin",
        "created_at": datetime.now(timezone.utc),
    })
    return {"id": str(result.inserted_id), "email": email_lower, "role": "master_admin"}

@api_router.get("/admin/admins")
async def list_admins(user: dict = Depends(get_current_user)):
    if not is_master_admin(user):
        raise HTTPException(status_code=403, detail="Only master admin can list admins")
    admins = await db.users.find(
        {"role": {"$in": ["master_admin", "super_admin", "site_admin", "city_admin"]}},
        {"_id": 1, "email": 1, "name": 1, "role": 1, "site_id": 1, "city_id": 1, "assigned_sites": 1, "created_at": 1}
    ).to_list(1000)
    for a in admins:
        a["id"] = str(a.pop("_id"))
        if isinstance(a.get("created_at"), datetime):
            a["created_at"] = a["created_at"].isoformat()
    return admins

@api_router.delete("/admin/admins/{admin_id}")
async def delete_admin(admin_id: str, user: dict = Depends(get_current_user)):
    if not is_master_admin(user):
        raise HTTPException(status_code=403, detail="Only master admin can delete admins")
    admin = await db.users.find_one({"_id": safe_objectid(admin_id, "Admin"), "role": {"$in": ["super_admin", "site_admin", "master_admin", "city_admin"]}})
    if not admin:
        raise HTTPException(status_code=404, detail="Admin not found")
    if admin.get("email") == os.environ.get("ADMIN_EMAIL", "admin@cravitoo.com"):
        raise HTTPException(status_code=400, detail="Cannot delete the seed master admin")
    await db.users.delete_one({"_id": safe_objectid(admin_id, "Admin")})
    return {"message": "Admin deleted"}

# Reports
@api_router.get("/reports/master-dashboard")
async def master_dashboard(user: dict = Depends(get_current_user)):
    if not is_master_admin(user):
        raise HTTPException(status_code=403, detail="Master admin only")
    total_sites = await db.sites.count_documents({"status": "active"})
    total_vendors = await db.vendors.count_documents({"status": "active"})
    total_users = await db.users.count_documents({})
    total_employees = await db.users.count_documents({"role": "employee"})
    total_orders = await db.orders.count_documents({})
    paid_orders = await db.orders.count_documents({"payment_status": "paid"})
    
    rev_pipe = [
        {"$match": {"payment_status": "paid"}},
        {"$group": {"_id": None, "total": {"$sum": "$total_amount"}}}
    ]
    rev = await db.orders.aggregate(rev_pipe).to_list(1)
    total_revenue = rev[0]["total"] if rev else 0
    
    # Top sites
    site_pipe = [
        {"$match": {"payment_status": "paid", "site_id": {"$exists": True, "$ne": None}}},
        {"$group": {"_id": "$site_id", "orders": {"$sum": 1}, "revenue": {"$sum": "$total_amount"}}},
        {"$sort": {"revenue": -1}},
        {"$limit": 5}
    ]
    top_sites_raw = await db.orders.aggregate(site_pipe).to_list(5)
    top_sites = []
    for ts in top_sites_raw:
        if not ts.get("_id") or not ObjectId.is_valid(ts["_id"]):
            continue
        site = await db.sites.find_one({"_id": ObjectId(ts["_id"])})
        if site:
            top_sites.append({
                "site_id": ts["_id"],
                "name": site.get("name", "Unknown"),
                "orders": ts["orders"],
                "revenue": ts["revenue"],
            })
    
    # Top vendors
    vendor_pipe = [
        {"$match": {"payment_status": "paid"}},
        {"$group": {"_id": "$vendor_id", "orders": {"$sum": 1}, "revenue": {"$sum": "$total_amount"}}},
        {"$sort": {"revenue": -1}},
        {"$limit": 5}
    ]
    top_vendors_raw = await db.orders.aggregate(vendor_pipe).to_list(5)
    top_vendors = []
    for tv in top_vendors_raw:
        if not tv.get("_id") or not ObjectId.is_valid(tv["_id"]):
            continue
        vendor = await db.vendors.find_one({"_id": ObjectId(tv["_id"])})
        if vendor:
            top_vendors.append({
                "vendor_id": tv["_id"],
                "name": vendor.get("name", "Unknown"),
                "orders": tv["orders"],
                "revenue": tv["revenue"],
            })
    
    return {
        "total_sites": total_sites,
        "total_vendors": total_vendors,
        "total_users": total_users,
        "total_employees": total_employees,
        "total_orders": total_orders,
        "paid_orders": paid_orders,
        "total_revenue": round(total_revenue, 2),
        "top_sites": top_sites,
        "top_vendors": top_vendors,
    }

@api_router.get("/reports/site/{site_id}")
async def site_report(site_id: str, user: dict = Depends(get_current_user)):
    if not can_access_site(user, site_id):
        raise HTTPException(status_code=403, detail="Access denied")
    total_orders = await db.orders.count_documents({"site_id": site_id})
    paid_orders = await db.orders.count_documents({"site_id": site_id, "payment_status": "paid"})
    rev_pipe = [
        {"$match": {"site_id": site_id, "payment_status": "paid"}},
        {"$group": {"_id": None, "total": {"$sum": "$total_amount"}}}
    ]
    rev = await db.orders.aggregate(rev_pipe).to_list(1)
    total_revenue = rev[0]["total"] if rev else 0
    
    # By vendor
    by_vendor_pipe = [
        {"$match": {"site_id": site_id, "payment_status": "paid"}},
        {"$group": {"_id": "$vendor_id", "orders": {"$sum": 1}, "revenue": {"$sum": "$total_amount"}}},
        {"$sort": {"revenue": -1}}
    ]
    by_vendor_raw = await db.orders.aggregate(by_vendor_pipe).to_list(100)
    by_vendor = []
    for bv in by_vendor_raw:
        vendor = await db.vendors.find_one({"_id": safe_objectid(bv["_id"], "Vendor")})
        if vendor:
            by_vendor.append({
                "vendor_id": bv["_id"],
                "name": vendor.get("name", "Unknown"),
                "orders": bv["orders"],
                "revenue": round(bv["revenue"], 2),
            })
    
    employees_at_site = await db.users.count_documents({"site_id": site_id, "role": "employee"})
    
    return {
        "site_id": site_id,
        "total_orders": total_orders,
        "paid_orders": paid_orders,
        "total_revenue": round(total_revenue, 2),
        "employees": employees_at_site,
        "vendors": by_vendor,
    }

# Add site_id to order creation
@api_router.get("/employee/my-site")
async def get_my_site(user: dict = Depends(get_current_user)):
    """Helper for employee app: returns the employee's site + vendors + meal schedule + ordering options."""
    if user.get("role") != "employee":
        raise HTTPException(status_code=403, detail="Employee only")
    site_id = user.get("site_id")
    if not site_id:
        raise HTTPException(status_code=404, detail="No site assigned to your account")
    site = await db.sites.find_one({"_id": safe_objectid(site_id, "Site")})
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    site["id"] = str(site.pop("_id"))
    if isinstance(site.get("created_at"), datetime):
        site["created_at"] = site["created_at"].isoformat()
    schedule = await db.meal_schedules.find_one({"site_id": site_id})
    schedules = schedule.get("schedules", []) if schedule else []
    current_period = current_meal_period(schedules)
    
    mappings = await db.vendor_site_mappings.find({"site_id": site_id, "status": "active"}).to_list(500)
    vendor_ids = [safe_objectid(m["vendor_id"], "Vendor") for m in mappings]
    vendors = []
    if vendor_ids:
        vlist = await db.vendors.find({"_id": {"$in": vendor_ids}, "status": "active"}).to_list(500)
        for v in vlist:
            v["id"] = str(v.pop("_id"))
            vendors.append(v)
    
    return {
        "site": site,
        "vendors": vendors,
        "meal_schedule": schedules,
        "current_meal_period": current_period,
        "ordering_modes": {
            "pre_order": site.get("allow_pre_order", True),
            "cash_carry": site.get("allow_cash_carry", True),
            "company_paid": site.get("allow_company_paid", False),
            "employee_paid": site.get("allow_employee_paid", True),
        },
    }

# ============== WEBSOCKETS ==============

@app.websocket("/ws/orders")
async def websocket_orders(websocket: WebSocket, token: str = Query(...)):
    """Employee WebSocket: subscribes to their own order updates."""
    payload = verify_ws_token(token)
    if not payload:
        await websocket.close(code=1008, reason="Unauthorized")
        return
    user_id = payload["sub"]
    await manager.connect_user(user_id, websocket)
    try:
        # Send initial ping
        await websocket.send_json({"type": "connected", "user_id": user_id})
        while True:
            # Keep alive - clients can send ping, we echo pong
            msg = await websocket.receive_text()
            if msg == "ping":
                await websocket.send_text("pong")
    except WebSocketDisconnect:
        manager.disconnect_user(user_id, websocket)
    except Exception as e:
        logger.error(f"WS error for user {user_id}: {e}")
        manager.disconnect_user(user_id, websocket)

@app.websocket("/ws/vendor")
async def websocket_vendor(websocket: WebSocket, token: str = Query(...)):
    """Vendor WebSocket: subscribes to their vendor's order events."""
    payload = verify_ws_token(token)
    if not payload:
        await websocket.close(code=1008, reason="Unauthorized")
        return
    user_id = payload["sub"]
    # Lookup vendor_id
    user = await db.users.find_one({"_id": safe_objectid(user_id, "User")})
    if not user or user.get("role") != "vendor" or not user.get("vendor_id"):
        await websocket.close(code=1008, reason="Not a vendor")
        return
    vendor_id = user["vendor_id"]
    await manager.connect_vendor(vendor_id, websocket)
    try:
        await websocket.send_json({"type": "connected", "vendor_id": vendor_id})
        while True:
            msg = await websocket.receive_text()
            if msg == "ping":
                await websocket.send_text("pong")
    except WebSocketDisconnect:
        manager.disconnect_vendor(vendor_id, websocket)
    except Exception as e:
        logger.error(f"WS error for vendor {vendor_id}: {e}")
        manager.disconnect_vendor(vendor_id, websocket)

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()