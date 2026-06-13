"""Iter8 Cravitoo P1 backend tests.

Covers:
- PATCH /api/admin/vendors/{id} (master vendor edit)
- GET /api/refunds (employee)
- POST/GET/DELETE /api/favorites (employee)
- GET /api/orders/last (employee)
- POST /api/onboarding/vendors/bulk-import (master, xlsx)
- POST /api/onboarding/vendors/{id}/menu/upload-excel
- GET /api/meal-period/current (public)
- GET /api/reports/city-leaderboard (master)
- Low-stock alert integration via order placement
"""
import io
import os
import time
import pytest
import requests
import openpyxl

# Read REACT_APP_BACKEND_URL from frontend/.env
def _read_base():
    base = os.environ.get("REACT_APP_BACKEND_URL")
    if base:
        return base.rstrip("/")
    try:
        with open("/app/frontend/.env") as f:
            for line in f:
                if line.startswith("REACT_APP_BACKEND_URL="):
                    return line.split("=", 1)[1].strip().rstrip("/")
    except Exception:
        pass
    raise RuntimeError("REACT_APP_BACKEND_URL not set")

BASE = _read_base()

CREDS = {
    "master": {"email": "admin@cravitoo.com", "password": "admin123"},
    "site_admin": {"email": "siteadmin@techcorp.com", "password": "site123"},
    "vendor": {"email": "vendor@spicekitchen.com", "password": "vendor123"},
    "employee": {"email": "employee@techcorp.com", "password": "employee123"},
}


def _login(role):
    r = requests.post(f"{BASE}/api/auth/login", json=CREDS[role], timeout=20)
    assert r.status_code == 200, f"login {role}: {r.status_code} {r.text}"
    body = r.json()
    return body.get("access_token") or body.get("token"), body


@pytest.fixture(scope="session")
def tokens():
    out = {}
    for role in CREDS:
        tok, body = _login(role)
        out[role] = tok
        out[f"{role}_user"] = body
    return out


def _hdr(tok):
    return {"Authorization": f"Bearer {tok}"}


@pytest.fixture(scope="session")
def vendor_id(tokens):
    # Find Spice Kitchen vendor id via /api/vendors (master)
    r = requests.get(f"{BASE}/api/vendors", headers=_hdr(tokens["master"]), timeout=20)
    assert r.status_code == 200, r.text
    arr = r.json()
    for v in arr:
        if "spice" in (v.get("name") or "").lower():
            return v.get("id") or v.get("_id")
    # fall back to first vendor
    assert arr, "no vendors exist"
    return arr[0].get("id") or arr[0].get("_id")


# ===================== PATCH /api/admin/vendors/{id} =====================
class TestVendorEdit:
    def test_master_update_success(self, tokens, vendor_id):
        payload = {"description": "TEST_ desc " + str(int(time.time())), "phone": "+919999999999"}
        r = requests.patch(f"{BASE}/api/admin/vendors/{vendor_id}", json=payload, headers=_hdr(tokens["master"]))
        assert r.status_code == 200, r.text
        # Verify via GET
        g = requests.get(f"{BASE}/api/vendors", headers=_hdr(tokens["master"]))
        v = next((x for x in g.json() if (x.get("id") or x.get("_id")) == vendor_id), None)
        assert v, "vendor not in list"
        assert v.get("description") == payload["description"]

    def test_status_validation(self, tokens, vendor_id):
        r = requests.patch(f"{BASE}/api/admin/vendors/{vendor_id}", json={"status": "bogus"}, headers=_hdr(tokens["master"]))
        assert r.status_code == 400
        # valid status
        r2 = requests.patch(f"{BASE}/api/admin/vendors/{vendor_id}", json={"status": "active"}, headers=_hdr(tokens["master"]))
        assert r2.status_code == 200

    def test_commission_pct_validation(self, tokens, vendor_id):
        r = requests.patch(f"{BASE}/api/admin/vendors/{vendor_id}", json={"commission_pct": 75}, headers=_hdr(tokens["master"]))
        assert r.status_code == 400

    def test_non_master_forbidden(self, tokens, vendor_id):
        r = requests.patch(f"{BASE}/api/admin/vendors/{vendor_id}", json={"name": "hack"}, headers=_hdr(tokens["employee"]))
        assert r.status_code == 403
        r2 = requests.patch(f"{BASE}/api/admin/vendors/{vendor_id}", json={"name": "hack"}, headers=_hdr(tokens["site_admin"]))
        assert r2.status_code == 403


# ===================== /api/refunds =====================
class TestRefunds:
    def test_employee_list(self, tokens):
        r = requests.get(f"{BASE}/api/refunds", headers=_hdr(tokens["employee"]))
        assert r.status_code == 200, r.text
        data = r.json()
        assert isinstance(data, list)
        for o in data:
            assert "order_id" in o
            assert "refund_status" in o or o.get("status") == "cancelled"
            # cancelled_at and refunded_at fields present (may be None)
            assert "cancelled_at" in o
            assert "refunded_at" in o

    def test_non_employee_forbidden(self, tokens):
        for role in ("master", "site_admin", "vendor"):
            r = requests.get(f"{BASE}/api/refunds", headers=_hdr(tokens[role]))
            assert r.status_code == 403, f"{role} got {r.status_code}"


# ===================== /api/favorites =====================
class TestFavorites:
    def test_add_favorite(self, tokens, vendor_id):
        # Clean up first to ensure first-add path
        requests.delete(f"{BASE}/api/favorites/{vendor_id}", headers=_hdr(tokens["employee"]))
        r = requests.post(f"{BASE}/api/favorites/{vendor_id}", headers=_hdr(tokens["employee"]))
        assert r.status_code == 200, r.text
        assert "Added" in r.json().get("message", "")

    def test_add_favorite_twice(self, tokens, vendor_id):
        r = requests.post(f"{BASE}/api/favorites/{vendor_id}", headers=_hdr(tokens["employee"]))
        assert r.status_code == 200
        assert "Already" in r.json().get("message", "")

    def test_list_favorites(self, tokens, vendor_id):
        r = requests.get(f"{BASE}/api/favorites", headers=_hdr(tokens["employee"]))
        assert r.status_code == 200, r.text
        favs = r.json()
        assert isinstance(favs, list)
        match = next((f for f in favs if f.get("vendor_id") == vendor_id), None)
        assert match, "vendor not found in favorites"
        # validate expected fields
        assert "name" in match
        assert "cuisine_type" in match
        assert "rating" in match
        assert "favorited_at" in match

    def test_delete_favorite(self, tokens, vendor_id):
        r = requests.delete(f"{BASE}/api/favorites/{vendor_id}", headers=_hdr(tokens["employee"]))
        assert r.status_code == 200
        # verify removed
        g = requests.get(f"{BASE}/api/favorites", headers=_hdr(tokens["employee"])).json()
        assert not any(f.get("vendor_id") == vendor_id for f in g)

    def test_non_employee_forbidden(self, tokens, vendor_id):
        for role in ("master", "site_admin", "vendor"):
            r = requests.post(f"{BASE}/api/favorites/{vendor_id}", headers=_hdr(tokens[role]))
            assert r.status_code == 403
            r2 = requests.get(f"{BASE}/api/favorites", headers=_hdr(tokens[role]))
            assert r2.status_code == 403
            r3 = requests.delete(f"{BASE}/api/favorites/{vendor_id}", headers=_hdr(tokens[role]))
            assert r3.status_code == 403


# ===================== /api/orders/last =====================
class TestLastOrder:
    def test_employee_last(self, tokens):
        r = requests.get(f"{BASE}/api/orders/last", headers=_hdr(tokens["employee"]))
        # 200 if has orders, 404 if none. Accept both, validate structure if 200
        assert r.status_code in (200, 404), r.text
        if r.status_code == 200:
            data = r.json()
            assert "vendor_id" in data
            assert "items" in data
            assert "total_amount" in data
        else:
            assert "No previous orders" in r.json().get("detail", "")

    def test_non_employee_forbidden(self, tokens):
        for role in ("master", "site_admin", "vendor"):
            r = requests.get(f"{BASE}/api/orders/last", headers=_hdr(tokens[role]))
            assert r.status_code == 403


# ===================== Bulk Import =====================
def _make_onboarding_xlsx(include_bad_row=True):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["vendor_name", "company_name", "contact_person", "mobile_number", "email", "business_address", "cuisine_type"])
    nonce = int(time.time())
    ws.append([f"TEST__vendor_A_{nonce}", "TEST_co", "Alice", "9000000001", f"a_{nonce}@t.com", "Addr A", "Indian"])
    ws.append([f"TEST__vendor_B_{nonce}", "TEST_co", "Bob",   "9000000002", f"b_{nonce}@t.com", "Addr B", "Chinese"])
    ws.append([f"TEST__vendor_C_{nonce}", "TEST_co", "Carl",  "9000000003", f"c_{nonce}@t.com", "Addr C", "Italian"])
    if include_bad_row:
        ws.append(["",                            "TEST_co", "Dave",  "9000000004", f"d_{nonce}@t.com", "Addr D", "Multi"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _make_missing_col_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    # missing 'email' column
    ws.append(["vendor_name", "company_name", "contact_person", "mobile_number", "business_address", "cuisine_type"])
    ws.append(["X", "Y", "Z", "9", "Addr", "Indian"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.fixture(scope="session")
def first_site_id(tokens):
    r = requests.get(f"{BASE}/api/sites", headers=_hdr(tokens["master"]))
    assert r.status_code == 200
    arr = r.json()
    assert arr, "no sites"
    return arr[0].get("id") or arr[0].get("_id")


class TestBulkImport:
    def test_bulk_import_3valid_1bad(self, tokens, first_site_id):
        buf = _make_onboarding_xlsx(include_bad_row=True)
        files = {"file": ("upload.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(
            f"{BASE}/api/onboarding/vendors/bulk-import",
            params={"site_id": first_site_id},
            files=files,
            headers=_hdr(tokens["master"]),
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["inserted"] == 3, body
        assert body["total_attempted"] == 4
        assert isinstance(body["errors"], list) and len(body["errors"]) == 1
        err = body["errors"][0]
        assert err.get("row") == 5
        assert "vendor_name" in err.get("error", "")

    def test_non_xlsx_400(self, tokens, first_site_id):
        files = {"file": ("upload.txt", io.BytesIO(b"hello"), "text/plain")}
        r = requests.post(
            f"{BASE}/api/onboarding/vendors/bulk-import",
            params={"site_id": first_site_id},
            files=files,
            headers=_hdr(tokens["master"]),
        )
        assert r.status_code == 400
        assert "xlsx" in r.json().get("detail", "").lower()

    def test_missing_column_400(self, tokens, first_site_id):
        buf = _make_missing_col_xlsx()
        files = {"file": ("upload.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(
            f"{BASE}/api/onboarding/vendors/bulk-import",
            params={"site_id": first_site_id},
            files=files,
            headers=_hdr(tokens["master"]),
        )
        assert r.status_code == 400
        assert "missing" in r.json().get("detail", "").lower() or "email" in r.json().get("detail", "").lower()


# ===================== Onboarding Menu Excel =====================
def _make_menu_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["name", "category", "price", "description", "is_vegetarian"])
    ws.append(["TEST_Paneer Tikka", "Starters", 250, "Cottage cheese", "yes"])
    ws.append(["TEST_Veg Biryani", "Mains", 350, "Spiced rice", "veg"])
    ws.append(["TEST_Gulab Jamun", "Dessert", 90, "Sweet", "yes"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.fixture(scope="session")
def draft_onboarding_id(tokens, first_site_id):
    """Create one in-progress onboarding to upload menu against."""
    nonce = int(time.time())
    payload = {
        "vendor_name": f"TEST__menuvendor_{nonce}",
        "company_name": "TEST_co",
        "contact_person": "TestPerson",
        "mobile_number": "9112233445",
        "email": f"menu_{nonce}@t.com",
        "business_address": "Test addr",
        "cuisine_type": "Indian",
        "site_id": first_site_id,
    }
    r = requests.post(f"{BASE}/api/onboarding/vendors", json=payload, headers=_hdr(tokens["master"]))
    assert r.status_code in (200, 201), r.text
    body = r.json()
    return body.get("id") or body.get("onboarding_id") or body.get("_id")


class TestOnboardingMenuExcel:
    def test_upload_menu_success(self, tokens, draft_onboarding_id):
        assert draft_onboarding_id, "no draft onboarding id"
        buf = _make_menu_xlsx()
        files = {"file": ("menu.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(
            f"{BASE}/api/onboarding/vendors/{draft_onboarding_id}/menu/upload-excel",
            files=files, headers=_hdr(tokens["master"])
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["inserted"] == 3
        # Verify checklist menu_uploaded=true and draft_menu stored
        g = requests.get(f"{BASE}/api/onboarding/vendors/{draft_onboarding_id}", headers=_hdr(tokens["master"]))
        assert g.status_code == 200, g.text
        ob = g.json()
        cl = ob.get("checklist", {})
        assert cl.get("menu_uploaded") is True, cl
        # NOTE: draft_menu is persisted in DB but NOT exposed by onboarding_to_dict.
        # We assert the upload reported 3 inserted + checklist.menu_uploaded=True.
        dm = ob.get("draft_menu") or []
        if dm:
            assert len(dm) == 3

    def test_upload_menu_approved_400(self, tokens, first_site_id):
        # Create onboarding and force status=approved via direct admin update endpoint if possible.
        # Otherwise skip — we just confirm the route guards against approved/active/rejected.
        # Try: PATCH onboarding to status='approved' — likely not allowed via update endpoint.
        # As a softer check: pick an approved/active onboarding from the list if present.
        r = requests.get(f"{BASE}/api/onboarding/vendors", headers=_hdr(tokens["master"]))
        assert r.status_code == 200
        approved = next((o for o in r.json() if o.get("status") in ("approved", "active", "rejected")), None)
        if not approved:
            pytest.skip("No approved/active/rejected onboarding present to test 400 guard")
        oid = approved.get("id") or approved.get("_id")
        buf = _make_menu_xlsx()
        files = {"file": ("menu.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r2 = requests.post(
            f"{BASE}/api/onboarding/vendors/{oid}/menu/upload-excel",
            files=files, headers=_hdr(tokens["master"])
        )
        assert r2.status_code == 400, r2.text


# ===================== /api/meal-period/current =====================
class TestMealPeriod:
    def test_public_endpoint(self):
        r = requests.get(f"{BASE}/api/meal-period/current")
        assert r.status_code == 200, r.text
        data = r.json()
        assert "period" in data
        assert data["period"] in ("breakfast", "lunch", "snacks", "dinner", None)


# ===================== City Leaderboard =====================
class TestCityLeaderboard:
    def test_master_success(self, tokens):
        r = requests.get(f"{BASE}/api/reports/city-leaderboard?days=30", headers=_hdr(tokens["master"]))
        assert r.status_code == 200, r.text
        body = r.json()
        assert "cities" in body
        assert isinstance(body["cities"], list)
        if body["cities"]:
            c0 = body["cities"][0]
            for k in ("city_id", "name", "site_count", "vendor_count", "pending_onboardings", "avg_checklist_pct", "revenue"):
                assert k in c0, f"missing key {k} in {c0}"
            # sort by revenue desc
            revs = [c["revenue"] for c in body["cities"]]
            assert revs == sorted(revs, reverse=True)

    def test_non_master_forbidden(self, tokens):
        for role in ("site_admin", "vendor", "employee"):
            r = requests.get(f"{BASE}/api/reports/city-leaderboard", headers=_hdr(tokens[role]))
            assert r.status_code == 403


# ===================== Low-stock Integration =====================
class TestLowStockAlert:
    def test_low_stock_alert_via_orders(self, tokens):
        # 1) Vendor sets low_stock_threshold = 1
        r = requests.patch(
            f"{BASE}/api/vendor/settings",
            json={"low_stock_threshold": 1},
            headers=_hdr(tokens["vendor"]),
        )
        assert r.status_code == 200, r.text

        # 2) Find vendor_id and an available menu item for this vendor
        me_r = requests.get(f"{BASE}/api/auth/me", headers=_hdr(tokens["vendor"]))
        vendor_id = me_r.json().get("vendor_id")
        assert vendor_id, "vendor user missing vendor_id"

        items_r = requests.get(f"{BASE}/api/menu/{vendor_id}", headers=_hdr(tokens["employee"]))
        assert items_r.status_code == 200, items_r.text
        menu = items_r.json()

        # Find an item that hasn't been ordered today (so qty after 1 order == threshold)
        from datetime import datetime, timezone, timedelta
        today_iso = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)

        # Get vendor's orders today to compute qty per menu_item_id
        ord_r = requests.get(f"{BASE}/api/orders", headers=_hdr(tokens["vendor"]))
        if ord_r.status_code != 200:
            pytest.skip("vendor orders fetch failed")
        qty_today = {}
        for o in ord_r.json():
            ca = o.get("created_at")
            try:
                if isinstance(ca, str):
                    s = ca.replace("Z", "+00:00")
                    dt = datetime.fromisoformat(s)
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                else:
                    dt = None
            except Exception:
                dt = None
            if not dt or dt < today_iso:
                continue
            if o.get("status") == "cancelled":
                continue
            for it in (o.get("items") or []):
                mid = it.get("menu_item_id")
                qty_today[mid] = qty_today.get(mid, 0) + (it.get("quantity") or 0)

        # Pick an available item with qty_today == 0
        item = next((m for m in menu if m.get("is_available", True) and qty_today.get(m.get("id") or m.get("_id"), 0) == 0), None)
        if not item:
            pytest.skip("No fresh menu items (all items already ordered today) — low-stock condition (qty==threshold) cannot be reproduced")
        item_id = item.get("id") or item.get("_id")

        # 3) Note current count of vendor notifications
        n_before = requests.get(f"{BASE}/api/notifications", headers=_hdr(tokens["vendor"]))
        before_low_stock = 0
        if n_before.status_code == 200:
            before_low_stock = sum(1 for n in n_before.json() if "Low stock" in (n.get("title") or ""))

        # 4) Employee places 1 order with qty=1 (= threshold)
        order_payload = {
            "vendor_id": vendor_id,
            "items": [{"menu_item_id": item_id, "quantity": 1, "price": item.get("price", 0)}],
            "delivery_type": "pickup",
            "special_instructions": "TEST_ low stock",
        }
        o = requests.post(f"{BASE}/api/orders", json=order_payload, headers=_hdr(tokens["employee"]))
        assert o.status_code in (200, 201), o.text

        # small sleep for async notification create
        time.sleep(1.5)

        # 5) Verify low-stock notification exists for vendor
        n_after = requests.get(f"{BASE}/api/notifications", headers=_hdr(tokens["vendor"]))
        assert n_after.status_code == 200
        after_low_stock = sum(1 for n in n_after.json() if "Low stock" in (n.get("title") or ""))
        assert after_low_stock > before_low_stock, f"No new Low stock notification (before={before_low_stock} after={after_low_stock})"


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
