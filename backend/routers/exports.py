"""
Report exports — Excel / CSV / PDF (PDF Module 14).

Endpoints:
  GET /api/exports/reservations?from=YYYY-MM-DD&to=YYYY-MM-DD&format=xlsx|csv|pdf
  GET /api/exports/orders?from=YYYY-MM-DD&to=YYYY-MM-DD&format=xlsx|csv|pdf
  GET /api/exports/vendor-sales?from=YYYY-MM-DD&to=YYYY-MM-DD&format=xlsx|csv|pdf
  GET /api/exports/meal-summary?from=YYYY-MM-DD&to=YYYY-MM-DD&format=xlsx|csv|pdf

Access rules:
  - master_admin, super_admin: see all rows
  - corporate_admin: scoped to their company
  - vendor: scoped to their vendor_id (vendor-sales only)
  - site_admin: scoped to their site_id
"""
from __future__ import annotations

import csv
import io
import logging
from datetime import datetime, timezone, timedelta, date as date_cls
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

logger = logging.getLogger(__name__)

MEAL_TYPE_LABELS = {
    "veg_meal": "Veg Meal",
    "non_veg_meal": "Non-Veg Meal",
    "veg_salad": "Veg Salad",
    "non_veg_salad": "Non-Veg Salad",
}


def _parse_date(s: Optional[str], default_offset_days: int) -> datetime:
    if s:
        try:
            d = datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    else:
        d = (datetime.now(timezone.utc).date() + timedelta(days=default_offset_days))
    return datetime(d.year, d.month, d.day, tzinfo=timezone.utc)


def _to_iso(v: Any) -> str:
    if isinstance(v, datetime):
        return v.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    return str(v or "")


def _csv_response(rows: List[Dict[str, Any]], filename: str, columns: List[str]) -> StreamingResponse:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(columns)
    for row in rows:
        w.writerow([row.get(c, "") for c in columns])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
    )


def _xlsx_response(rows: List[Dict[str, Any]], filename: str, columns: List[str], sheet_name: str = "Report") -> StreamingResponse:
    from openpyxl import Workbook  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:30]

    header_fill = PatternFill(start_color="FF5A1F", end_color="FF5A1F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")

    for r, row in enumerate(rows, start=2):
        for i, col in enumerate(columns, start=1):
            ws.cell(row=r, column=i, value=row.get(col, ""))

    for i, col in enumerate(columns, start=1):
        max_len = max(
            len(str(col)),
            *(len(str(row.get(col, ""))) for row in rows[:200])
        )
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max(max_len + 2, 10), 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


def _pdf_response(
    rows: List[Dict[str, Any]],
    filename: str,
    columns: List[str],
    title: str,
    subtitle: str = "",
) -> StreamingResponse:
    from reportlab.lib.pagesizes import A4, landscape  # type: ignore
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer  # type: ignore
    from reportlab.lib import colors  # type: ignore
    from reportlab.lib.styles import getSampleStyleSheet  # type: ignore

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"<b>{title}</b>", styles["Title"]),
    ]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
        elements.append(Spacer(1, 8))
    elements.append(Spacer(1, 8))

    if not rows:
        elements.append(Paragraph("No data for this period.", styles["Normal"]))
    else:
        data = [columns] + [[str(row.get(c, "")) for c in columns] for row in rows]
        # Cap rows in PDF (keep under ~5 pages)
        if len(data) > 400:
            data = data[:401] + [["…", *(["" for _ in columns[1:]])]]
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF5A1F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFF7ED")]),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}.pdf"'},
    )


def _send(rows: List[Dict[str, Any]], columns: List[str], filename: str, fmt: str, title: str, subtitle: str) -> StreamingResponse:
    fmt = (fmt or "xlsx").lower()
    if fmt == "csv":
        return _csv_response(rows, filename, columns)
    if fmt == "xlsx":
        return _xlsx_response(rows, filename, columns, sheet_name=title[:30])
    if fmt == "pdf":
        return _pdf_response(rows, filename, columns, title, subtitle)
    raise HTTPException(status_code=400, detail="format must be one of: xlsx, csv, pdf")


def make_router(db, safe_objectid, get_current_user):
    r = APIRouter()

    async def _scope_for_corporate(user: dict) -> Optional[List[str]]:
        """Return employee_ids list to scope queries for corporate_admin, or None for global access."""
        if user["role"] in ("master_admin", "super_admin"):
            return None
        if user["role"] == "corporate_admin":
            company_id = user.get("company_id")
            if not company_id:
                return []
            emp_ids = [str(u["_id"]) async for u in db.users.find({"company_id": company_id, "role": "employee"}, {"_id": 1})]
            return emp_ids
        raise HTTPException(status_code=403, detail="Not allowed to export this report")

    @r.get("/exports/reservations")
    async def export_reservations(
        format: str = Query("xlsx"),
        date_from: Optional[str] = Query(None, alias="from"),
        date_to: Optional[str] = Query(None, alias="to"),
        user: dict = Depends(get_current_user),
    ):
        scope = await _scope_for_corporate(user)
        d_from = _parse_date(date_from, -30)
        d_to = _parse_date(date_to, 0) + timedelta(days=1)
        query: Dict[str, Any] = {"delivery_date": {"$gte": d_from, "$lt": d_to}}
        if user["role"] == "corporate_admin":
            if not scope:
                rows: List[Dict[str, Any]] = []
            else:
                # company-bound reservations either by employee_id OR by corporate_bulk's company_id
                query["$or"] = [
                    {"employee_id": {"$in": scope}},
                    {"source": "corporate_bulk", "company_id": user.get("company_id")},
                ]
        docs = await db.reservations.find(query).sort("delivery_date", -1).to_list(5000)
        rows = []
        for d in docs:
            rows.append({
                "Reservation ID": str(d.get("_id")),
                "Delivery Date": _to_iso(d.get("delivery_date")).split(" ")[0],
                "Meal Period": d.get("meal_period", "") or "",
                "Meal Type": MEAL_TYPE_LABELS.get(d.get("meal_type") or "", d.get("meal_type") or ""),
                "Status": d.get("status", "") or "",
                "Source": d.get("source", "employee") or "employee",
                "Employee Name": d.get("employee_name", "") or "",
                "Employee Email": d.get("employee_email", "") or "",
                "Vendor": d.get("vendor_name", "") or "",
                "Site ID": d.get("site_id", "") or "",
                "Created At": _to_iso(d.get("created_at")),
                "Consumed At": _to_iso(d.get("consumed_at")),
            })
        cols = ["Reservation ID", "Delivery Date", "Meal Period", "Meal Type", "Status", "Source", "Employee Name", "Employee Email", "Vendor", "Site ID", "Created At", "Consumed At"]
        subtitle = f"{d_from.date().isoformat()} → {(d_to - timedelta(days=1)).date().isoformat()} · {len(rows)} rows"
        return _send(rows, cols, f"cravitoo-reservations-{datetime.now(timezone.utc).strftime('%Y%m%d')}", format, "Cravitoo — Reservations Report", subtitle)

    @r.get("/exports/orders")
    async def export_orders(
        format: str = Query("xlsx"),
        date_from: Optional[str] = Query(None, alias="from"),
        date_to: Optional[str] = Query(None, alias="to"),
        user: dict = Depends(get_current_user),
    ):
        scope = await _scope_for_corporate(user)
        d_from = _parse_date(date_from, -30)
        d_to = _parse_date(date_to, 0) + timedelta(days=1)
        query: Dict[str, Any] = {"created_at": {"$gte": d_from, "$lt": d_to}}
        if user["role"] == "corporate_admin":
            if not scope:
                rows: List[Dict[str, Any]] = []
            else:
                query["user_id"] = {"$in": scope}
        docs = await db.orders.find(query).sort("created_at", -1).to_list(5000)
        rows = []
        for d in docs:
            items = d.get("items") or []
            items_summary = "; ".join(f"{(it.get('name') or it.get('menu_item_name') or '?')} x{it.get('quantity', 1)}" for it in items[:10])
            rows.append({
                "Order ID": str(d.get("_id")),
                "Created At": _to_iso(d.get("created_at")),
                "Status": d.get("status", "") or "",
                "Payment Status": d.get("payment_status", "") or "",
                "Customer Email": d.get("user_email", "") or "",
                "Vendor": d.get("vendor_name", "") or "",
                "Items": items_summary,
                "Subtotal": float(d.get("subtotal", d.get("total_amount", 0)) or 0),
                "Total": float(d.get("total_amount", 0) or 0),
                "Site ID": d.get("site_id", "") or "",
            })
        cols = ["Order ID", "Created At", "Status", "Payment Status", "Customer Email", "Vendor", "Items", "Subtotal", "Total", "Site ID"]
        subtitle = f"{d_from.date().isoformat()} → {(d_to - timedelta(days=1)).date().isoformat()} · {len(rows)} rows"
        return _send(rows, cols, f"cravitoo-orders-{datetime.now(timezone.utc).strftime('%Y%m%d')}", format, "Cravitoo — Orders Report", subtitle)

    @r.get("/exports/vendor-sales")
    async def export_vendor_sales(
        format: str = Query("xlsx"),
        date_from: Optional[str] = Query(None, alias="from"),
        date_to: Optional[str] = Query(None, alias="to"),
        user: dict = Depends(get_current_user),
    ):
        d_from = _parse_date(date_from, -30)
        d_to = _parse_date(date_to, 0) + timedelta(days=1)
        if user["role"] == "vendor":
            vendor_filter = {"vendor_id": user.get("vendor_id")}
        elif user["role"] in ("master_admin", "super_admin"):
            vendor_filter = {}
        else:
            raise HTTPException(status_code=403, detail="Not allowed to export this report")

        pipeline = [
            {"$match": {**vendor_filter, "created_at": {"$gte": d_from, "$lt": d_to}, "payment_status": "paid"}},
            {"$group": {
                "_id": {"vendor_id": "$vendor_id", "vendor_name": "$vendor_name"},
                "orders": {"$sum": 1},
                "revenue": {"$sum": "$total_amount"},
            }},
            {"$sort": {"revenue": -1}},
        ]
        cursor = db.orders.aggregate(pipeline)
        rows = []
        async for g in cursor:
            rows.append({
                "Vendor": g["_id"].get("vendor_name", "") or g["_id"].get("vendor_id", ""),
                "Vendor ID": g["_id"].get("vendor_id", ""),
                "Orders": g.get("orders", 0),
                "Revenue (INR)": round(float(g.get("revenue", 0) or 0), 2),
                "AOV (INR)": round(float(g.get("revenue", 0) or 0) / max(int(g.get("orders", 0) or 0), 1), 2),
            })
        cols = ["Vendor", "Vendor ID", "Orders", "Revenue (INR)", "AOV (INR)"]
        subtitle = f"{d_from.date().isoformat()} → {(d_to - timedelta(days=1)).date().isoformat()} · {len(rows)} vendors"
        return _send(rows, cols, f"cravitoo-vendor-sales-{datetime.now(timezone.utc).strftime('%Y%m%d')}", format, "Cravitoo — Vendor Sales Report", subtitle)

    @r.get("/exports/meal-summary")
    async def export_meal_summary(
        format: str = Query("xlsx"),
        date_from: Optional[str] = Query(None, alias="from"),
        date_to: Optional[str] = Query(None, alias="to"),
        user: dict = Depends(get_current_user),
    ):
        """Reservation count rolled up by site × meal_period × meal_type. Useful for billing & ops."""
        scope = await _scope_for_corporate(user)
        d_from = _parse_date(date_from, -30)
        d_to = _parse_date(date_to, 0) + timedelta(days=1)
        match: Dict[str, Any] = {"delivery_date": {"$gte": d_from, "$lt": d_to}, "status": {"$in": ["reserved", "consumed"]}}
        if user["role"] == "corporate_admin":
            if not scope:
                match["__no_match__"] = True
            else:
                match["$or"] = [
                    {"employee_id": {"$in": scope}},
                    {"source": "corporate_bulk", "company_id": user.get("company_id")},
                ]
        pipeline = [
            {"$match": match},
            {"$group": {
                "_id": {
                    "site_id": "$site_id",
                    "meal_period": "$meal_period",
                    "meal_type": "$meal_type",
                },
                "count": {"$sum": 1},
                "consumed": {"$sum": {"$cond": [{"$eq": ["$status", "consumed"]}, 1, 0]}},
            }},
            {"$sort": {"_id.site_id": 1, "_id.meal_period": 1, "_id.meal_type": 1}},
        ]
        sites = {str(s["_id"]): s.get("name", "") async for s in db.sites.find({}, {"name": 1})}
        rows = []
        async for g in db.reservations.aggregate(pipeline):
            sid = (g["_id"] or {}).get("site_id") or ""
            rows.append({
                "Site": sites.get(sid, sid),
                "Site ID": sid,
                "Meal Period": (g["_id"] or {}).get("meal_period", ""),
                "Meal Type": MEAL_TYPE_LABELS.get((g["_id"] or {}).get("meal_type") or "", (g["_id"] or {}).get("meal_type") or ""),
                "Reserved": g.get("count", 0),
                "Consumed": g.get("consumed", 0),
            })
        cols = ["Site", "Site ID", "Meal Period", "Meal Type", "Reserved", "Consumed"]
        subtitle = f"{d_from.date().isoformat()} → {(d_to - timedelta(days=1)).date().isoformat()} · {len(rows)} groups"
        return _send(rows, cols, f"cravitoo-meal-summary-{datetime.now(timezone.utc).strftime('%Y%m%d')}", format, "Cravitoo — Meal Summary Report", subtitle)

    return r
