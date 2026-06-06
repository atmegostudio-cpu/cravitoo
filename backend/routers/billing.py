"""
Monthly Billing Engine (PDF Module 15 / Module 8).

Cron: 1st of each month at 06:00 IST runs `run_billing_for_period(previous_month)`.

For each corporate client (active, with reservations in the period):
  1. Aggregate `reservations` where status in ('reserved', 'consumed') and
     either (employee in client.employees) or (source='corporate_bulk' and company_id matches).
  2. Multiply by per-site meal-type prices from `sites.meal_prices` (defaults applied).
  3. Generate Excel (line-level) + PDF (summary) and store in `invoices` collection.
  4. Email Excel + PDF to billing_contact_email (or contact_email fallback).

Endpoints:
  GET    /api/billing/invoices             — Master: list all; Corporate Admin: list theirs
  POST   /api/billing/run                  — Master: manual trigger {month: 'YYYY-MM'}
  GET    /api/billing/invoices/{id}/download?format=xlsx|pdf
"""
from __future__ import annotations

import io
import logging
import os
from datetime import datetime, timezone, timedelta, date as date_cls
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

logger = logging.getLogger(__name__)

# Default pricing per meal type (INR) — used when a site has no `meal_prices` config.
DEFAULT_MEAL_PRICES: Dict[str, float] = {
    "veg_meal": 120.0,
    "non_veg_meal": 150.0,
    "veg_salad": 100.0,
    "non_veg_salad": 130.0,
}
MEAL_TYPE_LABELS = {
    "veg_meal": "Veg Meal",
    "non_veg_meal": "Non-Veg Meal",
    "veg_salad": "Veg Salad",
    "non_veg_salad": "Non-Veg Salad",
}

IST_TZ_OFFSET = timedelta(hours=5, minutes=30)


def _ist_now() -> datetime:
    return datetime.now(timezone.utc) + IST_TZ_OFFSET


def _month_bounds_ist(year: int, month: int) -> Tuple[datetime, datetime]:
    """Return [start, end_exclusive) of the given month in UTC (computed from IST calendar)."""
    from calendar import monthrange
    last_day = monthrange(year, month)[1]
    start_ist = datetime(year, month, 1, 0, 0)
    end_ist = datetime(year, month, last_day, 23, 59, 59)
    # Treat IST midnight as IST then shift to UTC
    start_utc = start_ist - IST_TZ_OFFSET
    end_utc = end_ist - IST_TZ_OFFSET + timedelta(seconds=1)
    return start_utc.replace(tzinfo=timezone.utc), end_utc.replace(tzinfo=timezone.utc)


def _site_prices(site_doc: Optional[Dict[str, Any]]) -> Dict[str, float]:
    """Merge site-level meal_prices with DEFAULT_MEAL_PRICES."""
    out = {**DEFAULT_MEAL_PRICES}
    if site_doc and isinstance(site_doc.get("meal_prices"), dict):
        for k, v in site_doc["meal_prices"].items():
            try:
                out[k] = float(v)
            except (TypeError, ValueError):
                continue
    return out


async def _compute_invoice_for_client(db, safe_objectid, client_doc: Dict[str, Any], start_utc: datetime, end_utc: datetime) -> Dict[str, Any]:
    """Compute totals for one corporate client for a date range."""
    company_id = str(client_doc.get("_id"))
    emp_ids = [str(u["_id"]) async for u in db.users.find({"company_id": company_id, "role": "employee"}, {"_id": 1})]

    # Pull all relevant reservations
    match: Dict[str, Any] = {
        "delivery_date": {"$gte": start_utc, "$lt": end_utc},
        "status": {"$in": ["reserved", "consumed"]},
        "$or": [
            {"employee_id": {"$in": emp_ids}},
            {"source": "corporate_bulk", "company_id": company_id},
        ],
    }
    reservations = await db.reservations.find(match).to_list(50000)

    # Group by site for pricing
    sites_cache: Dict[str, Dict[str, Any]] = {}
    line_items: List[Dict[str, Any]] = []
    totals_by_type: Dict[str, int] = {mt: 0 for mt in MEAL_TYPE_LABELS}
    grand_total = 0.0
    for r in reservations:
        site_id = r.get("site_id") or ""
        if site_id and site_id not in sites_cache:
            try:
                site = await db.sites.find_one({"_id": safe_objectid(site_id, "Site")})
                sites_cache[site_id] = site or {}
            except Exception:
                sites_cache[site_id] = {}
        prices = _site_prices(sites_cache.get(site_id))
        mt = r.get("meal_type") or "non_veg_meal"
        price = prices.get(mt, DEFAULT_MEAL_PRICES.get(mt, 0.0))
        line_items.append({
            "delivery_date": r.get("delivery_date"),
            "meal_period": r.get("meal_period") or "",
            "meal_type": mt,
            "meal_type_label": MEAL_TYPE_LABELS.get(mt, mt),
            "employee_name": r.get("employee_name") or "Corporate Bulk",
            "employee_email": r.get("employee_email") or "",
            "vendor": r.get("vendor_name") or "",
            "site_id": site_id,
            "site_name": (sites_cache.get(site_id) or {}).get("name") or "",
            "source": r.get("source") or "employee",
            "unit_price": price,
            "amount": price,
        })
        totals_by_type[mt] = totals_by_type.get(mt, 0) + 1
        grand_total += price

    return {
        "client_id": company_id,
        "client_name": client_doc.get("name", "Client"),
        "client_billing_email": client_doc.get("billing_contact_email") or client_doc.get("contact_email") or "",
        "client_billing_name": client_doc.get("billing_contact_name") or "",
        "period_start": start_utc,
        "period_end": end_utc,
        "line_items": line_items,
        "totals_by_type": totals_by_type,
        "grand_total": round(grand_total, 2),
        "line_item_count": len(line_items),
    }


def _build_excel(invoice: Dict[str, Any]) -> bytes:
    from openpyxl import Workbook  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Detail"

    header_fill = PatternFill(start_color="FF5A1F", end_color="FF5A1F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    cols = ["Delivery Date", "Meal Period", "Meal Type", "Employee Name", "Employee Email", "Vendor", "Site", "Source", "Unit Price (INR)", "Amount (INR)"]
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.fill = header_fill
        cell.font = header_font
    for r, item in enumerate(invoice["line_items"], start=2):
        dt = item["delivery_date"]
        dt_s = dt.strftime("%Y-%m-%d") if isinstance(dt, datetime) else str(dt)
        ws.cell(row=r, column=1, value=dt_s)
        ws.cell(row=r, column=2, value=item["meal_period"])
        ws.cell(row=r, column=3, value=item["meal_type_label"])
        ws.cell(row=r, column=4, value=item["employee_name"])
        ws.cell(row=r, column=5, value=item["employee_email"])
        ws.cell(row=r, column=6, value=item["vendor"])
        ws.cell(row=r, column=7, value=item["site_name"])
        ws.cell(row=r, column=8, value=item["source"])
        ws.cell(row=r, column=9, value=item["unit_price"])
        ws.cell(row=r, column=10, value=item["amount"])
    for i, c in enumerate(cols, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(c) + 2)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = f"Invoice for {invoice['client_name']}"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A2"] = f"Period: {invoice['period_start'].strftime('%b %d, %Y')} – {(invoice['period_end'] - timedelta(seconds=1)).strftime('%b %d, %Y')}"
    ws2["A3"] = f"Billing email: {invoice['client_billing_email']}"
    ws2["A5"] = "Meal Type"
    ws2["B5"] = "Quantity"
    ws2["A5"].font = ws2["B5"].font = Font(bold=True)
    row = 6
    for mt, qty in invoice["totals_by_type"].items():
        if qty == 0:
            continue
        ws2[f"A{row}"] = MEAL_TYPE_LABELS.get(mt, mt)
        ws2[f"B{row}"] = qty
        row += 1
    ws2[f"A{row+1}"] = "Total Meals"
    ws2[f"B{row+1}"] = invoice["line_item_count"]
    ws2[f"A{row+1}"].font = ws2[f"B{row+1}"].font = Font(bold=True)
    ws2[f"A{row+2}"] = "Grand Total (INR)"
    ws2[f"B{row+2}"] = invoice["grand_total"]
    ws2[f"A{row+2}"].font = ws2[f"B{row+2}"].font = Font(bold=True)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_pdf(invoice: Dict[str, Any]) -> bytes:
    from reportlab.lib.pagesizes import A4  # type: ignore
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer  # type: ignore
    from reportlab.lib import colors  # type: ignore
    from reportlab.lib.styles import getSampleStyleSheet  # type: ignore

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()

    period_label = f"{invoice['period_start'].strftime('%b %d, %Y')} – {(invoice['period_end'] - timedelta(seconds=1)).strftime('%b %d, %Y')}"
    elements = [
        Paragraph(f"<b>Cravitoo Invoice</b>", styles["Title"]),
        Spacer(1, 6),
        Paragraph(f"<b>{invoice['client_name']}</b>", styles["Heading2"]),
        Paragraph(f"Period: {period_label}", styles["Normal"]),
        Paragraph(f"Billing contact: {invoice['client_billing_name'] or '—'} &lt;{invoice['client_billing_email'] or '—'}&gt;", styles["Normal"]),
        Spacer(1, 18),
        Paragraph("<b>Summary</b>", styles["Heading3"]),
    ]
    summary_rows = [["Meal Type", "Quantity"]]
    for mt, qty in invoice["totals_by_type"].items():
        if qty == 0:
            continue
        summary_rows.append([MEAL_TYPE_LABELS.get(mt, mt), qty])
    summary_rows.append(["TOTAL MEALS", invoice["line_item_count"]])
    summary_rows.append(["GRAND TOTAL (INR)", f"₹ {invoice['grand_total']:,.2f}"])
    t = Table(summary_rows, colWidths=[280, 100])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF5A1F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -2), (-1, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFEDD5")),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        "This is a system-generated invoice. For questions email <b>billing@cravitoo.com</b>.",
        styles["Normal"],
    ))

    doc.build(elements)
    buf.seek(0)
    return buf.read()


async def run_billing_for_period(db, safe_objectid, year: int, month: int, triggered_by: Optional[str] = None) -> Dict[str, Any]:
    """Generate and store invoices for all 'active' or 'approved' corporate clients for the given month."""
    start, end = _month_bounds_ist(year, month)
    period_label = f"{year:04d}-{month:02d}"
    logger.info(f"Billing: starting run for {period_label} ({start.isoformat()} → {end.isoformat()})")

    eligible_clients = await db.companies.find({
        "$or": [
            {"lifecycle_status": {"$in": ["approved", "active"]}},
            {"lifecycle_status": {"$exists": False}, "status": "active"},
        ],
    }).to_list(500)

    generated: List[Dict[str, Any]] = []
    for client in eligible_clients:
        invoice = await _compute_invoice_for_client(db, safe_objectid, client, start, end)
        if invoice["line_item_count"] == 0:
            continue  # skip clients with no activity
        xlsx_bytes = _build_excel(invoice)
        pdf_bytes = _build_pdf(invoice)
        invoice_doc = {
            "client_id": invoice["client_id"],
            "client_name": invoice["client_name"],
            "period": period_label,
            "period_start": invoice["period_start"],
            "period_end": invoice["period_end"],
            "totals_by_type": invoice["totals_by_type"],
            "grand_total": invoice["grand_total"],
            "line_item_count": invoice["line_item_count"],
            "billing_email": invoice["client_billing_email"],
            "xlsx_blob": xlsx_bytes,
            "pdf_blob": pdf_bytes,
            "email_sent": False,
            "email_error": None,
            "generated_at": datetime.now(timezone.utc),
            "triggered_by": triggered_by or "cron",
        }
        # Upsert: one invoice per (client, period)
        result = await db.invoices.find_one_and_replace(
            {"client_id": invoice["client_id"], "period": period_label},
            invoice_doc,
            upsert=True,
            return_document=True,
        )
        invoice_id = str((result or invoice_doc).get("_id"))

        # Email (best-effort)
        try:
            if invoice["client_billing_email"]:
                import email_service as _email_service
                subject = f"Cravitoo Invoice — {invoice['client_name']} — {period_label}"
                html = (
                    f"<p>Hi {invoice['client_billing_name'] or 'there'},</p>"
                    f"<p>Please find attached your Cravitoo invoice for <b>{period_label}</b> covering "
                    f"<b>{invoice['line_item_count']}</b> meals across your team. Grand total: "
                    f"<b>₹ {invoice['grand_total']:,.2f}</b>.</p>"
                    f"<p>Two attachments: Excel (line-level details) and PDF (summary).</p>"
                    f"<p>— Team Cravitoo</p>"
                )
                text = (
                    f"Cravitoo Invoice — {invoice['client_name']} — {period_label}\n"
                    f"{invoice['line_item_count']} meals · ₹ {invoice['grand_total']:,.2f}\n"
                )
                attachments = [
                    {"filename": f"cravitoo-invoice-{period_label}.xlsx", "content": xlsx_bytes, "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
                    {"filename": f"cravitoo-invoice-{period_label}.pdf", "content": pdf_bytes, "content_type": "application/pdf"},
                ]
                ok, err = _email_service.send_email(invoice["client_billing_email"], subject, html, text, attachments=attachments)
                await db.invoices.update_one({"_id": (result or invoice_doc).get("_id")}, {"$set": {"email_sent": bool(ok), "email_error": err}})
        except Exception as e:
            logger.warning(f"Billing email failed for {invoice['client_billing_email']}: {e}")
            await db.invoices.update_one({"_id": (result or invoice_doc).get("_id")}, {"$set": {"email_sent": False, "email_error": str(e)}})

        generated.append({
            "id": invoice_id,
            "client_name": invoice["client_name"],
            "period": period_label,
            "grand_total": invoice["grand_total"],
            "line_item_count": invoice["line_item_count"],
        })

    logger.info(f"Billing: {len(generated)} invoices for {period_label}")
    return {"period": period_label, "invoices_generated": generated, "skipped_clients": len(eligible_clients) - len(generated)}


def _is_master(user: dict) -> bool:
    return user["role"] in ("master_admin", "super_admin")


def _is_corp_admin(user: dict) -> bool:
    return user["role"] == "corporate_admin"


class ManualBillingRun(BaseModel):
    month: str  # YYYY-MM


def make_router(db, safe_objectid, get_current_user):
    r = APIRouter()

    @r.post("/billing/run")
    async def manual_run(payload: ManualBillingRun, user: dict = Depends(get_current_user)):
        if not _is_master(user):
            raise HTTPException(status_code=403, detail="Only Master Admin can trigger billing")
        try:
            year_s, month_s = payload.month.split("-")
            year = int(year_s); month = int(month_s)
        except Exception:
            raise HTTPException(status_code=400, detail="month must be 'YYYY-MM' (e.g. 2026-05)")
        return await run_billing_for_period(db, safe_objectid, year, month, triggered_by=user.get("email"))

    @r.get("/billing/invoices")
    async def list_invoices(
        period: Optional[str] = Query(None, description="Filter by 'YYYY-MM'"),
        user: dict = Depends(get_current_user),
    ):
        query: Dict[str, Any] = {}
        if period:
            query["period"] = period
        if _is_master(user):
            pass
        elif _is_corp_admin(user):
            query["client_id"] = user.get("company_id")
        else:
            raise HTTPException(status_code=403, detail="Not allowed")
        invs = await db.invoices.find(query, {"xlsx_blob": 0, "pdf_blob": 0}).sort("generated_at", -1).to_list(500)
        out = []
        for d in invs:
            d["id"] = str(d.pop("_id"))
            for k in ("generated_at", "period_start", "period_end"):
                if isinstance(d.get(k), datetime):
                    d[k] = d[k].isoformat()
            out.append(d)
        return out

    @r.get("/billing/invoices/{invoice_id}/download")
    async def download_invoice(invoice_id: str, format: str = Query("xlsx"), user: dict = Depends(get_current_user)):
        fmt = (format or "xlsx").lower()
        if fmt not in ("xlsx", "pdf"):
            raise HTTPException(status_code=400, detail="format must be xlsx or pdf")
        doc = await db.invoices.find_one({"_id": safe_objectid(invoice_id, "Invoice")})
        if not doc:
            raise HTTPException(status_code=404, detail="Invoice not found")
        if _is_corp_admin(user) and doc.get("client_id") != user.get("company_id"):
            raise HTTPException(status_code=403, detail="Not your invoice")
        if not _is_master(user) and not _is_corp_admin(user):
            raise HTTPException(status_code=403, detail="Not allowed")
        blob_key = "xlsx_blob" if fmt == "xlsx" else "pdf_blob"
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if fmt == "xlsx" else "application/pdf"
        blob = doc.get(blob_key)
        if not blob:
            raise HTTPException(status_code=404, detail="Invoice format unavailable")
        fname = f"cravitoo-invoice-{doc.get('client_name', 'client').replace(' ', '-')}-{doc.get('period')}.{fmt}"
        return StreamingResponse(
            iter([bytes(blob)]),
            media_type=media,
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    @r.post("/billing/invoices/{invoice_id}/resend")
    async def resend_invoice(invoice_id: str, user: dict = Depends(get_current_user)):
        if not _is_master(user):
            raise HTTPException(status_code=403, detail="Only Master Admin can resend invoices")
        doc = await db.invoices.find_one({"_id": safe_objectid(invoice_id, "Invoice")})
        if not doc:
            raise HTTPException(status_code=404, detail="Invoice not found")
        if not doc.get("billing_email"):
            raise HTTPException(status_code=400, detail="Invoice has no billing email")
        try:
            import email_service as _email_service
            subject = f"Cravitoo Invoice (resent) — {doc.get('client_name')} — {doc.get('period')}"
            html = f"<p>Resending your invoice for <b>{doc.get('period')}</b>. Grand total: <b>₹ {doc.get('grand_total', 0):,.2f}</b>.</p>"
            text = f"Cravitoo Invoice resent for {doc.get('period')}"
            attachments = [
                {"filename": f"cravitoo-invoice-{doc.get('period')}.xlsx", "content": bytes(doc.get("xlsx_blob") or b""), "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
                {"filename": f"cravitoo-invoice-{doc.get('period')}.pdf", "content": bytes(doc.get("pdf_blob") or b""), "content_type": "application/pdf"},
            ]
            ok, err = _email_service.send_email(doc["billing_email"], subject, html, text, attachments=attachments)
            await db.invoices.update_one({"_id": doc["_id"]}, {"$set": {"email_sent": bool(ok), "email_error": err}})
            return {"ok": bool(ok), "error": err}
        except Exception as e:
            logger.warning(f"Resend invoice failed: {e}")
            raise HTTPException(status_code=500, detail=str(e))

    return r
