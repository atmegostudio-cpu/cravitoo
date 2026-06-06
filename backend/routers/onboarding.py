"""
Vendor Onboarding routes (extracted from server.py — iter17 phase 2 refactor).

Endpoints:
  POST   /api/onboarding/vendors/bulk-import
  POST   /api/onboarding/vendors/{onb_id}/menu/upload-excel
  POST   /api/onboarding/vendors
  GET    /api/onboarding/vendors
  GET    /api/onboarding/vendors/{onb_id}
  PATCH  /api/onboarding/vendors/{onb_id}
  PATCH  /api/onboarding/vendors/{onb_id}/checklist
  POST   /api/onboarding/vendors/{onb_id}/documents/{doc_type}
  DELETE /api/onboarding/vendors/{onb_id}/documents/{doc_type}
  POST   /api/onboarding/vendors/{onb_id}/submit-to-master
  POST   /api/onboarding/vendors/{onb_id}/site-review
  POST   /api/onboarding/vendors/{onb_id}/master-decision
  GET    /api/onboarding/vendors/{onb_id}/audit-trail
  GET    /api/onboarding/dashboard

Built as a make_router(...) factory to avoid circular imports.
"""

from __future__ import annotations

import io
import logging
import os
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile

from models import (
    CHECKLIST_FIELDS,
    DOC_TYPES,
    ONBOARDING_STATUSES,
    ChecklistUpdate,
    OnboardingDecision,
    VendorOnboardingBasic,
    VendorOnboardingUpdate,
)

logger = logging.getLogger(__name__)


def calc_checklist_pct(checklist: dict) -> int:
    if not checklist:
        return 0
    done = sum(1 for f in CHECKLIST_FIELDS if checklist.get(f))
    return int(done * 100 / len(CHECKLIST_FIELDS))


def onboarding_to_dict(o):
    return {
        "id": str(o["_id"]),
        "vendor_name": o.get("vendor_name"),
        "company_name": o.get("company_name"),
        "contact_person": o.get("contact_person"),
        "mobile_number": o.get("mobile_number"),
        "email": o.get("email"),
        "business_address": o.get("business_address"),
        "cuisine_type": o.get("cuisine_type"),
        "site_id": o.get("site_id"),
        "city_id": o.get("city_id"),
        "status": o.get("status", "draft"),
        "checklist": o.get("checklist", {}),
        "checklist_pct": calc_checklist_pct(o.get("checklist", {})),
        "documents": o.get("documents", {}),
        "draft_menu": o.get("draft_menu", []),
        "vendor_id": o.get("vendor_id"),  # set when approved
        "remarks": o.get("remarks", []),
        "created_by": o.get("created_by"),
        "created_at": o.get("created_at").isoformat() if o.get("created_at") else None,
        "updated_at": o.get("updated_at").isoformat() if o.get("updated_at") else None,
    }


def _is_master(user: dict) -> bool:
    return user.get("role") == "master_admin"


def make_router(db, safe_objectid, get_current_user, audit_log, UPLOAD_DIR: Path):
    r = APIRouter()

    @r.post("/onboarding/vendors/bulk-import")
    async def bulk_import_onboardings(
        file: UploadFile = File(...),
        site_id: str = "",
        user: dict = Depends(get_current_user),
    ):
        """Master/Site/City admin uploads an Excel sheet to bulk-create onboarding records.
        Columns: vendor_name, company_name, contact_person, mobile_number, email, business_address, cuisine_type"""
        if user["role"] not in ("master_admin", "city_admin", "site_admin"):
            raise HTTPException(status_code=403, detail="Access denied")
        if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
            raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")
        if user["role"] == "site_admin":
            site_id = user.get("site_id") or site_id
        if not site_id:
            raise HTTPException(status_code=400, detail="site_id required")
        site = await db.sites.find_one({"_id": safe_objectid(site_id, "Site")})
        if not site:
            raise HTTPException(status_code=404, detail="Site not found")
        if user["role"] == "site_admin" and site_id != user.get("site_id"):
            raise HTTPException(status_code=403, detail="Not your site")

        content = await file.read()
        if len(content) > 2 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File must be under 2 MB")

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid Excel: {e}")

        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="Excel must have header + at least 1 data row")
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        required = ["vendor_name", "company_name", "contact_person", "mobile_number", "email", "business_address"]
        missing = [r for r in required if r not in headers]
        if missing:
            raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(missing)}")

        inserted, errors = 0, []
        for idx, row in enumerate(rows[1:], start=2):
            try:
                rec = {headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(min(len(headers), len(row)))}
                if not rec.get("vendor_name") or not rec.get("email"):
                    errors.append({"row": idx, "error": "vendor_name and email required"})
                    continue
                doc = {
                    "vendor_name": rec.get("vendor_name"),
                    "company_name": rec.get("company_name", ""),
                    "contact_person": rec.get("contact_person", ""),
                    "mobile_number": rec.get("mobile_number", ""),
                    "email": rec.get("email"),
                    "business_address": rec.get("business_address", ""),
                    "cuisine_type": rec.get("cuisine_type", "Multi-cuisine"),
                    "site_id": site_id,
                    "city_id": site.get("city_id"),
                    "status": "draft",
                    "checklist": {},
                    "documents": {},
                    "remarks": [],
                    "created_by": user["id"],
                    "created_at": datetime.now(timezone.utc),
                    "updated_at": datetime.now(timezone.utc),
                }
                res = await db.vendor_onboarding.insert_one(doc)
                await audit_log(user, "vendor_onboarding", str(res.inserted_id), "bulk_imported", {"vendor_name": rec.get("vendor_name")})
                inserted += 1
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})

        return {"inserted": inserted, "errors": errors, "total_attempted": len(rows) - 1}

    @r.post("/onboarding/vendors/{onb_id}/menu/upload-excel")
    async def onboarding_menu_excel(
        onb_id: str,
        file: UploadFile = File(...),
        user: dict = Depends(get_current_user),
    ):
        """Pre-load menu items as 'draft' under onboarding — get activated when vendor is approved."""
        if user["role"] not in ("master_admin", "city_admin", "site_admin"):
            raise HTTPException(status_code=403, detail="Access denied")
        o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
        if not o:
            raise HTTPException(status_code=404, detail="Onboarding not found")
        if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
            raise HTTPException(status_code=403, detail="Not your site")
        if o.get("status") in ("approved", "active", "rejected"):
            raise HTTPException(status_code=400, detail=f"Cannot edit menu — status is '{o.get('status')}'")
        if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
            raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")
        content = await file.read()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid Excel: {e}")
        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="Excel must have header + at least 1 data row")
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        required = ["name", "category", "price"]
        missing = [rq for rq in required if rq not in headers]
        if missing:
            raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(missing)}")
        items = []
        errors = []
        for idx, row in enumerate(rows[1:], start=2):
            try:
                rec = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
                name = (str(rec.get("name") or "")).strip()
                if not name:
                    errors.append({"row": idx, "error": "Missing name"})
                    continue
                price = float(rec.get("price") or 0)
                if price <= 0:
                    errors.append({"row": idx, "error": f"Invalid price for {name}"})
                    continue
                items.append({
                    "name": name,
                    "description": str(rec.get("description") or ""),
                    "category": str(rec.get("category") or "Main").strip(),
                    "price": price,
                    "is_vegetarian": str(rec.get("is_vegetarian", "")).lower() in ("true", "yes", "1", "veg"),
                    "image_url": str(rec.get("image_url") or "") or None,
                    "is_available": True,
                })
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})
        await db.vendor_onboarding.update_one({"_id": o["_id"]}, {"$set": {"draft_menu": items, "updated_at": datetime.now(timezone.utc)}})
        # Auto-tick "menu_uploaded" in checklist
        new_checklist = {**(o.get("checklist", {})), "menu_uploaded": True}
        await db.vendor_onboarding.update_one({"_id": o["_id"]}, {"$set": {"checklist": new_checklist}})
        await audit_log(user, "vendor_onboarding", onb_id, "menu_uploaded", {"items": len(items)})
        return {"inserted": len(items), "errors": errors, "total_attempted": len(rows) - 1}

    @r.post("/onboarding/vendors")
    async def create_vendor_onboarding(data: VendorOnboardingBasic, user: dict = Depends(get_current_user)):
        if user["role"] not in ("site_admin", "master_admin", "city_admin"):
            raise HTTPException(status_code=403, detail="Only site_admin, city_admin, or master_admin can onboard vendors")
        # Site admin can only onboard for their own site
        if user["role"] == "site_admin" and user.get("site_id") != data.site_id:
            raise HTTPException(status_code=403, detail="You can only onboard vendors for your own site")
        site = await db.sites.find_one({"_id": safe_objectid(data.site_id, "Site")})
        if not site:
            raise HTTPException(status_code=404, detail="Site not found")
        if user["role"] == "city_admin" and site.get("city_id") != user.get("city_id"):
            raise HTTPException(status_code=403, detail="Site is not in your city")
        doc = {
            **data.model_dump(),
            "city_id": site.get("city_id"),
            "status": "draft",
            "checklist": {},
            "documents": {},
            "remarks": [],
            "created_by": user["id"],
            "created_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc),
        }
        res = await db.vendor_onboarding.insert_one(doc)
        onb_id = str(res.inserted_id)
        await audit_log(user, "vendor_onboarding", onb_id, "created", {"vendor_name": data.vendor_name, "site_id": data.site_id})
        doc["_id"] = res.inserted_id
        return onboarding_to_dict(doc)

    @r.get("/onboarding/vendors")
    async def list_vendor_onboardings(
        status: Optional[str] = None,
        user: dict = Depends(get_current_user),
    ):
        """List based on role:
        - master_admin: all
        - city_admin: only in their city
        - site_admin: only for their site
        - others: 403"""
        filt = {}
        if user["role"] == "master_admin":
            pass
        elif user["role"] == "city_admin":
            filt["city_id"] = user.get("city_id")
        elif user["role"] == "site_admin":
            filt["site_id"] = user.get("site_id")
        else:
            raise HTTPException(status_code=403, detail="Access denied")
        if status and status in ONBOARDING_STATUSES:
            filt["status"] = status
        cursor = db.vendor_onboarding.find(filt).sort("created_at", -1).limit(200)
        return [onboarding_to_dict(o) async for o in cursor]

    @r.get("/onboarding/vendors/{onb_id}")
    async def get_vendor_onboarding(onb_id: str, user: dict = Depends(get_current_user)):
        o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
        if not o:
            raise HTTPException(status_code=404, detail="Onboarding not found")
        # Role-based access
        if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
            raise HTTPException(status_code=403, detail="Not your site")
        if user["role"] == "city_admin" and o.get("city_id") != user.get("city_id"):
            raise HTTPException(status_code=403, detail="Not your city")
        if user["role"] not in ("master_admin", "city_admin", "site_admin"):
            raise HTTPException(status_code=403, detail="Access denied")
        return onboarding_to_dict(o)

    @r.patch("/onboarding/vendors/{onb_id}")
    async def update_vendor_onboarding(onb_id: str, data: VendorOnboardingUpdate, user: dict = Depends(get_current_user)):
        o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
        if not o:
            raise HTTPException(status_code=404, detail="Onboarding not found")
        if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
            raise HTTPException(status_code=403, detail="Not your site")
        if user["role"] not in ("master_admin", "city_admin", "site_admin"):
            raise HTTPException(status_code=403, detail="Access denied")
        if o.get("status") in ("approved", "active", "rejected"):
            raise HTTPException(status_code=400, detail=f"Cannot edit onboarding with status '{o.get('status')}'")
        updates = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
        if not updates:
            raise HTTPException(status_code=400, detail="No fields to update")
        updates["updated_at"] = datetime.now(timezone.utc)
        await db.vendor_onboarding.update_one({"_id": o["_id"]}, {"$set": updates})
        await audit_log(user, "vendor_onboarding", onb_id, "updated", updates)
        return {"message": "Updated"}

    @r.patch("/onboarding/vendors/{onb_id}/checklist")
    async def update_checklist(onb_id: str, data: ChecklistUpdate, user: dict = Depends(get_current_user)):
        o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
        if not o:
            raise HTTPException(status_code=404, detail="Onboarding not found")
        if user["role"] not in ("master_admin", "city_admin", "site_admin"):
            raise HTTPException(status_code=403, detail="Access denied")
        if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
            raise HTTPException(status_code=403, detail="Not your site")
        updates = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
        if not updates:
            raise HTTPException(status_code=400, detail="No fields to update")
        new_checklist = {**(o.get("checklist", {})), **{k: v for k, v in updates.items() if k != "notes"}}
        set_doc = {"checklist": new_checklist, "updated_at": datetime.now(timezone.utc)}
        if "notes" in updates:
            set_doc["checklist_notes"] = updates["notes"]
        await db.vendor_onboarding.update_one({"_id": o["_id"]}, {"$set": set_doc})
        pct = calc_checklist_pct(new_checklist)
        await audit_log(user, "vendor_onboarding", onb_id, "checklist_updated", {"updates": updates, "pct": pct})
        return {"checklist": new_checklist, "checklist_pct": pct}

    @r.post("/onboarding/vendors/{onb_id}/documents/{doc_type}")
    async def upload_onboarding_doc(
        onb_id: str,
        doc_type: str,
        file: UploadFile = File(...),
        user: dict = Depends(get_current_user),
    ):
        if doc_type not in DOC_TYPES:
            raise HTTPException(status_code=400, detail=f"Invalid doc_type. Must be one of: {', '.join(DOC_TYPES)}")
        o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
        if not o:
            raise HTTPException(status_code=404, detail="Onboarding not found")
        if user["role"] not in ("master_admin", "city_admin", "site_admin"):
            raise HTTPException(status_code=403, detail="Access denied")
        if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
            raise HTTPException(status_code=403, detail="Not your site")

        # File size limit 10 MB for compliance docs (can be larger PDFs)
        content = await file.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File must be under 10 MB")
        ext = (file.filename or "doc").rsplit(".", 1)[-1].lower()[:5]
        if ext not in ("pdf", "png", "jpg", "jpeg", "webp"):
            raise HTTPException(status_code=400, detail="Allowed types: PDF, PNG, JPG, JPEG, WEBP")
        fname = f"onb_{uuid.uuid4().hex}.{ext}"
        fpath = UPLOAD_DIR / fname
        with open(fpath, "wb") as f:
            f.write(content)
        base = os.environ.get('PUBLIC_BACKEND_URL', '').rstrip('/')
        url = f"{base}/api/uploads/{fname}" if base else f"/api/uploads/{fname}"
        docs = o.get("documents", {})
        docs[doc_type] = {
            "url": url,
            "filename": fname,
            "original_name": file.filename,
            "uploaded_by": user["email"],
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "size": len(content),
        }
        set_doc = {"documents": docs, "updated_at": datetime.now(timezone.utc)}
        # Auto-flip status from draft → documents_pending after first upload
        if o.get("status") == "draft":
            set_doc["status"] = "documents_pending"
        await db.vendor_onboarding.update_one({"_id": o["_id"]}, {"$set": set_doc})
        await audit_log(user, "vendor_onboarding", onb_id, "uploaded_doc", {"doc_type": doc_type, "filename": fname})
        return {"doc_type": doc_type, "url": url, "filename": fname}

    @r.delete("/onboarding/vendors/{onb_id}/documents/{doc_type}")
    async def delete_onboarding_doc(onb_id: str, doc_type: str, user: dict = Depends(get_current_user)):
        o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
        if not o:
            raise HTTPException(status_code=404, detail="Onboarding not found")
        if user["role"] not in ("master_admin", "city_admin", "site_admin"):
            raise HTTPException(status_code=403, detail="Access denied")
        if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
            raise HTTPException(status_code=403, detail="Not your site")
        if o.get("status") in ("approved", "active"):
            raise HTTPException(status_code=400, detail="Cannot delete docs after approval")
        docs = o.get("documents", {})
        if doc_type not in docs:
            raise HTTPException(status_code=404, detail="Document not found")
        docs.pop(doc_type, None)
        await db.vendor_onboarding.update_one({"_id": o["_id"]}, {"$set": {"documents": docs, "updated_at": datetime.now(timezone.utc)}})
        await audit_log(user, "vendor_onboarding", onb_id, "deleted_doc", {"doc_type": doc_type})
        return {"message": "Deleted"}

    @r.post("/onboarding/vendors/{onb_id}/submit-to-master")
    async def submit_to_master(onb_id: str, user: dict = Depends(get_current_user)):
        """Site admin submits to master after their review."""
        o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
        if not o:
            raise HTTPException(status_code=404, detail="Onboarding not found")
        if user["role"] not in ("site_admin", "city_admin", "master_admin"):
            raise HTTPException(status_code=403, detail="Access denied")
        if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
            raise HTTPException(status_code=403, detail="Not your site")
        if o.get("status") not in ("documents_pending", "under_site_review", "changes_requested"):
            raise HTTPException(status_code=400, detail=f"Cannot submit from status '{o.get('status')}'")
        pct = calc_checklist_pct(o.get("checklist", {}))
        if pct < 80:
            raise HTTPException(status_code=400, detail=f"Checklist must be at least 80% complete (currently {pct}%)")
        await db.vendor_onboarding.update_one(
            {"_id": o["_id"]},
            {"$set": {"status": "under_master_review", "site_reviewed_by": user["id"], "site_reviewed_at": datetime.now(timezone.utc), "updated_at": datetime.now(timezone.utc)}}
        )
        await audit_log(user, "vendor_onboarding", onb_id, "submitted_to_master", {"checklist_pct": pct})
        return {"message": "Submitted to master admin for final approval", "status": "under_master_review"}

    @r.post("/onboarding/vendors/{onb_id}/site-review")
    async def site_review(onb_id: str, data: OnboardingDecision, user: dict = Depends(get_current_user)):
        """Site admin/city admin reviews — Approve→sub_to_master, Reject, Request changes."""
        o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
        if not o:
            raise HTTPException(status_code=404, detail="Onboarding not found")
        if user["role"] not in ("site_admin", "city_admin", "master_admin"):
            raise HTTPException(status_code=403, detail="Access denied")
        if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
            raise HTTPException(status_code=403, detail="Not your site")
        if data.decision not in ("approve", "reject", "request_changes"):
            raise HTTPException(status_code=400, detail="decision must be approve|reject|request_changes")
        remark = {
            "stage": "site_review",
            "by": user["email"],
            "decision": data.decision,
            "remarks": data.remarks or "",
            "at": datetime.now(timezone.utc).isoformat(),
        }
        new_remarks = o.get("remarks", []) + [remark]
        set_doc = {"remarks": new_remarks, "updated_at": datetime.now(timezone.utc)}
        if data.decision == "approve":
            pct = calc_checklist_pct(o.get("checklist", {}))
            if pct < 80:
                raise HTTPException(status_code=400, detail=f"Checklist must be at least 80% complete to approve (currently {pct}%)")
            set_doc["status"] = "under_master_review"
            set_doc["site_reviewed_by"] = user["id"]
            set_doc["site_reviewed_at"] = datetime.now(timezone.utc)
        elif data.decision == "reject":
            set_doc["status"] = "rejected"
        else:  # request_changes
            set_doc["status"] = "changes_requested"
        await db.vendor_onboarding.update_one({"_id": o["_id"]}, {"$set": set_doc})
        await audit_log(user, "vendor_onboarding", onb_id, f"site_{data.decision}", {"remarks": data.remarks})
        return {"message": f"Site review: {data.decision}", "status": set_doc["status"]}

    @r.post("/onboarding/vendors/{onb_id}/master-decision")
    async def master_decision(onb_id: str, data: OnboardingDecision, user: dict = Depends(get_current_user)):
        """Master admin final approval/rejection."""
        if not _is_master(user):
            raise HTTPException(status_code=403, detail="Only master admin")
        o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
        if not o:
            raise HTTPException(status_code=404, detail="Onboarding not found")
        if o.get("status") != "under_master_review":
            raise HTTPException(status_code=400, detail=f"Cannot finalize — current status is '{o.get('status')}', must be 'under_master_review'")
        if data.decision not in ("approve", "reject"):
            raise HTTPException(status_code=400, detail="decision must be approve|reject")
        remark = {
            "stage": "master_review",
            "by": user["email"],
            "decision": data.decision,
            "remarks": data.remarks or "",
            "at": datetime.now(timezone.utc).isoformat(),
        }
        new_remarks = o.get("remarks", []) + [remark]
        set_doc = {"remarks": new_remarks, "updated_at": datetime.now(timezone.utc),
                   "master_reviewed_by": user["id"], "master_reviewed_at": datetime.now(timezone.utc)}
        if data.decision == "approve":
            # Create real Vendor business record + map to site
            vendor_doc = {
                "name": o.get("vendor_name"),
                "description": o.get("company_name", ""),
                "cuisine_type": o.get("cuisine_type", "Multi-cuisine"),
                "phone": o.get("mobile_number"),
                "email": o.get("email"),
                "address": o.get("business_address"),
                "rating": 0.0,
                "status": "active",
                "commission_pct": 15.0,
                "onboarding_id": str(o["_id"]),
                "created_at": datetime.now(timezone.utc),
            }
            vres = await db.vendors.insert_one(vendor_doc)
            vendor_id = str(vres.inserted_id)
            # Site-vendor mapping (uses canonical collection name `vendor_site_mappings`)
            await db.vendor_site_mappings.insert_one({
                "site_id": o.get("site_id"),
                "vendor_id": vendor_id,
                "status": "active",
                "created_at": datetime.now(timezone.utc),
            })
            # Create vendor LOGIN user (passwordless via email OTP) if none exists yet
            vendor_email = (o.get("email") or "").lower().strip()
            invite_sent = False
            if vendor_email:
                existing_user = await db.users.find_one({"email": vendor_email})
                if not existing_user:
                    # No password sharing — vendor signs in via Email OTP. Set a random
                    # password hash so the field is populated; vendor will never use it.
                    import secrets as _secrets
                    random_pwd = _secrets.token_urlsafe(24)
                    try:
                        from passlib.hash import bcrypt as _bcrypt
                        pwd_hash = _bcrypt.hash(random_pwd)
                    except Exception:
                        pwd_hash = ""
                    await db.users.insert_one({
                        "email": vendor_email,
                        "password_hash": pwd_hash,
                        "name": o.get("contact_person") or o.get("vendor_name") or "Vendor",
                        "role": "vendor",
                        "vendor_id": vendor_id,
                        "created_at": datetime.now(timezone.utc),
                        "failed_attempts": 0,
                        "created_via": "vendor_onboarding_approval",
                    })
                else:
                    # Existing user — ensure they're linked to this vendor
                    await db.users.update_one(
                        {"_id": existing_user["_id"]},
                        {"$set": {"vendor_id": vendor_id, "role": "vendor"}},
                    )
                # Fire invitation email + branded vendor decision email (best-effort)
                try:
                    import email_service as _email_service
                    inv_name = o.get("contact_person") or o.get("vendor_name") or "Partner"
                    inv_html, inv_text = _email_service.render_invitation_email(
                        name=inv_name, email=vendor_email, role="vendor",
                    )
                    if _email_service.send_email(vendor_email, "Welcome to Cravitoo Partner — your account is ready", inv_html, inv_text):
                        invite_sent = True
                except Exception as e:
                    logger.warning(f"Vendor invitation email failed for {vendor_email}: {e}")
                # Branded "Vendor Approved" decision email (PDF Module 13)
                try:
                    import email_service as _email_service2
                    v_name = o.get("contact_person") or o.get("vendor_name") or "Partner"
                    d_html, d_text = _email_service2.render_vendor_decision_email(
                        name=v_name,
                        vendor_name=o.get("vendor_name") or "Your business",
                        decision="approve",
                        remarks=data.remarks or "",
                    )
                    _email_service2.send_email(
                        vendor_email,
                        "Your Cravitoo Partner application is approved 🎉",
                        d_html, d_text,
                    )
                except Exception as e:
                    logger.warning(f"Vendor approved-decision email failed for {vendor_email}: {e}")
            set_doc["status"] = "active"
            set_doc["vendor_id"] = vendor_id
            set_doc["vendor_user_invited"] = invite_sent
        else:
            set_doc["status"] = "rejected"
            # Branded "Vendor Rejected" email (PDF Module 13)
            try:
                vendor_email_rej = (o.get("email") or "").lower().strip()
                if vendor_email_rej:
                    import email_service as _email_service3
                    v_name_rej = o.get("contact_person") or o.get("vendor_name") or "Partner"
                    rd_html, rd_text = _email_service3.render_vendor_decision_email(
                        name=v_name_rej,
                        vendor_name=o.get("vendor_name") or "Your business",
                        decision="reject",
                        remarks=data.remarks or "",
                    )
                    _email_service3.send_email(
                        vendor_email_rej,
                        "Update on your Cravitoo Partner application",
                        rd_html, rd_text,
                    )
            except Exception as e:
                logger.warning(f"Vendor rejected-decision email failed: {e}")
        await db.vendor_onboarding.update_one({"_id": o["_id"]}, {"$set": set_doc})
        await audit_log(user, "vendor_onboarding", onb_id, f"master_{data.decision}", {"remarks": data.remarks, "vendor_id": set_doc.get("vendor_id")})
        return {"message": f"Master decision: {data.decision}", "status": set_doc["status"], "vendor_id": set_doc.get("vendor_id"), "vendor_user_invited": set_doc.get("vendor_user_invited", False)}

    @r.get("/onboarding/vendors/{onb_id}/audit-trail")
    async def onboarding_audit_trail(onb_id: str, user: dict = Depends(get_current_user)):
        o = await db.vendor_onboarding.find_one({"_id": safe_objectid(onb_id, "Onboarding")})
        if not o:
            raise HTTPException(status_code=404, detail="Onboarding not found")
        if user["role"] not in ("master_admin", "city_admin", "site_admin"):
            raise HTTPException(status_code=403, detail="Access denied")
        if user["role"] == "site_admin" and o.get("site_id") != user.get("site_id"):
            raise HTTPException(status_code=403, detail="Not your site")
        cursor = db.audit_log.find({"entity_type": "vendor_onboarding", "entity_id": onb_id}).sort("created_at", 1)
        log = []
        async for entry in cursor:
            log.append({
                "user_email": entry.get("user_email"),
                "user_role": entry.get("user_role"),
                "action": entry.get("action"),
                "details": entry.get("details", {}),
                "created_at": entry.get("created_at").isoformat() if entry.get("created_at") else None,
            })
        return {"audit_trail": log}

    @r.get("/onboarding/dashboard")
    async def onboarding_dashboard(user: dict = Depends(get_current_user)):
        """Dashboard stats based on role."""
        filt = {}
        if user["role"] == "master_admin":
            pass
        elif user["role"] == "city_admin":
            filt["city_id"] = user.get("city_id")
        elif user["role"] == "site_admin":
            filt["site_id"] = user.get("site_id")
        else:
            raise HTTPException(status_code=403, detail="Access denied")
        pipe = [
            {"$match": filt} if filt else {"$match": {}},
            {"$group": {"_id": "$status", "count": {"$sum": 1}}},
        ]
        by_status = {row["_id"]: row["count"] async for row in db.vendor_onboarding.aggregate(pipe)}
        total = sum(by_status.values())
        pending = by_status.get("documents_pending", 0) + by_status.get("under_site_review", 0) + by_status.get("under_master_review", 0)
        # Average checklist pct of in-progress onboardings
        cursor = db.vendor_onboarding.find({**filt, "status": {"$nin": ["approved", "active", "rejected"]}})
        pcts = []
        async for o in cursor:
            pcts.append(calc_checklist_pct(o.get("checklist", {})))
        avg_pct = round(sum(pcts) / len(pcts), 1) if pcts else 0.0
        return {
            "total": total,
            "by_status": by_status,
            "pending_approvals": pending,
            "approved": by_status.get("approved", 0) + by_status.get("active", 0),
            "rejected": by_status.get("rejected", 0),
            "avg_checklist_pct": avg_pct,
        }

    return r
